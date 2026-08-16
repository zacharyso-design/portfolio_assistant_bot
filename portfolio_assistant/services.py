from __future__ import annotations

import csv
import hashlib
import io
import json
import logging
import os
import re
import shutil
import sqlite3
import subprocess
import threading
import uuid
from datetime import date, datetime, time, timedelta, timezone
from email.utils import getaddresses
from pathlib import Path
from typing import Any, BinaryIO, Iterable

from openpyxl import load_workbook

from .archive import (
    SCHEMA_VERSION,
    atomic_write_json,
    atomic_write_text,
    ensure_archive_roots,
    ingestion_folder_name,
    project_folder_name,
    read_json,
    relative_to_root,
    stable_id,
    update_manifest,
    write_project_files,
)
from .config import Settings
from .db import Database, utc_now
from .extraction import (
    ExtractionFailure,
    TRANSCRIPT_SUFFIXES,
    UnsupportedSource,
    extract_source,
    inspect_native_id,
    normalize_text,
    safe_filename,
    sha256_file,
)
from .llm import LlmAdapter, LlmContractError, LlmUnavailable


PRIORITIES = ("Critical", "High", "Medium", "Low")
STATUSES = ("Green", "Yellow", "Red", "Complete")
ASSIGNEE_TYPES = ("me", "person", "team_office")
ACTION_STATES = ("open", "blocked", "complete")
RECOGNIZED_SNOW_COLUMNS = (
    "Number", "Opened", "Alternate Contact", "Assigned to", "Short description",
    "Priority", "State", "Category", "Subcategory", "Assignment group", "Updated",
    "Created", "Contact", "Configuration item", "Comments and Work notes", "Tags",
)
REQUIRED_SNOW_COLUMNS = (
    "Number", "Short description", "Assignment group", "Updated", "Comments and Work notes",
)
SNOW_ENTRY_HEADER = re.compile(
    r"(?m)^(?P<stamp>\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\s*-\s*(?P<author>.+?)\s*\((?P<kind>[^)]+)\)\s*$"
)
RULE_TYPES = {"ticket_number", "project_name", "filename_phrase", "sender_subject", "meeting_workstream"}
PROJECT_FIT_SELECTED_REVIEW_THRESHOLD = 0.45
PROJECT_FIT_ALTERNATIVE_CONFIDENCE_THRESHOLD = 0.65
LOGGER = logging.getLogger(__name__)


class NotFoundError(LookupError):
    pass


class ConflictError(ValueError):
    pass


class ValidationError(ValueError):
    pass


class UnexpectedSummaryError(RuntimeError):
    """Programming failure deliberately left without an API handler so direct regeneration returns 500."""

    pass


def _like_pattern(value: str, limit: int = 200) -> str:
    """Build a contains-pattern with LIKE wildcards escaped.

    Callers must pair this with ESCAPE '\\' so that a query containing % or _
    matches those characters literally instead of matching every row.
    """
    escaped = value[:limit].replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
    return f"%{escaped}%"


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def _decode(value: str | None, fallback: Any) -> Any:
    try:
        return json.loads(value) if value else fallback
    except json.JSONDecodeError:
        return fallback


def _safe_error(exc: Exception) -> str:
    if isinstance(exc, (UnsupportedSource, ExtractionFailure, ValidationError, LlmContractError, LlmUnavailable)):
        return str(exc)[:400]
    return "The operation failed. See application diagnostics for the safe error code."


class PortfolioService:
    def __init__(self, settings: Settings, db: Database, llm: LlmAdapter):
        self.settings = settings
        self.db = db
        self.llm = llm
        self._llm_model_catalog_lock = threading.Lock()
        self._llm_model_catalog_result: dict[str, Any] | None = None
        self._llm_model_catalog_generation = 0
        self._llm_model_catalog_inflight: threading.Event | None = None
        self._routing_publication_lock = threading.Lock()
        self.archive_paths = ensure_archive_roots(settings.app.one_drive_root)
        self.migrate_archive()
        self._recover_routing_publications()

    def _under_root(self, path: Path) -> Path:
        root = self.settings.app.one_drive_root.resolve()
        resolved = path.resolve()
        try:
            resolved.relative_to(root)
        except ValueError as exc:
            raise ValidationError("Resolved path is outside the configured OneDrive root") from exc
        return resolved

    def _project(self, connection: sqlite3.Connection, project_id: str) -> sqlite3.Row:
        row = connection.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
        if not row:
            raise NotFoundError("Project not found")
        return row

    @staticmethod
    def _archive_project_id(project_id: str) -> str:
        compact = re.sub(r"[^A-Fa-f0-9]", "", project_id)
        if len(compact) < 8:
            compact = hashlib.sha256(project_id.encode("utf-8")).hexdigest()
        return f"P-{compact[:8].upper()}"

    @staticmethod
    def _legacy_ingestion_id(source: sqlite3.Row) -> str:
        material = f"{source['id']}:{source['sha256']}".encode("utf-8")
        return f"I-{hashlib.sha256(material).hexdigest()[:8].upper()}"

    def _write_project_archive(self, project: sqlite3.Row | dict[str, Any]) -> Path:
        archive_id = str(project["archive_id"] or self._archive_project_id(str(project["id"])))
        existing = Path(str(project["folder_path"]))
        descriptor_path = existing / "_Assistant" / "project.json"
        descriptor = None
        if descriptor_path.is_file():
            try:
                descriptor = json.loads(descriptor_path.read_text(encoding="utf-8"))
            except (OSError, UnicodeDecodeError, json.JSONDecodeError):
                descriptor = None
        folder = (
            self._under_root(existing)
            if isinstance(descriptor, dict) and descriptor.get("archive_id") == archive_id
            else self._under_root(
                self.archive_paths["projects"] / project_folder_name(str(project["name"]), archive_id)
            )
        )
        folder.mkdir(parents=True, exist_ok=True)
        write_project_files(
            folder,
            project_id=str(project["id"]),
            archive_id=archive_id,
            name=str(project["name"]),
            created_at=str(project["created_at"]),
        )
        return folder

    def _initial_assistant_files(self, package: Path, ingestion_id: str, project_id: str | None, title: str) -> None:
        assistant = package / "Assistant"
        (assistant / "Extracted").mkdir(parents=True, exist_ok=True)
        atomic_write_text(
            assistant / "source-summary.md",
            f"# {title}\n\nProcessing is pending. No derived source summary is available yet.\n",
        )
        atomic_write_json(
            assistant / "index.json",
            {
                "ingestion_id": ingestion_id,
                "project_id": project_id,
                "title": title,
                "summary": "Processing pending.",
                "topics": [],
                "people": [],
                "organizations": [],
                "ticket_numbers": [],
                "suggested_project_ids": [project_id] if project_id else [],
                "original_files": [],
                "processing_version": 1,
            },
        )
        atomic_write_json(assistant / "knowledge-items.json", [])
        atomic_write_json(assistant / "citations.json", [])
        atomic_write_text(assistant / "source-lifecycle.jsonl", "")

    def _routing_staging_root(self) -> Path:
        root = self._under_root(
            self.settings.app.one_drive_root / "_PortfolioAssistant" / "staging" / "routing"
        )
        root.mkdir(parents=True, exist_ok=True)
        return root

    @staticmethod
    def _discard_routing_staging(staging: Path | None) -> None:
        if staging and staging.exists():
            try:
                shutil.rmtree(staging)
            except OSError:
                # This location is outside every archive-rescan glob. A startup
                # recovery pass will retry without exposing uncommitted memory.
                LOGGER.warning(
                    "Could not remove rolled-back routing staging %s; will retry",
                    staging,
                    exc_info=True,
                )

    @staticmethod
    def _routing_manifest_matches(manifest: Any, identity: Any) -> bool:
        if not isinstance(manifest, dict) or not isinstance(identity, dict):
            return False
        required = (
            "ingestion_id", "linked_ingestion_id", "database_project_id", "project_id",
        )
        return (
            manifest.get("source_type") == "linked-source"
            and identity.get("source_type") == "linked-source"
            and manifest.get("canonical_source") is False
            and all(
                identity.get(field) and manifest.get(field) == identity.get(field)
                for field in required
            )
        )

    def _quarantine_routing_path(self, path: Path, label: str) -> Path:
        quarantine_root = self._under_root(
            self.settings.app.one_drive_root / "_PortfolioAssistant" / "quarantine" / "routing"
        )
        quarantine_root.mkdir(parents=True, exist_ok=True)
        destination = self._under_root(quarantine_root / f"{label}-{stable_id('Q')}")
        os.replace(path, destination)
        return destination

    def _publish_routing_staging(self, staging: Path) -> bool:
        publication = read_json(staging / "publication.json", None)
        if not isinstance(publication, dict):
            return False
        review_id = int(publication.get("review_id") or 0)
        final_relative = publication.get("final_relative_path")
        mode = publication.get("mode")
        package_identity = publication.get("package_identity")
        if (
            not review_id or not isinstance(final_relative, str)
            or mode not in {"package", "segments"}
            or not isinstance(package_identity, dict)
        ):
            return False
        final = self._under_root(self.settings.app.one_drive_root / final_relative)
        with self.db.connect() as connection:
            review = connection.execute(
                "SELECT status, resolution_json FROM review_items WHERE id = ?", (review_id,)
            ).fetchone()
            resolution = _decode(review["resolution_json"], {}) if review else {}
            derived_source_id = resolution.get("derived_source_id") if isinstance(resolution, dict) else None
            derived = connection.execute(
                "SELECT ingestion_path FROM sources WHERE id = ?", (derived_source_id,)
            ).fetchone() if derived_source_id else None
        if (
            not review or review["status"] != "resolved" or not derived
            or not derived["ingestion_path"]
            or Path(derived["ingestion_path"]).resolve() != final.resolve()
        ):
            return False
        with self._routing_publication_lock:
            if mode == "package":
                staged_package = staging / "package"
                if final.exists():
                    final_manifest = read_json(final / "manifest.json", None)
                    if not self._routing_manifest_matches(final_manifest, package_identity):
                        raise ConflictError("Routing publication destination contains unrelated data")
                    if staged_package.is_dir():
                        staged_manifest = read_json(staged_package / "manifest.json", None)
                        if not self._routing_manifest_matches(staged_manifest, package_identity):
                            raise ConflictError("Committed routing package staging has invalid ownership")
                        self._merge_linked_segments(
                            staged_package / "Assistant" / "linked-segments.json",
                            final / "Assistant" / "linked-segments.json",
                        )
                    elif not self._linked_segments_include(
                        final / "Assistant" / "linked-segments.json", review_id
                    ):
                        raise ConflictError("Committed routing package staging is unavailable")
                elif staged_package.is_dir():
                    staged_manifest = read_json(staged_package / "manifest.json", None)
                    if not self._routing_manifest_matches(staged_manifest, package_identity):
                        raise ConflictError("Committed routing package staging has invalid ownership")
                    os.replace(staged_package, final)
                else:
                    raise ConflictError("Committed routing package staging is unavailable")
            else:
                staged_segments = staging / "linked-segments.json"
                if staged_segments.is_file():
                    if not final.is_dir():
                        raise ConflictError("Committed linked routing package is unavailable")
                    if not self._routing_manifest_matches(
                        read_json(final / "manifest.json", None), package_identity
                    ):
                        raise ConflictError("Routing publication destination contains unrelated data")
                    self._merge_linked_segments(
                        staged_segments, final / "Assistant" / "linked-segments.json"
                    )
                elif not self._linked_segments_include(
                    final / "Assistant" / "linked-segments.json", review_id
                ):
                    raise ConflictError("Committed linked routing segment staging is unavailable")
            shutil.rmtree(staging)
        return True

    @staticmethod
    def _linked_segments_include(path: Path, review_id: int) -> bool:
        segments = read_json(path, None)
        return isinstance(segments, list) and any(
            isinstance(item, dict) and item.get("review_id") == review_id
            for item in segments
        )

    @staticmethod
    def _merge_linked_segments(staged_path: Path, final_path: Path) -> None:
        staged = read_json(staged_path, None)
        existing = read_json(final_path, [])
        if not isinstance(staged, list) or not isinstance(existing, list):
            raise ConflictError("Linked routing segment metadata is invalid")
        merged = list(existing)
        for item in staged:
            review_id = item.get("review_id") if isinstance(item, dict) else None
            if review_id is not None:
                merged = [
                    current for current in merged
                    if not isinstance(current, dict) or current.get("review_id") != review_id
                ]
            elif item in merged:
                continue
            merged.append(item)
        atomic_write_json(final_path, merged)

    def _recover_routing_publications(self) -> None:
        try:
            staging_root = self._routing_staging_root()
            candidates = list(staging_root.iterdir())
        except OSError:
            LOGGER.warning("Could not enumerate routing publication staging", exc_info=True)
            return
        for staging in candidates:
            try:
                if not staging.is_dir():
                    continue
                if not self._publish_routing_staging(staging):
                    LOGGER.warning(
                        "Retaining ambiguous routing publication staging %s for manual recovery",
                        staging,
                    )
            except Exception:
                LOGGER.warning(
                    "Could not recover routing publication %s; will retry", staging,
                    exc_info=True,
                )

    @staticmethod
    def _matching_snow_import(imports: Path, expected_sha256: str) -> Path | None:
        for candidate in imports.iterdir():
            if candidate.name.startswith("."):
                continue
            try:
                if candidate.is_file() and sha256_file(candidate) == expected_sha256:
                    return candidate
            except OSError:
                LOGGER.warning(
                    "Skipping unavailable ServiceNow import candidate %s", candidate,
                    exc_info=True,
                )
        return None

    def _cleanup_repaired_snow_quarantines(self, imports: Path) -> None:
        """Retry non-essential package deletion after a prior database commit."""
        try:
            candidates = list(imports.iterdir())
        except OSError:
            LOGGER.warning("Could not enumerate ServiceNow repair quarantines", exc_info=True)
            return
        for quarantine in candidates:
            match = re.fullmatch(r"\.legacy-snow-package-(\d+)", quarantine.name)
            if not match:
                continue
            try:
                if not quarantine.is_dir():
                    continue
                with self.db.connect() as connection:
                    source = connection.execute(
                        """SELECT capture_method, original_path, sha256 FROM sources
                           WHERE id = ?""",
                        (int(match.group(1)),),
                    ).fetchone()
                if not source or source["capture_method"] != "snow_import":
                    continue
                authoritative = self._under_root(Path(source["original_path"]))
                if (
                    not authoritative.is_file()
                    or sha256_file(authoritative) != source["sha256"]
                ):
                    continue
                shutil.rmtree(quarantine)
            except Exception:
                # Quarantine is under shared imports, not a project. Leaving it for
                # the next start is safer than making a transient sync lock fatal.
                LOGGER.warning(
                    "Could not remove repaired ServiceNow quarantine %s; will retry",
                    quarantine,
                    exc_info=True,
                )

    def _repair_migrated_snow_source(
        self, source: sqlite3.Row, imports: Path,
    ) -> None:
        original_package = self._under_root(Path(source["ingestion_path"]))
        quarantine = self._under_root(imports / f".legacy-snow-package-{source['id']}")
        package = original_package if original_package.is_dir() else quarantine
        if not package.is_dir():
            return
        manifest = read_json(package / "manifest.json", None)
        if not isinstance(manifest, dict) or any((
            manifest.get("source_type") != "snow_comments",
            manifest.get("capture_method") != "legacy_migration",
            str(manifest.get("ingestion_id")) != str(source["ingestion_id"]),
            str(manifest.get("database_project_id")) != str(source["project_id"]),
        )):
            return
        if package == original_package:
            project_folder = self._under_root(Path(source["project_folder"]))
            try:
                relative_package = package.relative_to(project_folder)
            except ValueError:
                return
            if not relative_package.parts:
                return
        originals = manifest.get("original_files")
        primary = originals[0] if isinstance(originals, list) and originals else None
        if not isinstance(primary, dict) or not primary.get("relative_path"):
            return
        leaked_original = self._under_root(package / str(primary["relative_path"]))
        try:
            leaked_original.relative_to(package)
        except ValueError:
            return
        expected_sha256 = str(source["sha256"])
        if not leaked_original.is_file() or sha256_file(leaked_original) != expected_sha256:
            return
        authoritative = self._matching_snow_import(imports, expected_sha256)
        if authoritative is None:
            suffix = Path(safe_filename(source["original_filename"], "snow-export")).suffix[:16]
            authoritative = self._under_root(
                imports / f"recovered-{source['id']}-{expected_sha256[:16]}{suffix}"
            )
            if authoritative.exists():
                if not authoritative.is_file() or sha256_file(authoritative) != expected_sha256:
                    raise ConflictError("ServiceNow recovery destination already contains different data")
            else:
                temporary = self._under_root(imports / f".{uuid.uuid4().hex}.tmp")
                try:
                    shutil.copy2(leaked_original, temporary)
                    if authoritative.exists():
                        if sha256_file(authoritative) != expected_sha256:
                            raise ConflictError(
                                "ServiceNow recovery destination already contains different data"
                            )
                    else:
                        os.replace(temporary, authoritative)
                finally:
                    try:
                        temporary.unlink(missing_ok=True)
                    except OSError:
                        LOGGER.warning(
                            "Could not remove temporary ServiceNow recovery file %s",
                            temporary,
                            exc_info=True,
                        )
        moved = False
        if package == original_package:
            if quarantine.exists():
                raise ConflictError("A prior ServiceNow repair quarantine still needs recovery")
            os.replace(original_package, quarantine)
            moved = True
        try:
            with self.db.transaction() as connection:
                connection.execute(
                    """UPDATE sources SET original_path = ?, ingestion_id = NULL,
                       ingestion_path = NULL, capture_method = 'snow_import'
                       WHERE id = ?""",
                    (str(authoritative), source["id"]),
                )
                connection.execute(
                    "DELETE FROM original_files WHERE source_id = ?", (source["id"],)
                )
                connection.execute(
                    """UPDATE citation_records
                       SET original_relative_path = ?, display_name = ?
                       WHERE source_id = ?""",
                    (authoritative.name, source["original_filename"], source["id"]),
                )
        except Exception:
            if moved and quarantine.exists() and not original_package.exists():
                os.replace(quarantine, original_package)
            raise
        try:
            shutil.rmtree(quarantine)
        except OSError:
            LOGGER.warning(
                "Could not remove repaired ServiceNow quarantine %s; will retry",
                quarantine,
                exc_info=True,
            )

    def _repair_migrated_snow_sources(self) -> None:
        """Remove legacy per-project copies of shared ServiceNow exports."""
        with self.db.connect() as connection:
            sources = connection.execute(
                """SELECT s.*, p.folder_path AS project_folder
                   FROM sources s JOIN projects p ON p.id = s.project_id
                   WHERE s.source_type = 'snow_comments'
                     AND s.ingestion_path IS NOT NULL
                     AND s.capture_method = 'legacy_migration'
                   ORDER BY s.id"""
            ).fetchall()
        imports = self._under_root(
            self.settings.app.one_drive_root / "_PortfolioAssistant" / "imports" / "snow"
        )
        imports.mkdir(parents=True, exist_ok=True)
        self._cleanup_repaired_snow_quarantines(imports)
        for source in sources:
            try:
                self._repair_migrated_snow_source(source, imports)
            except Exception:
                # The source remains in its legacy state and will be retried on the
                # next start. A repair-only OneDrive failure must not block the app.
                LOGGER.warning(
                    "Could not repair legacy ServiceNow source %s; will retry",
                    source["id"],
                    exc_info=True,
                )

    def migrate_archive(self) -> dict[str, int]:
        """Idempotently materialize legacy database records in the durable OneDrive archive."""
        counts = {"projects": 0, "sources": 0, "missing_originals": 0}
        with self.db.connect() as connection:
            projects = connection.execute("SELECT * FROM projects ORDER BY created_at, id").fetchall()
        for project in projects:
            archive_id = str(project["archive_id"] or self._archive_project_id(str(project["id"])))
            folder = self._write_project_archive({**dict(project), "archive_id": archive_id})
            if project["archive_id"] != archive_id or Path(project["folder_path"]) != folder:
                with self.db.transaction() as connection:
                    connection.execute(
                        "UPDATE projects SET archive_id = ?, folder_path = ? WHERE id = ?",
                        (archive_id, str(folder), project["id"]),
                    )
                counts["projects"] += 1

        self._repair_migrated_snow_sources()

        with self.db.connect() as connection:
            # snow_comments rows point original_path at the shared multi-ticket
            # export and legitimately carry no ingestion_path, so migrating them
            # here would copy every ticket's export into each ticket's project
            # folder — leaking other projects' work notes across the portfolio.
            roots = connection.execute(
                "SELECT * FROM sources WHERE parent_source_id IS NULL AND ingestion_path IS NULL"
                " AND source_type <> 'snow_comments' ORDER BY id"
            ).fetchall()
        for source in roots:
            ingestion_id = str(source["ingestion_id"] or self._legacy_ingestion_id(source))
            with self.db.connect() as connection:
                project = self._project(connection, source["project_id"]) if source["project_id"] else None
            destination = Path(project["folder_path"]) if project else self.archive_paths["shared_intake"]
            created = datetime.fromisoformat(str(source["created_at"]).replace("Z", "+00:00"))
            title = str(source["meeting_name"] or Path(source["original_filename"]).stem)
            folder_name = ingestion_folder_name(created, source["source_type"], title, ingestion_id)
            final = self._under_root(destination / folder_name)
            incomplete = self._under_root(destination / f"_INCOMPLETE_{ingestion_id}")
            if not final.exists():
                incomplete.mkdir(parents=True, exist_ok=True)
                original_dir = incomplete / "Original"
                original_dir.mkdir(exist_ok=True)
                self._initial_assistant_files(incomplete, ingestion_id, source["project_id"], title)
                old = Path(source["original_path"])
                originals: list[dict[str, Any]] = []
                errors: list[str] = []
                if old.is_file():
                    stored = safe_filename(source["original_filename"], "source")
                    new_original = original_dir / stored
                    shutil.copy2(old, new_original)
                    originals.append({
                        "relative_path": f"Original/{stored}",
                        "original_name": source["original_filename"],
                        "stored_name": stored,
                        "size_bytes": new_original.stat().st_size,
                        "sha256": sha256_file(new_original),
                    })
                else:
                    errors.append("Legacy original is unavailable; no replacement was fabricated.")
                    counts["missing_originals"] += 1
                atomic_write_json(incomplete / "manifest.json", {
                    "schema_version": SCHEMA_VERSION,
                    "ingestion_id": ingestion_id,
                    "project_id": project["archive_id"] if project else None,
                    "database_project_id": source["project_id"],
                    "source_type": source["source_type"],
                    "title": title,
                    "created_at": source["created_at"],
                    "source_date": source["meeting_date"],
                    "capture_method": "legacy_migration",
                    "canonical_source": True,
                    "linked_ingestion_id": None,
                    "processing_status": source["processing_state"],
                    "original_files": originals,
                    "assistant_files": [],
                    "extractor_version": "1.0",
                    "errors": errors,
                })
                os.replace(incomplete, final)
            counts["missing_originals"] += self._migrate_legacy_attachments(int(source["id"]), final)
            manifest = read_json(final / "manifest.json", {})
            originals = manifest.get("original_files", [])
            primary = final / str(originals[0]["relative_path"]) if originals else Path(source["original_path"])
            with self.db.transaction() as connection:
                connection.execute(
                    """UPDATE sources SET ingestion_id = ?, ingestion_path = ?, source_title = ?,
                       capture_method = 'legacy_migration', original_path = ? WHERE id = ?""",
                    (ingestion_id, str(final), title, str(primary), source["id"]),
                )
                for item in originals:
                    connection.execute(
                        """INSERT OR IGNORE INTO original_files(
                           source_id, relative_path, original_name, stored_name, size_bytes, sha256,
                           is_attachment, created_at
                           ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                        (source["id"], item["relative_path"], item["original_name"], item["stored_name"],
                         item["size_bytes"], item["sha256"], int(bool(item.get("is_attachment"))),
                         source["created_at"]),
                    )
            self._refresh_source_archive(int(source["id"]))
            counts["sources"] += 1
        with self.db.connect() as connection:
            legacy_updates = connection.execute(
                """SELECT u.* FROM project_updates u
                   WHERE u.source_id IS NOT NULL AND u.update_type IN ('knowledge', 'routed_review')
                   AND NOT EXISTS (
                     SELECT 1 FROM knowledge_items k
                     WHERE k.source_id = u.source_id AND k.text = u.text
                   ) ORDER BY u.created_at, u.id"""
            ).fetchall()
        touched_projects: set[str] = set()
        touched_sources: set[int] = set()
        for update in legacy_updates:
            with self.db.transaction() as connection:
                source = connection.execute("SELECT * FROM sources WHERE id = ?", (update["source_id"],)).fetchone()
                if not source:
                    continue
                citation_ids = []
                for citation in _decode(update["citations_json"], []):
                    try:
                        validated = {"source_id": int(citation["source_id"]), "chunk_id": int(citation["chunk_id"])}
                        durable = self._ensure_citation_record(connection, source, validated, update["created_at"])
                        citation_ids.append(durable["citation_id"])
                    except (KeyError, TypeError, ValueError, LlmContractError):
                        continue
                if not citation_ids:
                    continue
                connection.execute(
                    """INSERT INTO knowledge_items(
                       id, project_id, source_id, text, category, source_date, citation_ids_json,
                       review_status, created_at
                       ) VALUES (?, ?, ?, ?, ?, ?, ?, 'unreviewed', ?)""",
                    (stable_id("K"), update["project_id"], update["source_id"], update["text"],
                     self._knowledge_category(update["text"]), source["source_date"] or source["created_at"],
                     _json(citation_ids), update["created_at"]),
                )
            touched_projects.add(str(update["project_id"]))
            touched_sources.add(int(update["source_id"]))
        with self.db.transaction() as connection:
            legacy_projects = connection.execute(
                """SELECT * FROM projects WHERE current_summary <> '' AND NOT EXISTS (
                     SELECT 1 FROM summary_versions v WHERE v.project_id = projects.id
                   )"""
            ).fetchall()
            for project in legacy_projects:
                content = {"legacy": True, "sections": [{
                    "section": "Legacy Summary",
                    "text": project["current_summary"],
                    "knowledge_item_ids": [],
                }]}
                connection.execute(
                    """INSERT INTO summary_versions(
                       project_id, revision, content_json, markdown, review_status,
                       generation_state, model_id, created_at
                       ) VALUES (?, 0, ?, ?, 'unreviewed', 'stale', NULL, ?)""",
                    (project["id"], _json(content), f"## Legacy Summary\n\n{project['current_summary']}",
                     project["updated_at"]),
                )
                connection.execute(
                    "UPDATE projects SET summary_generation_state = 'stale' WHERE id = ?", (project["id"],)
                )
                legacy_path = (
                    Path(project["folder_path"]) / "_Assistant" / "living-summary" / "versions" /
                    "000000__legacy.json"
                )
                atomic_write_json(legacy_path, {
                    "schema_version": SCHEMA_VERSION, "project_id": project["archive_id"],
                    "revision": 0, "review_status": "unreviewed", "generation_state": "stale",
                    "created_at": project["updated_at"], **content,
                })
        for source_id in touched_sources:
            self._refresh_source_archive(source_id)
        for project_id in touched_projects:
            self._write_knowledge_history(project_id)
        return counts

    def _migrate_legacy_attachments(self, root_source_id: int, package: Path) -> int:
        """Copy legacy attachment source rows into their root source's durable package."""
        manifest_path = package / "manifest.json"
        manifest = read_json(manifest_path, {})
        if not isinstance(manifest, dict):
            manifest = {}
        originals = manifest.get("original_files")
        if not isinstance(originals, list):
            originals = []
        errors = manifest.get("errors")
        if not isinstance(errors, list):
            errors = []
        with self.db.connect() as connection:
            children = connection.execute(
                """WITH RECURSIVE descendants(id) AS (
                     SELECT id FROM sources WHERE parent_source_id = ?
                     UNION ALL
                     SELECT s.id FROM sources s JOIN descendants d ON s.parent_source_id = d.id
                   )
                   SELECT s.* FROM sources s WHERE s.id IN (SELECT id FROM descendants)
                   ORDER BY s.id""",
                (root_source_id,),
            ).fetchall()
        used = {str(item.get("relative_path", "")).casefold() for item in originals}
        missing = 0
        changed = False
        for child in children:
            existing_item = next(
                (item for item in originals if item.get("legacy_source_id") == int(child["id"])), None
            )
            old_path = Path(child["original_path"])
            if existing_item:
                migrated_path = self._under_root(package / str(existing_item["relative_path"]))
                if migrated_path.is_file():
                    with self.db.transaction() as connection:
                        connection.execute(
                            """UPDATE sources SET original_path = ?, ingestion_path = ?,
                               capture_method = 'legacy_attachment_migration' WHERE id = ?""",
                            (str(migrated_path), str(package), child["id"]),
                        )
                    continue
            if not old_path.is_file():
                message = f"Legacy attachment is unavailable: {child['original_filename']}"
                if message not in errors:
                    errors.append(message)
                    changed = True
                missing += 1
                continue
            name = safe_filename(child["original_filename"], f"attachment-{child['id']}")
            candidate = Path("Original") / "Attachments" / name
            suffix = 2
            while candidate.as_posix().casefold() in used:
                candidate = candidate.with_name(f"{Path(name).stem}-{suffix}{Path(name).suffix}")
                suffix += 1
            migrated_path = self._under_root(package / candidate)
            temporary = migrated_path.parent / f".{uuid.uuid4().hex}.tmp"
            try:
                migrated_path.parent.mkdir(parents=True, exist_ok=True)
                digest = sha256_file(old_path)
                shutil.copy2(old_path, temporary)
                if sha256_file(temporary) != digest:
                    raise ConflictError("Legacy attachment copy failed its integrity check")
                os.replace(temporary, migrated_path)
                item = {
                    "relative_path": candidate.as_posix(),
                    "original_name": child["original_filename"],
                    "stored_name": migrated_path.name,
                    "size_bytes": migrated_path.stat().st_size,
                    "sha256": digest,
                    "is_attachment": True,
                    "legacy_source_id": int(child["id"]),
                }
            except ConflictError:
                message = (
                    "Legacy attachment failed its integrity check after copy: "
                    f"{child['original_filename']}"
                )
                if message not in errors:
                    errors.append(message)
                    changed = True
                missing += 1
                continue
            except OSError:
                message = (
                    "Legacy attachment could not be archived due to a file-system error: "
                    f"{child['original_filename']}"
                )
                if message not in errors:
                    errors.append(message)
                    changed = True
                missing += 1
                continue
            finally:
                try:
                    temporary.unlink(missing_ok=True)
                except OSError:
                    message = f"Legacy attachment temporary cleanup failed: {child['original_filename']}"
                    if message not in errors:
                        errors.append(message)
                        changed = True
            used.add(candidate.as_posix().casefold())
            originals.append(item)
            with self.db.transaction() as connection:
                connection.execute(
                    """INSERT OR IGNORE INTO original_files(
                       source_id, relative_path, original_name, stored_name, size_bytes, sha256,
                       is_attachment, created_at
                       ) VALUES (?, ?, ?, ?, ?, ?, 1, ?)""",
                    (root_source_id, item["relative_path"], item["original_name"], item["stored_name"],
                     item["size_bytes"], item["sha256"], child["created_at"]),
                )
                connection.execute(
                    """UPDATE sources SET original_path = ?, ingestion_path = ?,
                       capture_method = 'legacy_attachment_migration' WHERE id = ?""",
                    (str(migrated_path), str(package), child["id"]),
                )
            changed = True
        if changed:
            manifest["original_files"] = originals
            manifest["errors"] = errors
            atomic_write_json(manifest_path, manifest)
        return missing

    @staticmethod
    def _knowledge_category(text: str) -> str:
        lowered = text.casefold()
        if any(word in lowered for word in ("approved", "decided", "decision")):
            return "decision"
        if any(word in lowered for word in ("risk", "issue", "blocked", "delay")):
            return "risk"
        if any(word in lowered for word in ("due", "milestone", "schedule", "target date")):
            return "milestone"
        if any(word in lowered for word in ("action", "follow up", "next step")):
            return "action"
        return "development"

    def _record_source_lifecycle(
        self,
        connection: sqlite3.Connection,
        source: sqlite3.Row,
        event_type: str,
        reason: str,
        *,
        from_project_id: str | None = None,
        to_project_id: str | None = None,
        details: dict[str, Any] | None = None,
        event_id: str | None = None,
        created_at: str | None = None,
    ) -> str:
        event_id = event_id or stable_id("E")
        connection.execute(
            """INSERT INTO source_lifecycle_events(
               id, source_id, ingestion_id, project_id, event_type, from_project_id,
               to_project_id, reason, details_json, created_at
               ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                event_id, source["id"], source["ingestion_id"], source["project_id"],
                event_type, from_project_id, to_project_id, normalize_text(reason)[:500],
                _json(details or {}), created_at or utc_now(),
            ),
        )
        return event_id

    def _refresh_source_archive(self, source_id: int, *, processing_status: str | None = None) -> None:
        with self.db.connect() as connection:
            source = connection.execute("SELECT * FROM sources WHERE id = ?", (source_id,)).fetchone()
            if not source or not source["ingestion_path"]:
                return
            related_ids = [source_id]
            if source["source_type"] == "routed_segment":
                related_ids = [int(row["id"]) for row in connection.execute(
                    "SELECT id FROM sources WHERE project_id = ? AND ingestion_path = ?",
                    (source["project_id"], source["ingestion_path"]),
                ).fetchall()]
            placeholders = ",".join("?" for _ in related_ids)
            originals = [dict(row) for row in connection.execute(
                f"""SELECT * FROM original_files WHERE source_id IN ({placeholders})
                    ORDER BY is_attachment, id""", related_ids
            ).fetchall()]
            chunks = [dict(row) for row in connection.execute(
                f"""WITH RECURSIVE source_tree(id) AS (
                       SELECT id FROM sources WHERE id IN ({placeholders})
                       UNION ALL
                       SELECT s.id FROM sources s JOIN source_tree t ON s.parent_source_id = t.id
                     )
                     SELECT c.* FROM source_chunks c WHERE c.source_id IN (SELECT id FROM source_tree)
                     ORDER BY c.source_id, c.sequence""",
                related_ids,
            ).fetchall()]
            knowledge = [dict(row) for row in connection.execute(
                f"SELECT * FROM knowledge_items WHERE source_id IN ({placeholders}) ORDER BY created_at, id",
                related_ids,
            ).fetchall()]
            citations = [dict(row) for row in connection.execute(
                f"SELECT * FROM citation_records WHERE source_id IN ({placeholders}) ORDER BY id", related_ids
            ).fetchall()]
            lifecycle = [dict(row) for row in connection.execute(
                "SELECT * FROM source_lifecycle_events WHERE source_id = ? ORDER BY created_at, id",
                (source_id,),
            ).fetchall()]
            project = (
                connection.execute("SELECT * FROM projects WHERE id = ?", (source["project_id"],)).fetchone()
                if source["project_id"] else None
            )
        package = self._under_root(Path(source["ingestion_path"]))
        metadata = _decode(source["metadata_json"], {})
        source_summary = source["source_summary"].strip()
        if not source_summary:
            preview = normalize_text(" ".join(chunk["text"] for chunk in chunks))[:900]
            source_summary = preview or "No searchable text was extracted; the original remains preserved."
        tickets = sorted(set(re.findall(r"\b(?:REQ|INC|CHG|RITM|TASK)[-_]?\d+\b", source_summary, re.IGNORECASE)))
        topics = sorted({self._knowledge_category(item["text"]) for item in knowledge})
        addresses = getaddresses([
            str(metadata.get("sender") or ""), str(metadata.get("recipients") or ""),
            str(metadata.get("cc") or ""),
        ])
        people = sorted({normalize_text(name) or address for name, address in addresses if name or address})
        organizations = {address.rsplit("@", 1)[1].casefold() for _, address in addresses if "@" in address}
        organizations.update(re.findall(r"\b[A-Z][A-Z0-9&-]{1,9}\b", source_summary))
        original_payload = [{
            "relative_path": item["relative_path"],
            "original_name": item["original_name"],
            "stored_name": item["stored_name"],
            "size_bytes": item["size_bytes"],
            "sha256": item["sha256"],
            "is_attachment": bool(item["is_attachment"]),
        } for item in originals]
        atomic_write_text(
            package / "Assistant" / "source-summary.md",
            f"# {source['source_title'] or source['original_filename']}\n\n"
            f"- **Source type:** {source['source_type']}\n"
            f"- **Source date:** {source['source_date'] or source['created_at']}\n"
            f"- **Project:** {project['name'] if project else 'Shared intake — awaiting routing'}\n\n"
            f"## Overview\n\n{source_summary}\n\n"
            f"## Categories\n\n{', '.join(topics) if topics else 'Not yet categorized'}\n",
        )
        atomic_write_json(package / "Assistant" / "index.json", {
            "ingestion_id": source["ingestion_id"],
            "project_id": project["archive_id"] if project else None,
            "title": source["source_title"] or source["original_filename"],
            "source_type": source["source_type"],
            "source_date": source["source_date"] or source["created_at"],
            "summary": source_summary,
            "topics": topics,
            "people": people,
            "organizations": sorted(organizations),
            "ticket_numbers": tickets,
            "suggested_project_ids": [project["archive_id"]] if project else [],
            "original_files": original_payload,
            "email_metadata": {key: metadata.get(key) for key in (
                "subject", "sender", "recipients", "cc", "timestamp", "message_id"
            ) if metadata.get(key)},
            "processing_version": source["processing_version"],
            "memory_state": source["memory_state"],
            "project_fit_confirmed": bool(source["project_fit_confirmed"]),
            "archived_previous_memory_state": source["archived_previous_memory_state"],
        })
        atomic_write_json(package / "Assistant" / "knowledge-items.json", [{
            "knowledge_item_id": item["id"],
            "project_id": project["archive_id"] if project else item["project_id"],
            "text": item["text"],
            "category": item["category"],
            "source_date": item["source_date"],
            "citation_ids": _decode(item["citation_ids_json"], []),
            "review_status": item["review_status"],
            "supersedes_knowledge_item_id": item["supersedes_knowledge_item_id"],
        } for item in knowledge])
        atomic_write_json(package / "Assistant" / "citations.json", [{
            "citation_id": item["id"],
            "original_relative_path": item["original_relative_path"],
            "display_name": item["display_name"],
            "source_type": item["source_type"],
            "locator": item["locator"],
            "excerpt": item["excerpt"],
            "source_date": item["source_date"],
        } for item in citations])
        lifecycle_lines = []
        for lifecycle_item in lifecycle:
            lifecycle_item["details"] = _decode(lifecycle_item.pop("details_json"), {})
            lifecycle_lines.append(_json(lifecycle_item))
        atomic_write_text(
            package / "Assistant" / "source-lifecycle.jsonl",
            ("\n".join(lifecycle_lines) + "\n") if lifecycle_lines else "",
        )
        existing_manifest = read_json(package / "manifest.json", {})
        archive_errors = list(existing_manifest.get("errors", [])) if isinstance(existing_manifest, dict) else []
        if source["error_message"] and source["error_message"] not in archive_errors:
            archive_errors.append(source["error_message"])
        update_manifest(
            package,
            processing_status=processing_status or source["processing_state"],
            original_files=original_payload,
            assistant_files=[
                "Assistant/source-summary.md", "Assistant/index.json",
                "Assistant/knowledge-items.json", "Assistant/citations.json",
                "Assistant/source-lifecycle.jsonl", "Assistant/Extracted",
            ],
            database_project_id=source["project_id"],
            project_id=project["archive_id"] if project else None,
            memory_state=source["memory_state"],
            project_fit_confirmed=bool(source["project_fit_confirmed"]),
            archived_previous_memory_state=source["archived_previous_memory_state"],
            errors=archive_errors,
        )

    def create_group(self, name: str, sort_order: int = 0) -> dict[str, Any]:
        clean = normalize_text(name)
        if not clean or len(clean) > 120:
            raise ValidationError("Portfolio group name must be 1-120 characters")
        with self.db.transaction() as connection:
            try:
                cursor = connection.execute(
                    "INSERT INTO portfolio_groups(name, sort_order, is_system, created_at) VALUES (?, ?, 0, ?)",
                    (clean, int(sort_order), utc_now()),
                )
            except sqlite3.IntegrityError as exc:
                raise ConflictError("A portfolio group with that name already exists") from exc
            row = connection.execute("SELECT * FROM portfolio_groups WHERE id = ?", (cursor.lastrowid,)).fetchone()
        return dict(row)

    def list_groups(self) -> list[dict[str, Any]]:
        with self.db.connect() as connection:
            rows = connection.execute(
                """
                SELECT g.*, count(p.id) AS project_count
                FROM portfolio_groups g LEFT JOIN projects p ON p.portfolio_group_id = g.id
                GROUP BY g.id ORDER BY g.sort_order, g.name COLLATE NOCASE
                """
            ).fetchall()
        return [dict(row) for row in rows]

    def update_group(self, group_id: int, *, name: str | None = None, sort_order: int | None = None) -> dict[str, Any]:
        with self.db.transaction() as connection:
            row = connection.execute("SELECT * FROM portfolio_groups WHERE id = ?", (group_id,)).fetchone()
            if not row:
                raise NotFoundError("Portfolio group not found")
            if row["is_system"] and name is not None and normalize_text(name).casefold() != "unassigned":
                raise ValidationError("The system Unassigned group cannot be renamed")
            new_name = normalize_text(name) if name is not None else row["name"]
            new_order = int(sort_order) if sort_order is not None else row["sort_order"]
            try:
                connection.execute(
                    "UPDATE portfolio_groups SET name = ?, sort_order = ? WHERE id = ?",
                    (new_name, new_order, group_id),
                )
            except sqlite3.IntegrityError as exc:
                raise ConflictError("A portfolio group with that name already exists") from exc
            updated = connection.execute("SELECT * FROM portfolio_groups WHERE id = ?", (group_id,)).fetchone()
        return dict(updated)

    def create_project(
        self,
        name: str,
        *,
        portfolio_group_id: int | None = None,
        snow_number: str | None = None,
        assignment_group: str | None = None,
        snow_metadata: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        clean = normalize_text(name)
        if not clean or len(clean) > 240:
            raise ValidationError("Project name must be 1-240 characters")
        project_id = str(uuid.uuid4())
        archive_id = self._archive_project_id(project_id)
        folder = self._under_root(
            self.archive_paths["projects"] / project_folder_name(clean, archive_id)
        )
        folder.mkdir(parents=True, exist_ok=False)
        now = utc_now()
        write_project_files(
            folder, project_id=project_id, archive_id=archive_id, name=clean, created_at=now
        )
        try:
            with self.db.transaction() as connection:
                if portfolio_group_id is None:
                    group = connection.execute(
                        "SELECT id FROM portfolio_groups WHERE name = 'Unassigned' COLLATE NOCASE"
                    ).fetchone()
                    portfolio_group_id = int(group["id"])
                elif not connection.execute("SELECT 1 FROM portfolio_groups WHERE id = ?", (portfolio_group_id,)).fetchone():
                    raise ValidationError("Portfolio group not found")
                connection.execute(
                    """
                    INSERT INTO projects(
                      id, archive_id, name, portfolio_group_id, snow_number, assignment_group,
                      snow_metadata_json, folder_path, created_at, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (project_id, archive_id, clean, portfolio_group_id, snow_number, assignment_group,
                     _json(snow_metadata or {}), str(folder), now, now),
                )
        except Exception:
            try:
                shutil.rmtree(folder)
            except OSError:
                pass
            raise
        return self.get_project(project_id)

    def list_projects(
        self,
        *,
        query: str = "",
        portfolio_group_id: int | None = None,
        assignment_group: str | None = None,
        status: str | None = None,
        priority: str | None = None,
        limit: int = 250,
    ) -> dict[str, Any]:
        clauses: list[str] = []
        params: list[Any] = []
        if query.strip():
            pattern = _like_pattern(query.strip())
            clauses.append(
                "(p.name LIKE ? ESCAPE '\\' OR p.snow_number LIKE ? ESCAPE '\\' OR p.owner_text LIKE ? ESCAPE '\\' OR p.assignment_group LIKE ? ESCAPE '\\' OR p.next_action LIKE ? ESCAPE '\\' OR p.latest_change LIKE ? ESCAPE '\\')"
            )
            params.extend([pattern] * 6)
        if portfolio_group_id is not None:
            clauses.append("p.portfolio_group_id = ?")
            params.append(portfolio_group_id)
        if assignment_group is not None:
            if assignment_group == "":
                clauses.append("coalesce(p.assignment_group, '') = ''")
            else:
                clauses.append("p.assignment_group = ? COLLATE NOCASE")
                params.append(assignment_group)
        if status:
            if status not in STATUSES:
                raise ValidationError("Invalid project status filter")
            clauses.append("p.status = ?")
            params.append(status)
        if priority:
            if priority not in PRIORITIES:
                raise ValidationError("Invalid project priority filter")
            clauses.append("p.priority = ?")
            params.append(priority)
        where = "WHERE " + " AND ".join(clauses) if clauses else ""
        limit = max(1, min(limit, 250))
        with self.db.connect() as connection:
            rows = connection.execute(
                f"""
                SELECT p.*, g.name AS portfolio_group_name
                FROM projects p JOIN portfolio_groups g ON g.id = p.portfolio_group_id
                {where}
                ORDER BY g.sort_order, g.name COLLATE NOCASE,
                  CASE p.priority WHEN 'Critical' THEN 0 WHEN 'High' THEN 1 WHEN 'Medium' THEN 2 ELSE 3 END,
                  p.name COLLATE NOCASE LIMIT ?
                """,
                (*params, limit),
            ).fetchall()
            total = connection.execute(f"SELECT count(*) FROM projects p {where}", params).fetchone()[0]
            assignments = [row[0] or "" for row in connection.execute(
                "SELECT DISTINCT assignment_group FROM projects ORDER BY assignment_group COLLATE NOCASE"
            ).fetchall()]
            metrics = dict(connection.execute(
                """SELECT count(*) AS total,
                   sum(CASE WHEN status <> 'Complete' THEN 1 ELSE 0 END) AS active,
                   sum(CASE WHEN status = 'Red' THEN 1 ELSE 0 END) AS red
                   FROM projects"""
            ).fetchone())
        return {
            "items": [self._decode_project(dict(row)) for row in rows],
            "total": total,
            "limit": limit,
            "assignment_groups": assignments,
            "metrics": metrics,
        }

    def _decode_project(self, project: dict[str, Any]) -> dict[str, Any]:
        project["snow_metadata"] = _decode(project.pop("snow_metadata_json", "{}"), {})
        project["name_manually_overridden"] = bool(project["name_manually_overridden"])
        return project

    def get_project(self, project_id: str) -> dict[str, Any]:
        with self.db.connect() as connection:
            row = connection.execute(
                """SELECT p.*, g.name AS portfolio_group_name
                   FROM projects p JOIN portfolio_groups g ON g.id = p.portfolio_group_id
                   WHERE p.id = ?""",
                (project_id,),
            ).fetchone()
            if not row:
                raise NotFoundError("Project not found")
            updates = connection.execute(
                """SELECT u.* FROM project_updates u LEFT JOIN sources s ON s.id = u.source_id
                   WHERE u.project_id = ? AND (u.source_id IS NULL OR s.memory_state = 'active')
                   ORDER BY u.created_at DESC, u.id DESC LIMIT 100""",
                (project_id,),
            ).fetchall()
            sources = connection.execute(
                """SELECT s.*, parent.original_filename AS parent_original_filename
                   FROM sources s LEFT JOIN sources parent ON parent.id = s.parent_source_id
                   WHERE s.project_id = ?
                   ORDER BY s.memory_state = 'removed', s.parent_source_id IS NOT NULL,
                            s.created_at DESC, s.id DESC LIMIT 200""",
                (project_id,),
            ).fetchall()
            actions = connection.execute(
                """SELECT a.* FROM action_items a LEFT JOIN sources s ON s.id = a.source_id
                   WHERE a.project_id = ?
                     AND (a.source_id IS NULL OR a.created_by = 'user' OR s.memory_state = 'active')
                   ORDER BY a.state = 'complete', a.due_date, a.id DESC""",
                (project_id,),
            ).fetchall()
            decoded_updates = [self._decode_update(item) for item in updates]
            decoded_actions = self._decode_active_actions(connection, actions)
            for update in decoded_updates:
                update["citations"] = self._enrich_citations(connection, update["citations"])
            for action in decoded_actions:
                action["citations"] = self._enrich_citations(connection, action["citations"])
        result = self._decode_project(dict(row))
        result["updates"] = decoded_updates
        result["sources"] = [self._decode_source(item) for item in sources]
        result["action_items"] = decoded_actions
        summary_citations: list[dict[str, Any]] = []
        seen: set[tuple[int, int]] = set()
        for update in reversed(decoded_updates):
            for citation in update["citations"]:
                pair = (int(citation["source_id"]), int(citation["chunk_id"]))
                if pair not in seen:
                    seen.add(pair)
                    summary_citations.append(citation)
        result["summary_citations"] = summary_citations
        result["knowledge_history"] = self.list_knowledge(project_id)
        result["living_summary"] = self.get_living_summary(project_id)
        return result

    def _enrich_citations(
        self, connection: sqlite3.Connection, citations: list[dict[str, Any]]
    ) -> list[dict[str, Any]]:
        enriched: list[dict[str, Any]] = []
        for citation in citations:
            try:
                source_id = int(citation["source_id"])
                chunk_id = int(citation["chunk_id"])
            except (KeyError, TypeError, ValueError):
                continue
            row = connection.execute(
                """SELECT c.text, c.locator, s.original_filename, s.meeting_name, s.meeting_date
                   FROM source_chunks c JOIN sources s ON s.id = c.source_id
                   WHERE c.id = ? AND c.source_id = ?""",
                (chunk_id, source_id),
            ).fetchone()
            if row:
                enriched.append({
                    "source_id": source_id,
                    "chunk_id": chunk_id,
                    "citation_id": citation.get("citation_id"),
                    "locator": row["locator"],
                    "original_filename": row["original_filename"],
                    "meeting_name": row["meeting_name"],
                    "meeting_date": row["meeting_date"],
                    "excerpt": row["text"][:600],
                })
        return enriched

    def _decode_update(self, row: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
        item = dict(row)
        item["citations"] = _decode(item.pop("citations_json", "[]"), [])
        return item

    def _decode_source(self, row: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
        item = dict(row)
        item["metadata"] = _decode(item.pop("metadata_json", "{}"), {})
        return item

    @staticmethod
    def _manifest_memory_state(manifest: dict[str, Any]) -> str:
        explicit = manifest.get("memory_state")
        if explicit in {"pending", "active", "removed"}:
            return str(explicit)
        return "active" if manifest.get("processing_status") == "complete" else "pending"

    def _decode_action(self, row: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
        item = dict(row)
        item["citations"] = _decode(item.pop("citations_json", None), [])
        return item

    def _decode_active_action(
        self, connection: sqlite3.Connection, row: sqlite3.Row | dict[str, Any]
    ) -> dict[str, Any]:
        return self._decode_active_actions(connection, [row])[0]

    def _decode_active_actions(
        self, connection: sqlite3.Connection,
        rows: Iterable[sqlite3.Row | dict[str, Any]],
    ) -> list[dict[str, Any]]:
        items = [self._decode_action(row) for row in rows]
        cited_source_ids: set[int] = set()
        for item in items:
            if not item.get("progress_text"):
                continue
            cited_source_ids.update(
                int(citation["source_id"])
                for citation in item["citations"]
                if isinstance(citation, dict) and citation.get("source_id") is not None
            )
        if not cited_source_ids:
            return items
        placeholders = ",".join("?" for _ in cited_source_ids)
        active_source_ids = {
            int(source["id"])
            for source in connection.execute(
                f"SELECT id FROM sources WHERE id IN ({placeholders}) AND memory_state = 'active'",
                tuple(cited_source_ids),
            ).fetchall()
        }
        for item in items:
            progress_source_ids = {
                int(citation["source_id"])
                for citation in item["citations"]
                if isinstance(citation, dict) and citation.get("source_id") is not None
            }
            if item.get("progress_text") and not progress_source_ids.issubset(active_source_ids):
                # The action itself may be user-owned; retract only its removed-source progress.
                item["progress_text"] = None
                item["citations"] = []
        return items

    def update_project(self, project_id: str, fields: dict[str, Any]) -> dict[str, Any]:
        allowed = {"name", "portfolio_group_id", "status", "priority", "owner_text", "next_action", "next_action_due"}
        unknown = set(fields) - allowed
        if unknown:
            raise ValidationError(f"Read-only or unknown project fields: {', '.join(sorted(unknown))}")
        changes: list[tuple[str, Any, Any]] = []
        now = utc_now()
        with self.db.transaction() as connection:
            project = self._project(connection, project_id)
            values: dict[str, Any] = {}
            for key, value in fields.items():
                if key == "name":
                    value = normalize_text(str(value))
                    if not value:
                        raise ValidationError("Project name cannot be blank")
                    values["name_manually_overridden"] = 1
                elif key == "status" and value not in STATUSES:
                    raise ValidationError("Invalid project status")
                elif key == "priority" and value not in PRIORITIES:
                    raise ValidationError("Invalid project priority")
                elif key == "portfolio_group_id":
                    if not connection.execute("SELECT 1 FROM portfolio_groups WHERE id = ?", (value,)).fetchone():
                        raise ValidationError("Portfolio group not found")
                elif key == "next_action_due" and value:
                    try:
                        date.fromisoformat(str(value))
                    except ValueError as exc:
                        raise ValidationError("next_action_due must be an ISO date") from exc
                if project[key] != value:
                    changes.append((key, project[key], value))
                values[key] = value
            if values:
                values["updated_at"] = now
                if "status" in values:
                    values["completed_at"] = now if values["status"] == "Complete" else None
                assignments = ", ".join(f"{key} = ?" for key in values)
                connection.execute(
                    f"UPDATE projects SET {assignments} WHERE id = ?",
                    (*values.values(), project_id),
                )
            for key, old, new in changes:
                if key in {"status", "priority"}:
                    text = f"User changed {key} from {old} to {new}."
                    connection.execute(
                        "INSERT INTO project_updates(project_id, source_id, update_type, text, citations_json, model_id, created_at) VALUES (?, NULL, 'manual_field', ?, '[]', NULL, ?)",
                        (project_id, text, now),
                    )
                    connection.execute("UPDATE projects SET latest_change = ? WHERE id = ?", (text, project_id))
        updated_project = self.get_project(project_id)
        if "name" in fields:
            write_project_files(
                Path(updated_project["folder_path"]),
                project_id=project_id,
                archive_id=updated_project["archive_id"],
                name=updated_project["name"],
                created_at=updated_project["created_at"],
            )
        return updated_project

    def _destination_dir(self, project: sqlite3.Row | None, multi_project: bool) -> Path:
        if multi_project:
            return self._under_root(self.archive_paths["shared_intake"])
        if project is None:
            raise ValidationError("A destination project is required")
        return self._under_root(Path(project["folder_path"]))

    @staticmethod
    def _find_existing_source(
        connection: sqlite3.Connection, *, project_id: str | None, digest: str,
        native_id: str | None, multi_project: bool,
    ) -> sqlite3.Row | None:
        if multi_project:
            return connection.execute(
                """SELECT * FROM sources WHERE project_id IS NULL AND sha256 = ?
                   AND parent_source_id IS NULL""",
                (digest,),
            ).fetchone()
        if native_id:
            return connection.execute(
                "SELECT * FROM sources WHERE project_id = ? AND native_id = ?", (project_id, native_id)
            ).fetchone()
        return connection.execute(
            """SELECT * FROM sources WHERE project_id = ? AND sha256 = ?
               AND parent_source_id IS NULL""",
            (project_id, digest),
        ).fetchone()

    def capture_source(
        self,
        stream: BinaryIO,
        filename: str,
        *,
        project_id: str | None,
        meeting_name: str | None = None,
        meeting_date: str | None = None,
        is_transcript: bool = False,
        multi_project: bool = False,
        capture_method: str = "file_upload",
    ) -> tuple[dict[str, Any], bool]:
        return self.capture_sources(
            [(stream, filename, filename)],
            project_id=project_id,
            meeting_name=meeting_name,
            meeting_date=meeting_date,
            is_transcript=is_transcript,
            multi_project=multi_project,
            capture_method=capture_method,
        )

    @staticmethod
    def _safe_relative_path(value: str, fallback: str) -> Path:
        candidate = str(value or fallback).replace("\\", "/")
        parts = [part for part in candidate.split("/") if part not in {"", "."}]
        if not parts or any(part == ".." for part in parts) or Path(candidate).is_absolute():
            parts = [fallback]
        return Path(*(safe_filename(part, "item") for part in parts))

    def capture_sources(
        self,
        files: list[tuple[BinaryIO, str, str]],
        *,
        project_id: str | None,
        meeting_name: str | None = None,
        meeting_date: str | None = None,
        is_transcript: bool = False,
        multi_project: bool = False,
        capture_method: str = "file_upload",
    ) -> tuple[dict[str, Any], bool]:
        if not files:
            raise ValidationError("At least one source file is required")
        first_name = safe_filename(files[0][1], "source")
        suffix = Path(first_name).suffix.casefold()
        transcript_in_selection = is_transcript or any(
            Path(safe_filename(name, "source")).suffix.casefold() in TRANSCRIPT_SUFFIXES
            for _, name, _ in files
        )
        if transcript_in_selection and (not meeting_name or not meeting_date):
            raise ValidationError("Meeting name and meeting date are required for transcripts")
        if meeting_date:
            try:
                date.fromisoformat(meeting_date)
            except ValueError as exc:
                raise ValidationError("meeting_date must be an ISO date") from exc
        with self.db.connect() as connection:
            project = self._project(connection, project_id) if project_id else None
        if not multi_project and project is None:
            raise ValidationError("A destination project is required")
        if multi_project and project_id is not None:
            raise ValidationError("Multi-project intake cannot have a project destination before review")
        destination = self._destination_dir(project, multi_project)
        destination.mkdir(parents=True, exist_ok=True)
        ingestion_id = stable_id("I")
        incomplete = self._under_root(destination / f"_INCOMPLETE_{ingestion_id}")
        incomplete.mkdir(parents=False, exist_ok=False)
        (incomplete / "Original").mkdir()
        title = normalize_text(meeting_name or Path(first_name).stem) or "Source"
        source_kind = (
            "meeting-transcript" if transcript_in_selection
            else "email" if suffix in {".msg", ".eml"}
            else "uploaded-folder" if any("/" in relative.replace("\\", "/") for _, _, relative in files)
            else "multiple-files" if len(files) > 1
            else suffix.lstrip(".") or "unknown"
        )
        created_local = datetime.now().astimezone()
        final_package = self._under_root(
            destination / ingestion_folder_name(created_local, source_kind, title, ingestion_id)
        )
        source_committed = False
        total = 0
        limit = self.settings.app.max_file_mb * 1024 * 1024
        stored_files: list[dict[str, Any]] = []
        used_paths: set[str] = set()
        try:
            for stream, original_name, relative_name in files:
                relative = self._safe_relative_path(relative_name, safe_filename(original_name, "source"))
                relative_key = relative.as_posix().casefold()
                original_relative = relative
                collision = 2
                while relative_key in used_paths:
                    relative = original_relative.with_name(
                        f"{original_relative.stem}-{collision}{original_relative.suffix}"
                    )
                    relative_key = relative.as_posix().casefold()
                    collision += 1
                used_paths.add(relative_key)
                output_path = self._under_root(incomplete / "Original" / relative)
                output_path.parent.mkdir(parents=True, exist_ok=True)
                with output_path.open("xb") as output:
                    while True:
                        block = stream.read(1024 * 1024)
                        if not block:
                            break
                        total += len(block)
                        if total > limit:
                            raise ValidationError(
                                f"Ingestion exceeds the configured {self.settings.app.max_file_mb} MB limit"
                            )
                        output.write(block)
                stored_files.append({
                    "relative_path": f"Original/{relative.as_posix()}",
                    "original_name": original_name,
                    "stored_name": relative.name,
                    "size_bytes": output_path.stat().st_size,
                    "sha256": sha256_file(output_path),
                })
            digest = (
                stored_files[0]["sha256"] if len(stored_files) == 1
                else hashlib.sha256(_json(sorted([
                    (item["relative_path"], item["sha256"]) for item in stored_files
                ])).encode("utf-8")).hexdigest()
            )
            primary_incomplete = incomplete / stored_files[0]["relative_path"]
            native_id = inspect_native_id(primary_incomplete) if len(stored_files) == 1 else None
            with self.db.connect() as connection:
                existing = self._find_existing_source(
                    connection, project_id=project_id, digest=digest,
                    native_id=native_id, multi_project=multi_project,
                )
            if existing:
                shutil.rmtree(incomplete)
                if existing["memory_state"] == "removed" and project_id is not None:
                    restored = self.restore_source_to_memory(
                        project_id,
                        int(existing["id"]),
                        reason="Restored because the same source was uploaded again.",
                    )
                    return restored, False
                return self._decode_source(existing), True
            now = utc_now()
            self._initial_assistant_files(incomplete, ingestion_id, project_id, title)
            atomic_write_json(incomplete / "manifest.json", {
                "schema_version": SCHEMA_VERSION,
                "ingestion_id": ingestion_id,
                "project_id": project["archive_id"] if project else None,
                "database_project_id": project_id,
                "source_type": source_kind,
                "title": title,
                "created_at": now,
                "source_date": meeting_date,
                "capture_method": capture_method,
                "canonical_source": True,
                "linked_ingestion_id": None,
                "processing_status": "captured",
                "memory_state": "pending",
                "project_fit_confirmed": bool(multi_project or project_id is None),
                "original_files": stored_files,
                "assistant_files": [
                    "Assistant/source-summary.md", "Assistant/index.json",
                    "Assistant/knowledge-items.json", "Assistant/citations.json",
                    "Assistant/source-lifecycle.jsonl",
                ],
                "extractor_version": "1.0",
                "errors": [],
            })
            os.replace(incomplete, final_package)
            primary_final = final_package / stored_files[0]["relative_path"]
            insert_error: sqlite3.IntegrityError | None = None
            concurrent = None
            with self.db.transaction() as connection:
                try:
                    cursor = connection.execute(
                        """
                        INSERT INTO sources(
                          project_id, source_type, native_id, sha256, original_filename, original_path,
                          metadata_json, meeting_name, meeting_date, processing_state, created_at,
                          ingestion_id, ingestion_path, source_title, source_date, capture_method,
                          canonical_source, processing_version, memory_state, project_fit_confirmed,
                          memory_state_changed_at
                        ) VALUES (?, ?, ?, ?, ?, ?, '{}', ?, ?, 'captured', ?, ?, ?, ?, ?, ?, 1, 1,
                                  'pending', ?, ?)
                        """,
                        (project_id, source_kind, native_id, digest, first_name, str(primary_final),
                         normalize_text(meeting_name or "") or None, meeting_date, now, ingestion_id,
                         str(final_package), title, meeting_date, capture_method,
                         int(multi_project or project_id is None), now),
                    )
                    source_id = int(cursor.lastrowid)
                    for item in stored_files:
                        connection.execute(
                            """INSERT INTO original_files(
                               source_id, relative_path, original_name, stored_name, size_bytes, sha256, created_at
                               ) VALUES (?, ?, ?, ?, ?, ?, ?)""",
                            (source_id, item["relative_path"], item["original_name"], item["stored_name"],
                             item["size_bytes"], item["sha256"], now),
                        )
                    row = connection.execute("SELECT * FROM sources WHERE id = ?", (cursor.lastrowid,)).fetchone()
                except sqlite3.IntegrityError as exc:
                    insert_error = exc
                    concurrent = self._find_existing_source(
                        connection, project_id=project_id, digest=digest,
                        native_id=native_id, multi_project=multi_project,
                    )
            if insert_error is not None:
                shutil.rmtree(final_package)
                if concurrent:
                    if concurrent["memory_state"] == "removed" and project_id is not None:
                        restored = self.restore_source_to_memory(
                            project_id,
                            int(concurrent["id"]),
                            reason="Restored because the same source was uploaded again.",
                        )
                        return restored, False
                    return self._decode_source(concurrent), True
                raise ConflictError("Source insert conflict") from insert_error
            source_committed = True
            return self._decode_source(row), False
        except Exception:
            if final_package.exists() and not source_committed:
                recovery = destination / f"_INCOMPLETE_{ingestion_id}"
                if not recovery.exists():
                    os.replace(final_package, recovery)
            elif incomplete.exists():
                shutil.rmtree(incomplete)
            raise

    def capture_note(
        self, project_id: str, text: str, title: str = "Manual note", *,
        is_transcript: bool = False, meeting_name: str | None = None,
        meeting_date: str | None = None,
    ) -> dict[str, Any]:
        body = text.strip()
        if not body:
            raise ValidationError("Manual note cannot be blank")
        filename = "transcript-as-submitted.txt" if is_transcript else f"{safe_filename(title, 'Manual note')}.txt"
        source, _ = self.capture_source(
            io.BytesIO(body.encode("utf-8")), filename, project_id=project_id,
            meeting_name=meeting_name, meeting_date=meeting_date, is_transcript=is_transcript,
            capture_method="pasted_text",
        )
        return source

    def recover_interrupted(self) -> int:
        with self.db.transaction() as connection:
            cursor = connection.execute(
                """UPDATE sources SET processing_state = 'captured', error_code = 'recovered_after_restart',
                   error_message = NULL WHERE processing_state = 'processing'
                   AND memory_state <> 'removed' AND parent_source_id IS NULL"""
            )
            children = [int(row["id"]) for row in connection.execute(
                """SELECT id FROM sources WHERE processing_state = 'processing'
                   AND memory_state <> 'removed' AND parent_source_id IS NOT NULL"""
            ).fetchall()]
        for child_id in children:
            self._extract_attachment_child(child_id)
        return cursor.rowcount + len(children)

    def process_pending(self, *, manual: bool = False, source_id: int | None = None, limit: int = 20) -> dict[str, int]:
        clauses = ["parent_source_id IS NULL", "memory_state <> 'removed'"]
        params: list[Any] = []
        if source_id is not None:
            clauses.append("id = ?")
            params.append(source_id)
            clauses.append("processing_state IN ('captured', 'pending_ai', 'error')")
        else:
            clauses.append("processing_state IN ('captured', 'pending_ai')")
            if not manual:
                clauses.append("retry_count < ?")
                params.append(self.settings.app.automatic_ai_attempts)
        with self.db.connect() as connection:
            rows = connection.execute(
                f"SELECT id, project_id FROM sources WHERE {' AND '.join(clauses)} ORDER BY created_at LIMIT ?",
                (*params, max(1, min(limit, 100))),
            ).fetchall()
        counts = {"processed": 0, "pending_ai": 0, "needs_review": 0, "unsupported": 0, "error": 0}
        for row in rows:
            state = self.process_source(int(row["id"])) if row["project_id"] else self.process_multi_source(int(row["id"]))
            if state == "complete":
                counts["processed"] += 1
            elif state in counts:
                counts[state] += 1
        return counts

    def retry_source(self, source_id: int) -> dict[str, int]:
        with self.db.connect() as connection:
            source = connection.execute("SELECT * FROM sources WHERE id = ?", (source_id,)).fetchone()
        if not source:
            raise NotFoundError("Source not found")
        if source["memory_state"] == "removed":
            raise ConflictError("Removed sources cannot be retried; restore the source first")
        if source["processing_state"] == "needs_review":
            with self.db.transaction() as connection:
                open_reviews = connection.execute(
                    "SELECT id, kind FROM review_items WHERE source_id = ? AND status = 'open'",
                    (source_id,),
                ).fetchall()
                if open_reviews and all(item["kind"] == "malformed_llm" for item in open_reviews):
                    now = utc_now()
                    connection.execute(
                        """UPDATE review_items SET status = 'dismissed', resolution_json = ?, resolved_at = ?
                           WHERE source_id = ? AND status = 'open' AND kind = 'malformed_llm'""",
                        (_json({"action": "retry"}), now, source_id),
                    )
                    connection.execute(
                        """UPDATE sources SET processing_state = 'captured', error_code = NULL,
                           error_message = NULL WHERE id = ?""",
                        (source_id,),
                    )
                    source = connection.execute(
                        "SELECT * FROM sources WHERE id = ?", (source_id,)
                    ).fetchone()
        counts = {
            "processed": 0, "pending_ai": 0, "needs_review": 0,
            "unsupported": 0, "error": 0, "already_complete": 0,
        }
        if source["parent_source_id"] is None:
            if source["processing_state"] == "complete":
                counts["already_complete"] = 1
                return counts
            if source["processing_state"] in {"needs_review", "unsupported"}:
                counts[source["processing_state"]] = 1
                return counts
            result = self.process_pending(manual=True, source_id=source_id, limit=1)
            counts.update(result)
            return counts
        if source["processing_state"] != "error":
            if source["processing_state"] == "complete":
                counts["already_complete"] = 1
            elif source["processing_state"] in counts:
                counts[source["processing_state"]] = 1
            return counts
        with self.db.transaction() as connection:
            connection.execute(
                "UPDATE sources SET processing_state = 'processing', error_code = NULL, error_message = NULL WHERE id = ?",
                (source_id,),
            )
        self._extract_attachment_child(source_id)
        with self.db.connect() as connection:
            state = connection.execute("SELECT processing_state FROM sources WHERE id = ?", (source_id,)).fetchone()[0]
        if state == "complete":
            counts["processed"] = 1
        elif state in counts:
            counts[state] = 1
        return counts

    def process_source(self, source_id: int) -> str:
        with self.db.transaction() as connection:
            source = connection.execute("SELECT * FROM sources WHERE id = ?", (source_id,)).fetchone()
            if not source:
                raise NotFoundError("Source not found")
            if source["memory_state"] == "removed":
                raise ConflictError("Removed sources cannot be processed; restore the source first")
            if source["parent_source_id"] is not None:
                raise ValidationError("Attachment sources are processed with their parent")
            if source["processing_state"] == "complete":
                return "complete"
            if source["processing_state"] == "processing":
                return "processing"
            if source["project_id"] is None:
                raise ValidationError("Use multi-project processing for unassigned intake")
            claimed = connection.execute(
                """UPDATE sources SET processing_state = 'processing', error_code = NULL, error_message = NULL
                   WHERE id = ? AND processing_state <> 'processing'""",
                (source_id,),
            )
            if claimed.rowcount != 1:
                return "processing"
        try:
            self._ensure_extracted(source_id)
            evidence = self._source_evidence(source_id)
            if not evidence:
                raise UnsupportedSource("No supported text could be extracted")
            with self.db.connect() as connection:
                source = connection.execute("SELECT * FROM sources WHERE id = ?", (source_id,)).fetchone()
                project = self._project(connection, source["project_id"])
            bounded, dropped = self._bounded_evidence(evidence)
            self._record_evidence_window(source_id, dropped)
            if not bounded:
                raise LlmContractError("No complete evidence chunk fits the configured evidence limit")
            if not bool(source["project_fit_confirmed"]) and self._require_project_fit_review(
                source_id, project, bounded
            ):
                return "needs_review"
            result = self.llm.knowledge_update(project["current_summary"], bounded)
            self._commit_knowledge(source_id, result, bounded)
            return "complete"
        except UnsupportedSource as exc:
            self._set_source_state(source_id, "unsupported", "unsupported_type", exc)
            return "unsupported"
        except LlmUnavailable as exc:
            self._set_source_state(source_id, "pending_ai", "llm_unavailable", exc, increment_retry=True)
            return "pending_ai"
        except (LlmContractError, ValidationError) as exc:
            self._create_review(
                kind="malformed_llm", source_id=source_id, project_id=self._source_project_id(source_id),
                question="How should this source update the project?", reason=_safe_error(exc),
                evidence=[], options=[], memory_preview="No routing rule will be created.",
            )
            self._set_source_state(source_id, "needs_review", "llm_contract", exc)
            return "needs_review"
        except ExtractionFailure as exc:
            self._set_source_state(source_id, "error", "extraction_failed", exc)
            return "error"
        except UnexpectedSummaryError:
            # Knowledge and the source were committed before summary generation began.
            return "complete"
        except Exception as exc:
            self._set_source_state(source_id, "error", "processing_failed", exc)
            return "error"

    def _require_project_fit_review(
        self, source_id: int, selected_project: sqlite3.Row, evidence: list[dict[str, Any]]
    ) -> bool:
        with self.db.connect() as connection:
            projects = [dict(row) for row in connection.execute(
                "SELECT id, name, snow_number FROM projects ORDER BY name COLLATE NOCASE"
            ).fetchall()]
        result = self.llm.project_fit(
            {
                "id": selected_project["id"], "name": selected_project["name"],
                "snow_number": selected_project["snow_number"],
            },
            evidence,
            projects,
        )
        if not isinstance(result, dict):
            raise LlmContractError("Project-fit response must be a JSON object")
        try:
            selected_confidence = float(result["selected_project_confidence"])
            confidence = float(result["confidence"])
        except (KeyError, TypeError, ValueError) as exc:
            raise LlmContractError("Project-fit confidence values are invalid") from exc
        if not (0 <= selected_confidence <= 1 and 0 <= confidence <= 1):
            raise LlmContractError("Project-fit confidence values must be between 0 and 1")
        project_ids = {str(item["id"]) for item in projects}
        recommended = str(result.get("recommended_project_id") or selected_project["id"])
        if recommended not in project_ids:
            raise LlmContractError("Project-fit response recommended an unknown project")
        allowed = {(int(item["source_id"]), int(item["chunk_id"])) for item in evidence}
        citations = self._validate_citations(result.get("citations"), allowed)
        needs_review = (
            bool(result.get("needs_review"))
            or selected_confidence < PROJECT_FIT_SELECTED_REVIEW_THRESHOLD
            or (
                recommended != selected_project["id"]
                and confidence >= PROJECT_FIT_ALTERNATIVE_CONFIDENCE_THRESHOLD
            )
        )
        reason = normalize_text(str(result.get("reason") or "The selected project fit is uncertain."))
        if needs_review:
            cited = {(item["source_id"], item["chunk_id"]) for item in citations}
            review_evidence = [{
                "text": item["text"],
                "citations": [{"source_id": item["source_id"], "chunk_id": item["chunk_id"]}],
                "cited_by_project_fit": (item["source_id"], item["chunk_id"]) in cited,
                "selected_project_confidence": selected_confidence,
                "recommended_project_id": recommended,
                "confidence": confidence,
                "model_id": self.llm.model_for("project_fit_check"),
            } for item in evidence]
            self._create_review(
                kind="project_fit", source_id=source_id, project_id=selected_project["id"],
                question=f"Does this source belong in {selected_project['name']}?",
                reason=reason,
                evidence=review_evidence,
                options=[
                    {"project_id": item["id"], "label": item["name"]}
                    for item in projects if item["id"] != selected_project["id"]
                ],
                memory_preview="No project memory changes until you confirm where this source belongs.",
            )
            self._set_source_state(
                source_id, "needs_review", "project_fit_uncertain",
                ValidationError("The selected project fit requires review"),
            )
            return True
        now = utc_now()
        with self.db.transaction() as connection:
            source = connection.execute("SELECT * FROM sources WHERE id = ?", (source_id,)).fetchone()
            if not source:
                raise NotFoundError("Source not found")
            connection.execute(
                "UPDATE sources SET project_fit_confirmed = 1, memory_state_changed_at = ? WHERE id = ?",
                (now, source_id),
            )
            self._record_source_lifecycle(
                connection, source, "project_fit_confirmed", reason,
                from_project_id=source["project_id"], to_project_id=source["project_id"],
                details={
                    "selected_project_confidence": selected_confidence,
                    "confidence": confidence,
                    "model_id": self.llm.model_for("project_fit_check"),
                },
            )
        self._refresh_source_archive(source_id)
        return False

    def _source_project_id(self, source_id: int) -> str | None:
        with self.db.connect() as connection:
            row = connection.execute("SELECT project_id FROM sources WHERE id = ?", (source_id,)).fetchone()
        return row["project_id"] if row else None

    def _set_source_state(
        self, source_id: int, state: str, code: str, error: Exception, *, increment_retry: bool = False
    ) -> None:
        with self.db.transaction() as connection:
            connection.execute(
                """UPDATE sources SET processing_state = ?, error_code = ?, error_message = ?,
                   retry_count = retry_count + ? WHERE id = ?""",
                (state, code, _safe_error(error), int(increment_retry), source_id),
            )
            row = connection.execute(
                "SELECT id, parent_source_id FROM sources WHERE id = ?", (source_id,)
            ).fetchone()
        if row:
            self._refresh_source_archive(int(row["parent_source_id"] or row["id"]), processing_status=state)

    def _ensure_extracted(self, source_id: int) -> None:
        with self.db.connect() as connection:
            source = connection.execute("SELECT * FROM sources WHERE id = ?", (source_id,)).fetchone()
            existing = connection.execute(
                "SELECT count(*) FROM source_chunks WHERE source_id = ? OR source_id IN (SELECT id FROM sources WHERE parent_source_id = ?)",
                (source_id, source_id),
            ).fetchone()[0]
            originals = connection.execute(
                "SELECT * FROM original_files WHERE source_id = ? AND is_attachment = 0 ORDER BY id",
                (source_id,),
            ).fetchall()
        metadata = _decode(source["metadata_json"], {})
        if existing and (metadata.get("_extraction_complete") is True or source["source_type"] in {"snow_comments", "routed_segment"}):
            return
        package = Path(source["ingestion_path"] or Path(source["original_path"]).parent)
        if not originals:
            originals = [{
                "relative_path": Path(source["original_path"]).name,
                "original_name": source["original_filename"],
            }]
        results: list[tuple[Any, Any, Path]] = []
        unsupported_errors: list[Exception] = []
        for original in originals:
            path = package / original["relative_path"] if source["ingestion_path"] else Path(source["original_path"])
            try:
                result = extract_source(
                    path,
                    max_attachments=self.settings.app.max_attachments,
                    max_text_bytes=self.settings.app.max_extracted_text_mb * 1024 * 1024,
                )
                results.append((original, result, path))
            except UnsupportedSource as exc:
                unsupported_errors.append(exc)
            except ExtractionFailure as exc:
                raise ExtractionFailure(f"{original['relative_path']}: {exc}") from exc
        if not results:
            raise unsupported_errors[0] if unsupported_errors else UnsupportedSource("No supported text could be extracted")
        combined_metadata = dict(metadata)
        sequence = 0
        with self.db.transaction() as connection:
            for original, result, path in results:
                combined_metadata.update(result.metadata)
                for chunk in result.chunks:
                    locator = (
                        chunk["locator"] if len(originals) == 1
                        else f"{original['relative_path']}: {chunk['locator']}"
                    )
                    connection.execute(
                        "INSERT OR IGNORE INTO source_chunks(source_id, project_id, sequence, text, locator, processing_state) VALUES (?, ?, ?, ?, ?, 'captured')",
                        (source_id, source["project_id"], sequence, chunk["text"], locator),
                    )
                    sequence += 1
            primary_result = results[0][1]
            connection.execute(
                "UPDATE sources SET native_id = coalesce(native_id, ?), source_type = ?, metadata_json = ? WHERE id = ?",
                (primary_result.native_id,
                 primary_result.source_type if len(results) == 1 else "multiple-files",
                 _json({**combined_metadata, "_extraction_complete": False}), source_id),
            )
        extracted_dir = package / "Assistant" / "Extracted"
        original_positions = {
            str(item["relative_path"]): index for index, item in enumerate(originals, start=1)
        }
        for original, result, path in results:
            index = original_positions[str(original["relative_path"])]
            extracted_name = f"{index:03d}-{safe_filename(Path(original['relative_path']).stem, 'source')}.txt"
            atomic_write_text(
                extracted_dir / extracted_name,
                "\n\n".join(chunk["text"] for chunk in result.chunks) + "\n",
            )
            if result.attachments:
                self._preserve_attachments(source_id, result.attachments)
        with self.db.transaction() as connection:
            connection.execute(
                "UPDATE sources SET metadata_json = ? WHERE id = ?",
                (_json({**combined_metadata, "_extraction_complete": True}), source_id),
            )
        self._refresh_source_archive(source_id, processing_status="processing")

    def _preserve_attachments(self, source_id: int, attachments: Iterable[Any]) -> None:
        with self.db.connect() as connection:
            source = connection.execute("SELECT * FROM sources WHERE id = ?", (source_id,)).fetchone()
        root_source = source
        while root_source["parent_source_id"]:
            with self.db.connect() as connection:
                root_source = connection.execute(
                    "SELECT * FROM sources WHERE id = ?", (root_source["parent_source_id"],)
                ).fetchone()
        package = Path(root_source["ingestion_path"])
        directory = self._under_root(package / "Original" / "Attachments")
        if source["parent_source_id"]:
            directory = self._under_root(directory / safe_filename(Path(source["original_filename"]).stem, "nested-email"))
        directory.mkdir(parents=True, exist_ok=True)
        for sequence, attachment in enumerate(attachments):
            native_id = f"attachment:{source_id}:{sequence}"
            with self.db.connect() as connection:
                existing_child = connection.execute(
                    """SELECT id, processing_state FROM sources
                       WHERE project_id IS ? AND parent_source_id = ? AND native_id = ?""",
                    (source["project_id"], source_id, native_id),
                ).fetchone()
            if existing_child:
                if existing_child["processing_state"] != "complete":
                    self._extract_attachment_child(int(existing_child["id"]))
                continue
            filename = safe_filename(attachment.filename, f"attachment-{sequence + 1}")
            final_path = self._under_root(directory / filename)
            digest = hashlib.sha256(attachment.data).hexdigest()
            collision = 2
            while final_path.exists():
                relative_path = relative_to_root(final_path, package)
                with self.db.connect() as connection:
                    claimed = connection.execute(
                        """SELECT 1 FROM original_files
                           WHERE source_id = ? AND relative_path = ?""",
                        (root_source["id"], relative_path),
                    ).fetchone()
                if not claimed and sha256_file(final_path) == digest:
                    # A prior attempt can rename the attachment before its source
                    # transaction begins. Reclaim that exact unindexed file rather
                    # than leaving it orphaned and creating a suffixed duplicate.
                    break
                final_path = self._under_root(
                    directory / f"{Path(filename).stem}-{collision}{Path(filename).suffix}"
                )
                collision += 1
            if not final_path.exists():
                temp_path = self._under_root(directory / f".{uuid.uuid4().hex}.tmp")
                temp_path.write_bytes(attachment.data)
                os.replace(temp_path, final_path)
            now = utc_now()
            with self.db.transaction() as connection:
                cursor = connection.execute(
                    """
                    INSERT OR IGNORE INTO sources(project_id, parent_source_id, source_type, native_id, sha256,
                      original_filename, original_path, metadata_json, processing_state, created_at,
                      memory_state, project_fit_confirmed, memory_state_changed_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'processing', ?, 'pending', 0, ?)
                    """,
                    (source["project_id"], source_id, Path(filename).suffix.casefold().lstrip(".") or "attachment",
                     native_id, digest, filename, str(final_path),
                     _json({"content_type": attachment.content_type}), now, now),
                )
                if cursor.rowcount == 1:
                    child_id = int(cursor.lastrowid)
                else:
                    existing_child = connection.execute(
                        """SELECT id, processing_state FROM sources
                           WHERE project_id IS ? AND parent_source_id = ? AND native_id = ?""",
                        (source["project_id"], source_id, native_id),
                    ).fetchone()
                    if not existing_child or existing_child["processing_state"] == "complete":
                        continue
                    child_id = int(existing_child["id"])
                relative_path = relative_to_root(final_path, package)
                connection.execute(
                    """INSERT OR IGNORE INTO original_files(
                       source_id, relative_path, original_name, stored_name, size_bytes, sha256, is_attachment, created_at
                       ) VALUES (?, ?, ?, ?, ?, ?, 1, ?)""",
                    (root_source["id"], relative_path, attachment.filename, final_path.name,
                     final_path.stat().st_size, digest, now),
                )
                connection.execute(
                    """UPDATE sources SET ingestion_path = ?, capture_method = 'email_attachment',
                       canonical_source = 0 WHERE id = ?""",
                    (str(package), child_id),
                )
            self._extract_attachment_child(child_id)
        self._refresh_source_archive(int(root_source["id"]), processing_status="processing")

    def _extract_attachment_child(self, child_id: int) -> None:
        with self.db.connect() as connection:
            child = connection.execute("SELECT * FROM sources WHERE id = ?", (child_id,)).fetchone()
            if not child or child["parent_source_id"] is None or child["processing_state"] == "complete":
                return
        try:
            result = extract_source(
                Path(child["original_path"]),
                max_attachments=self.settings.app.max_attachments,
                max_text_bytes=self.settings.app.max_extracted_text_mb * 1024 * 1024,
            )
            prior_metadata = _decode(child["metadata_json"], {})
            with self.db.transaction() as connection:
                for sequence, chunk in enumerate(result.chunks):
                    connection.execute(
                        """INSERT OR IGNORE INTO source_chunks(
                           source_id, project_id, sequence, text, locator, processing_state
                           ) VALUES (?, ?, ?, ?, ?, 'captured')""",
                        (child_id, child["project_id"], sequence, chunk["text"],
                         f"Attachment {child['original_filename']}: {chunk['locator']}"),
                    )
                connection.execute(
                    """UPDATE sources SET source_type = ?, metadata_json = ?, processing_state = 'complete',
                       processed_at = ?, error_code = NULL, error_message = NULL WHERE id = ?""",
                    (result.source_type, _json({**prior_metadata, **result.metadata}), utc_now(), child_id),
                )
            if result.attachments:
                self._preserve_attachments(child_id, result.attachments)
            parent = child["parent_source_id"]
            with self.db.connect() as connection:
                root = connection.execute("SELECT * FROM sources WHERE id = ?", (parent,)).fetchone()
            if root and root["ingestion_path"]:
                atomic_write_text(
                    Path(root["ingestion_path"]) / "Assistant" / "Extracted" /
                    f"attachment-{child_id}-{safe_filename(Path(child['original_filename']).stem, 'attachment')}.txt",
                    "\n\n".join(chunk["text"] for chunk in result.chunks) + "\n",
                )
        except UnsupportedSource as exc:
            self._set_source_state(child_id, "unsupported", "unsupported_attachment", exc)
        except ExtractionFailure as exc:
            self._set_source_state(child_id, "error", "attachment_extraction_failed", exc)
        except Exception as exc:
            self._set_source_state(child_id, "error", "attachment_processing_failed", exc)

    def _source_evidence(self, source_id: int) -> list[dict[str, Any]]:
        with self.db.connect() as connection:
            rows = connection.execute(
                """
                WITH RECURSIVE source_tree(id) AS (
                  SELECT ?
                  UNION ALL
                  SELECT s.id FROM sources s JOIN source_tree t ON s.parent_source_id = t.id
                )
                SELECT c.id AS chunk_id, c.source_id, c.text, c.locator, c.comment_at, c.author,
                       s.original_filename, s.meeting_name, s.meeting_date
                FROM source_chunks c JOIN sources s ON s.id = c.source_id
                WHERE c.source_id IN (SELECT id FROM source_tree)
                ORDER BY c.source_id = ? DESC, c.source_id, c.sequence
                """,
                (source_id, source_id),
            ).fetchall()
        return [dict(row) for row in rows]

    def _bounded_evidence(self, evidence: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
        remaining = self.settings.llm.max_evidence_chars
        bounded: list[dict[str, Any]] = []
        dropped = 0
        for item in evidence:
            copy = dict(item)
            text = str(copy["text"])
            if len(text) > remaining:
                dropped += 1
                continue
            copy["text"] = text
            remaining -= len(text)
            bounded.append(copy)
        return bounded, dropped

    def _record_evidence_window(self, source_id: int, dropped_chunks: int) -> None:
        with self.db.transaction() as connection:
            cursor = connection.execute(
                """UPDATE sources
                   SET metadata_json = json_set(coalesce(metadata_json, '{}'), '$.evidence_dropped_chunks', ?)
                   WHERE id = ?""",
                (dropped_chunks, source_id),
            )
            if cursor.rowcount != 1:
                raise NotFoundError("Source not found")

    def _validate_citations(self, citations: Any, allowed: set[tuple[int, int]]) -> list[dict[str, int]]:
        if not isinstance(citations, list) or not citations:
            raise LlmContractError("Every material update requires a citation")
        validated = []
        for citation in citations:
            if not isinstance(citation, dict):
                raise LlmContractError("Citation must be an object")
            try:
                pair = (int(citation["source_id"]), int(citation["chunk_id"]))
            except (KeyError, TypeError, ValueError) as exc:
                raise LlmContractError("Citation identifiers are invalid") from exc
            if pair not in allowed:
                raise LlmContractError("AI response cited evidence that was not supplied")
            validated.append({"source_id": pair[0], "chunk_id": pair[1]})
        return validated

    def _commit_knowledge(self, source_id: int, result: dict[str, Any], evidence: list[dict[str, Any]]) -> None:
        if not isinstance(result, dict):
            raise LlmContractError("Knowledge update must be a JSON object")
        if result.get("needs_review"):
            raise LlmContractError("The AI response requires user review")
        summary = result.get("updated_summary")
        updates = result.get("updates")
        if not isinstance(summary, str) or not summary.strip() or not isinstance(updates, list) or not updates:
            raise LlmContractError("Knowledge update is missing a summary or cited updates")
        allowed = {(int(item["source_id"]), int(item["chunk_id"])) for item in evidence}
        prepared_updates = []
        for update in updates:
            if not isinstance(update, dict) or not normalize_text(str(update.get("text", ""))):
                raise LlmContractError("Knowledge update text is missing")
            prepared_updates.append((normalize_text(update["text"]), self._validate_citations(update.get("citations"), allowed)))
        operations = result.get("action_item_operations", [])
        if not isinstance(operations, list):
            raise LlmContractError("action_item_operations must be an array")
        recommendations = result.get("project_field_recommendations", [])
        if not isinstance(recommendations, list):
            raise LlmContractError("project_field_recommendations must be an array")
        now = utc_now()
        with self.db.transaction() as connection:
            source = connection.execute("SELECT * FROM sources WHERE id = ?", (source_id,)).fetchone()
            if not source:
                raise NotFoundError("Source not found")
            if source["processing_state"] == "complete":
                return
            project_id = source["project_id"]
            project = self._project(connection, project_id)
            for text, citations in prepared_updates:
                durable_citations = [self._ensure_citation_record(connection, source, citation, now) for citation in citations]
                connection.execute(
                    "INSERT INTO project_updates(project_id, source_id, update_type, text, citations_json, model_id, created_at) VALUES (?, ?, 'knowledge', ?, ?, ?, ?)",
                    (project_id, source_id, text, _json(durable_citations), self.llm.model_id, now),
                )
                connection.execute(
                    """INSERT INTO knowledge_items(
                       id, project_id, source_id, text, category, source_date, citation_ids_json,
                       review_status, created_at
                       ) VALUES (?, ?, ?, ?, ?, ?, ?, 'unreviewed', ?)""",
                    (stable_id("K"), project_id, source_id, text, self._knowledge_category(text),
                     source["source_date"] or source["meeting_date"] or source["created_at"],
                     _json([citation["citation_id"] for citation in durable_citations]), now),
                )
            connection.execute(
                """UPDATE projects SET latest_change = ?, updated_at = ?,
                   summary_revision = summary_revision + 1, summary_generation_state = 'stale',
                   summary_error = NULL WHERE id = ?""",
                (prepared_updates[-1][0], now, project_id),
            )
            for operation in operations:
                self._apply_ai_action(connection, project_id, source_id, operation, allowed, now)
            for recommendation in recommendations:
                self._insert_review(
                    connection, "project_field_recommendation", source_id, project_id,
                    "Do you want to change a manual project field?",
                    "AI recommendations never change project status or priority automatically.",
                    [recommendation], [], "No routing rule will be created.", now,
                )
            connection.execute(
                """WITH RECURSIVE source_tree(id) AS (
                     SELECT ? UNION ALL
                     SELECT s.id FROM sources s JOIN source_tree t ON s.parent_source_id = t.id
                   )
                   UPDATE source_chunks SET processing_state = 'complete', processed_at = ?
                   WHERE source_id IN (SELECT id FROM source_tree)""",
                (source_id, now),
            )
            connection.execute(
                """WITH RECURSIVE source_tree(id) AS (
                     SELECT ? UNION ALL
                     SELECT s.id FROM sources s JOIN source_tree t ON s.parent_source_id = t.id
                   )
                   UPDATE sources SET memory_state = 'active', memory_state_changed_at = ?
                   WHERE id IN (SELECT id FROM source_tree)""",
                (source_id, now),
            )
            connection.execute(
                """UPDATE sources SET processing_state = 'complete', model_id = ?, processed_at = ?,
                   error_code = NULL, error_message = NULL, retry_count = 0 WHERE id = ?""",
                (self.llm.model_id, now, source_id),
            )
            metadata = _decode(source["metadata_json"], {})
            if source["source_type"] == "snow_comments" and metadata.get("cell_hash"):
                connection.execute(
                    "UPDATE projects SET snow_comments_cell_hash = ? WHERE id = ?",
                    (metadata["cell_hash"], project_id),
                )
            source_summary = " ".join(text for text, _ in prepared_updates)[:1800]
            connection.execute(
                "UPDATE sources SET source_summary = ? WHERE id = ?", (source_summary, source_id)
            )
        self._refresh_source_archive(source_id, processing_status="complete")
        self._write_knowledge_history(project_id)
        self.regenerate_living_summary(project_id, advance_revision=False)

    @staticmethod
    def _store_citation_record(
        connection: sqlite3.Connection, *, source_id: int, chunk_id: int,
        original_relative_path: str, display_name: str, source_type: str,
        locator: str, excerpt: str, source_date: str | None, created_at: str,
        identifier_key: str | None = None, preferred_id: str | None = None,
    ) -> str:
        existing = connection.execute(
            "SELECT id FROM citation_records WHERE source_id = ? AND chunk_id = ?",
            (source_id, chunk_id),
        ).fetchone()
        if existing:
            return str(existing["id"])

        key = identifier_key or f"{source_id}:{chunk_id}"
        digest = hashlib.sha256(key.encode()).hexdigest().upper()
        candidates = ([preferred_id] if preferred_id else []) + [
            f"C-{digest[:length]}" for length in range(8, len(digest) + 1, 4)
        ]
        for candidate in candidates:
            occupied = connection.execute(
                "SELECT source_id, chunk_id FROM citation_records WHERE id = ?", (candidate,)
            ).fetchone()
            if occupied:
                if (int(occupied["source_id"]), int(occupied["chunk_id"])) == (source_id, chunk_id):
                    return candidate
                continue
            connection.execute(
                """INSERT INTO citation_records(
                   id, source_id, chunk_id, original_relative_path, display_name, source_type,
                   locator, excerpt, source_date, created_at
                   ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    candidate, source_id, chunk_id, original_relative_path, display_name,
                    source_type, locator, excerpt, source_date, created_at,
                ),
            )
            return candidate
        raise ConflictError("Could not allocate a unique citation identifier")

    def _ensure_citation_record(
        self, connection: sqlite3.Connection, root_source: sqlite3.Row,
        citation: dict[str, int], now: str,
    ) -> dict[str, Any]:
        row = connection.execute(
            """SELECT c.*, s.original_filename, s.original_path, s.source_type,
                      s.meeting_date, s.created_at, s.parent_source_id
               FROM source_chunks c JOIN sources s ON s.id = c.source_id
               WHERE c.id = ? AND c.source_id = ?""",
            (citation["chunk_id"], citation["source_id"]),
        ).fetchone()
        if not row:
            raise LlmContractError("Citation source excerpt is unavailable")
        package = Path(root_source["ingestion_path"] or Path(root_source["original_path"]).parent)
        original_path = Path(row["original_path"])
        try:
            original_relative = original_path.resolve().relative_to(package.resolve()).as_posix()
        except ValueError:
            original_relative = f"Original/{safe_filename(row['original_filename'], 'source')}"
        citation_id = self._store_citation_record(
            connection,
            source_id=int(root_source["id"]),
            chunk_id=int(row["id"]),
            original_relative_path=original_relative,
            display_name=row["original_filename"],
            source_type=row["source_type"],
            locator=row["locator"],
            excerpt=row["text"][:1200],
            source_date=row["meeting_date"] or row["created_at"],
            created_at=now,
            identifier_key=f"{citation['source_id']}:{citation['chunk_id']}",
        )
        return {
            "source_id": int(citation["source_id"]),
            "chunk_id": int(citation["chunk_id"]),
            "citation_id": citation_id,
        }

    def _write_knowledge_history(self, project_id: str) -> None:
        with self.db.connect() as connection:
            project = self._project(connection, project_id)
            rows = connection.execute(
                """SELECT k.*, s.source_title, s.source_type, s.ingestion_id
                   FROM knowledge_items k JOIN sources s ON s.id = k.source_id
                   WHERE k.project_id = ? AND s.memory_state = 'active'
                   ORDER BY k.created_at, k.id""",
                (project_id,),
            ).fetchall()
        lines = []
        for row in rows:
            item = dict(row)
            item["citation_ids"] = _decode(item.pop("citation_ids_json"), [])
            lines.append(_json(item))
        atomic_write_text(
            Path(project["folder_path"]) / "_Assistant" / "knowledge-history.jsonl",
            ("\n".join(lines) + "\n") if lines else "",
        )

    def _mark_summary_failed(self, project_id: str, revision: int, exc: Exception) -> None:
        with self.db.transaction() as connection:
            current = self._project(connection, project_id)
            if int(current["summary_revision"]) == revision:
                connection.execute(
                    """UPDATE projects SET summary_generation_state = 'failed', summary_error = ?
                       WHERE id = ?""",
                    (_safe_error(exc), project_id),
                )

    def regenerate_living_summary(self, project_id: str, *, advance_revision: bool = True) -> dict[str, Any]:
        with self.db.transaction() as connection:
            project = self._project(connection, project_id)
            revision = int(project["summary_revision"]) + (1 if advance_revision else 0)
            connection.execute(
                """UPDATE projects SET summary_revision = ?, summary_generation_state = 'updating',
                   summary_error = NULL WHERE id = ?""",
                (revision, project_id),
            )
        try:
            with self.db.connect() as connection:
                project = self._project(connection, project_id)
                knowledge = [dict(row) for row in connection.execute(
                    """SELECT k.id, k.text, k.category, k.source_date, k.review_status, k.created_at
                       FROM knowledge_items k JOIN sources s ON s.id = k.source_id
                       WHERE k.project_id = ? AND k.review_status <> 'flagged'
                         AND s.memory_state = 'active'
                       ORDER BY k.created_at, k.id""",
                    (project_id,),
                ).fetchall()]
            generated = self.llm.living_summary(
                {"id": project["archive_id"], "name": project["name"]}, knowledge
            )
            sections = generated.get("sections") if isinstance(generated, dict) else None
            if not isinstance(sections, list):
                raise LlmContractError("Living summary response did not contain sections")
            allowed = {item["id"] for item in knowledge}
            prepared: list[dict[str, Any]] = []
            for section in sections:
                if not isinstance(section, dict):
                    raise LlmContractError("Living summary section must be an object")
                heading = normalize_text(str(section.get("section", "")))
                text = normalize_text(str(section.get("text", "")))
                ids = section.get("knowledge_item_ids")
                if not heading or not text or not isinstance(ids, list) or not ids:
                    raise LlmContractError("Living summary section is missing text or knowledge-item citations")
                normalized_ids = [str(item) for item in ids]
                if any(item not in allowed for item in normalized_ids):
                    raise LlmContractError("Living summary cited an unknown or wrong-project knowledge item")
                prepared.append({"section": heading, "text": text, "knowledge_item_ids": normalized_ids})
            markdown = "\n\n".join(
                f"## {section['section']}\n\n{section['text']}\n\n"
                f"Knowledge: {', '.join(section['knowledge_item_ids'])}"
                for section in prepared
            )
            summary_text = " ".join(section["text"] for section in prepared)
            now = utc_now()
            with self.db.transaction() as connection:
                current = self._project(connection, project_id)
                if int(current["summary_revision"]) != revision:
                    connection.execute(
                        "UPDATE projects SET summary_generation_state = 'stale' WHERE id = ?", (project_id,)
                    )
                    raise ConflictError("New knowledge arrived while the summary was being generated")
                connection.execute(
                    """INSERT INTO summary_versions(
                       project_id, revision, content_json, markdown, review_status,
                       generation_state, model_id, created_at
                       ) VALUES (?, ?, ?, ?, 'unreviewed', 'current', ?, ?)""",
                    (project_id, revision, _json({"sections": prepared}), markdown, self.llm.model_id, now),
                )
                connection.execute(
                    """UPDATE projects SET current_summary = ?, summary_generation_state = 'current',
                       summary_review_status = 'unreviewed', summary_generated_at = ?, summary_error = NULL
                       WHERE id = ?""",
                    (summary_text, now, project_id),
                )
            self._write_summary_files(project_id, revision, prepared, markdown, now)
            return self.get_living_summary(project_id)
        except Exception as exc:
            self._mark_summary_failed(project_id, revision, exc)
            # Operational LLM/filesystem failures are retryable UI states; database and code errors surface.
            if isinstance(exc, (LlmUnavailable, LlmContractError, ConflictError, OSError)):
                return self.get_living_summary(project_id)
            raise UnexpectedSummaryError("Unexpected Living Summary generation failure") from exc

    def _write_summary_files(
        self, project_id: str, revision: int, sections: list[dict[str, Any]],
        markdown: str, created_at: str,
    ) -> None:
        with self.db.connect() as connection:
            project = self._project(connection, project_id)
        folder = Path(project["folder_path"]) / "_Assistant" / "living-summary"
        payload = {
            "schema_version": SCHEMA_VERSION,
            "project_id": project["archive_id"],
            "revision": revision,
            "review_status": "unreviewed",
            "generation_state": "current",
            "created_at": created_at,
            "sections": sections,
        }
        atomic_write_text(folder / "current.md", f"# {project['name']} — Living Summary\n\n{markdown}\n")
        atomic_write_json(folder / "current.json", payload)
        stamp = datetime.fromisoformat(created_at.replace("Z", "+00:00")).strftime("%Y-%m-%d_%H%M%S")
        atomic_write_json(folder / "versions" / f"{revision:06d}__{stamp}.json", payload)

    def get_living_summary(self, project_id: str) -> dict[str, Any]:
        with self.db.connect() as connection:
            project = self._project(connection, project_id)
            versions = connection.execute(
                "SELECT * FROM summary_versions WHERE project_id = ? ORDER BY revision DESC", (project_id,)
            ).fetchall()
        current_row = next(
            (row for row in versions if int(row["revision"]) == int(project["summary_revision"])), None
        )
        current = dict(current_row) if current_row else None
        if current:
            current["content"] = _decode(current.pop("content_json"), {"sections": []})
        return {
            "project_id": project_id,
            "generation_state": project["summary_generation_state"],
            "review_status": project["summary_review_status"],
            "revision": project["summary_revision"],
            "generated_at": project["summary_generated_at"],
            "error": project["summary_error"],
            "current": current,
            "versions": [{
                "id": row["id"], "revision": row["revision"], "review_status": row["review_status"],
                "generation_state": row["generation_state"], "created_at": row["created_at"],
            } for row in versions],
        }

    def get_summary_version(self, project_id: str, revision: int) -> dict[str, Any]:
        with self.db.connect() as connection:
            self._project(connection, project_id)
            row = connection.execute(
                "SELECT * FROM summary_versions WHERE project_id = ? AND revision = ?",
                (project_id, revision),
            ).fetchone()
        if not row:
            raise NotFoundError("Living summary version not found")
        result = dict(row)
        result["content"] = _decode(result.pop("content_json"), {"sections": []})
        return result

    def list_knowledge(
        self, project_id: str, *, review_status: str = "all", category: str = "", query: str = ""
    ) -> list[dict[str, Any]]:
        if review_status not in {"all", "unreviewed", "approved", "flagged"}:
            raise ValidationError("Invalid knowledge review status")
        clauses = ["k.project_id = ?", "s.memory_state = 'active'"]
        params: list[Any] = [project_id]
        if review_status != "all":
            clauses.append("k.review_status = ?")
            params.append(review_status)
        if category:
            clauses.append("k.category = ?")
            params.append(category)
        if query.strip():
            clauses.append("k.text LIKE ? ESCAPE '\\'")
            params.append(_like_pattern(query.strip()))
        with self.db.connect() as connection:
            self._project(connection, project_id)
            rows = connection.execute(
                f"""SELECT k.*, s.source_title, s.source_type, s.original_filename, s.ingestion_id
                    FROM knowledge_items k JOIN sources s ON s.id = k.source_id
                    WHERE {' AND '.join(clauses)} ORDER BY k.created_at DESC, k.id DESC""",
                params,
            ).fetchall()
            result = []
            for row in rows:
                item = dict(row)
                ids = _decode(item.pop("citation_ids_json"), [])
                citation_rows = []
                for citation_id in ids:
                    citation = connection.execute(
                        "SELECT * FROM citation_records WHERE id = ?", (citation_id,)
                    ).fetchone()
                    if citation:
                        citation_item = dict(citation)
                        citation_item["citation_id"] = citation_item.pop("id")
                        citation_rows.append(citation_item)
                item["citations"] = citation_rows
                result.append(item)
        return result

    def review_knowledge(self, project_id: str, knowledge_id: str, status: str) -> dict[str, Any]:
        if status not in {"unreviewed", "approved", "flagged"}:
            raise ValidationError("Invalid knowledge review status")
        with self.db.transaction() as connection:
            row = connection.execute(
                """SELECT k.* FROM knowledge_items k JOIN sources s ON s.id = k.source_id
                   WHERE k.id = ? AND k.project_id = ? AND s.memory_state = 'active'""",
                (knowledge_id, project_id),
            ).fetchone()
            if not row:
                raise NotFoundError("Knowledge item not found")
            connection.execute("UPDATE knowledge_items SET review_status = ? WHERE id = ?", (status, knowledge_id))
            source_id = int(row["source_id"])
            connection.execute(
                """UPDATE projects SET summary_revision = summary_revision + 1,
                   summary_generation_state = 'stale' WHERE id = ?""",
                (project_id,),
            )
        self._refresh_source_archive(source_id)
        self._write_knowledge_history(project_id)
        try:
            self.regenerate_living_summary(project_id, advance_revision=False)
        except UnexpectedSummaryError:
            # The knowledge review committed; its summary remains visibly failed and retryable.
            pass
        return next(item for item in self.list_knowledge(project_id) if item["id"] == knowledge_id)

    def review_summary(self, project_id: str, status: str) -> dict[str, Any]:
        if status not in {"unreviewed", "approved", "flagged"}:
            raise ValidationError("Invalid summary review status")
        with self.db.transaction() as connection:
            project = self._project(connection, project_id)
            version = connection.execute(
                "SELECT * FROM summary_versions WHERE project_id = ? AND revision = ?",
                (project_id, project["summary_revision"]),
            ).fetchone()
            if not version:
                raise ConflictError("No current living summary version is available to review")
            connection.execute(
                "UPDATE summary_versions SET review_status = ? WHERE id = ?", (status, version["id"])
            )
            connection.execute(
                "UPDATE projects SET summary_review_status = ? WHERE id = ?", (status, project_id)
            )
        current_path = Path(project["folder_path"]) / "_Assistant" / "living-summary" / "current.json"
        fallback_payload = {
            "schema_version": SCHEMA_VERSION, "project_id": project["archive_id"],
            "revision": int(version["revision"]), "generation_state": version["generation_state"],
            "created_at": version["created_at"],
            "sections": _decode(version["content_json"], {"sections": []}).get("sections", []),
        }
        payload = read_json(current_path, {}) if current_path.exists() else {}
        if not isinstance(payload, dict) or not isinstance(payload.get("sections"), list):
            payload = dict(fallback_payload)
        payload["review_status"] = status
        atomic_write_json(current_path, payload)
        versions_folder = current_path.parent / "versions"
        for version_path in versions_folder.glob(f"{int(version['revision']):06d}__*.json"):
            version_payload = read_json(version_path, {})
            if not isinstance(version_payload, dict) or not isinstance(version_payload.get("sections"), list):
                version_payload = dict(fallback_payload)
            version_payload["review_status"] = status
            atomic_write_json(version_path, version_payload)
        return self.get_living_summary(project_id)

    def refresh_source_derivatives(self, source_id: int) -> dict[str, Any]:
        """Recreate replaceable extracted text and sidecars without modifying originals."""
        with self.db.connect() as connection:
            source = connection.execute("SELECT * FROM sources WHERE id = ?", (source_id,)).fetchone()
            if not source:
                raise NotFoundError("Source not found")
            while source["parent_source_id"]:
                source = connection.execute(
                    "SELECT * FROM sources WHERE id = ?", (source["parent_source_id"],)
                ).fetchone()
            originals = connection.execute(
                "SELECT * FROM original_files WHERE source_id = ? ORDER BY is_attachment, id", (source["id"],)
            ).fetchall()
        if not source["canonical_source"]:
            self._refresh_source_archive(int(source["id"]))
            return {"source_id": int(source["id"]), "files_refreshed": 0, "unsupported_files": 0}
        if not originals:
            raise ValidationError("No preserved originals are available to rebuild derived files")
        package = self._under_root(Path(source["ingestion_path"]))
        extracted_dir = self._under_root(package / "Assistant" / "Extracted")
        replacement = self._under_root(package / "Assistant" / f"_DERIVED_{uuid.uuid4().hex}")
        replacement.mkdir(parents=True, exist_ok=False)
        generated: list[Path] = []
        unsupported = 0
        try:
            for index, original in enumerate(originals, start=1):
                path = self._under_root(package / original["relative_path"])
                if not path.is_file():
                    raise ValidationError(
                        f"Preserved original is unavailable locally: {original['original_name']}"
                    )
                if sha256_file(path) != original["sha256"]:
                    raise ConflictError(
                        f"Preserved original failed its SHA-256 integrity check: {original['original_name']}"
                    )
                try:
                    result = extract_source(
                        path,
                        max_attachments=self.settings.app.max_attachments,
                        max_text_bytes=self.settings.app.max_extracted_text_mb * 1024 * 1024,
                    )
                except UnsupportedSource:
                    unsupported += 1
                    continue
                output = replacement / (
                    f"{index:03d}-{safe_filename(Path(original['relative_path']).stem, 'source')}.txt"
                )
                atomic_write_text(output, "\n\n".join(chunk["text"] for chunk in result.chunks) + "\n")
                generated.append(output)
            extracted_dir.mkdir(parents=True, exist_ok=True)
            expected_names = {path.name for path in generated}
            for existing in extracted_dir.glob("*.txt"):
                if existing.name not in expected_names:
                    existing.unlink()
            for generated_path in generated:
                os.replace(generated_path, extracted_dir / generated_path.name)
        finally:
            if replacement.exists():
                shutil.rmtree(replacement)
        with self.db.transaction() as connection:
            connection.execute(
                "UPDATE sources SET processing_version = processing_version + 1 WHERE id = ?",
                (source["id"],),
            )
        self._refresh_source_archive(int(source["id"]))
        if source["project_id"]:
            self._write_knowledge_history(str(source["project_id"]))
        return {
            "source_id": int(source["id"]), "files_refreshed": len(generated),
            "unsupported_files": unsupported,
        }

    def source_detail(self, source_id: int) -> dict[str, Any]:
        with self.db.connect() as connection:
            source = connection.execute("SELECT * FROM sources WHERE id = ?", (source_id,)).fetchone()
            if not source:
                raise NotFoundError("Source not found")
            originals = [dict(row) for row in connection.execute(
                "SELECT * FROM original_files WHERE source_id = ? ORDER BY is_attachment, id", (source_id,)
            ).fetchall()]
            lifecycle = [dict(row) for row in connection.execute(
                "SELECT * FROM source_lifecycle_events WHERE source_id = ? ORDER BY created_at DESC, id DESC",
                (source_id,),
            ).fetchall()]
        item = self._decode_source(source)
        item["original_files"] = originals
        for event in lifecycle:
            event["details"] = _decode(event.pop("details_json"), {})
        item["lifecycle"] = lifecycle
        manifest = Path(source["ingestion_path"] or "") / "manifest.json"
        item["manifest"] = read_json(manifest) if manifest.is_file() else None
        return item

    @staticmethod
    def _path_after_package_move(value: str | None, old_package: Path, new_package: Path) -> str | None:
        if not value:
            return value
        candidate = Path(value)
        try:
            relative = candidate.resolve().relative_to(old_package.resolve())
        except ValueError:
            return value
        return str(new_package / relative)

    def _move_source_package(
        self,
        source_id: int,
        destination: Path,
        *,
        memory_state: str,
        event_type: str,
        reason: str,
        to_project_id: str | None = None,
        project_fit_confirmed: bool | None = None,
    ) -> dict[str, Any]:
        with self.db.connect() as connection:
            source = connection.execute("SELECT * FROM sources WHERE id = ?", (source_id,)).fetchone()
            if not source:
                raise NotFoundError("Source not found")
            if source["parent_source_id"] is not None and source["source_type"] != "routed_segment":
                raise ValidationError("Use the top-level source package for memory changes")
            if source["processing_state"] == "processing":
                raise ConflictError("Wait for source extraction to finish before moving or archiving it")
            target_project_id = to_project_id if to_project_id is not None else source["project_id"]
            if to_project_id is not None and to_project_id != source["project_id"]:
                duplicate = self._find_existing_source(
                    connection,
                    project_id=to_project_id,
                    digest=source["sha256"],
                    native_id=source["native_id"],
                    multi_project=False,
                )
                if duplicate and int(duplicate["id"]) != source_id:
                    raise ConflictError("This source is already preserved in the selected project")
            target_project = self._project(connection, target_project_id)
            source_project = self._project(connection, source["project_id"])
            if source["source_type"] == "routed_segment":
                descendants = connection.execute(
                    "SELECT * FROM sources WHERE project_id = ? AND ingestion_path = ?",
                    (source["project_id"], source["ingestion_path"]),
                ).fetchall()
            else:
                descendants = connection.execute(
                    """WITH RECURSIVE source_tree(id) AS (
                         SELECT ? UNION ALL
                         SELECT s.id FROM sources s JOIN source_tree t ON s.parent_source_id = t.id
                         WHERE s.project_id = ? AND s.ingestion_path = ?
                       ) SELECT * FROM sources WHERE id IN (SELECT id FROM source_tree)""",
                    (source_id, source["project_id"], source["ingestion_path"]),
                ).fetchall()
        if not source["ingestion_path"]:
            raise ConflictError("This legacy source has no movable OneDrive package")
        old_package = self._under_root(Path(source["ingestion_path"]))
        destination = self._under_root(destination)
        if old_package == destination:
            raise ConflictError("Source package is already in that location")
        if not old_package.is_dir():
            raise ConflictError("The preserved source package is missing from OneDrive")
        if destination.exists():
            raise ConflictError("A source package with this archive identity already exists at the destination")
        old_manifest_path = old_package / "manifest.json"
        old_lifecycle_path = old_package / "Assistant" / "source-lifecycle.jsonl"
        manifest_before = old_manifest_path.read_text(encoding="utf-8") if old_manifest_path.is_file() else "{}\n"
        lifecycle_before = old_lifecycle_path.read_text(encoding="utf-8") if old_lifecycle_path.is_file() else ""
        summary_folder = Path(source_project["folder_path"]) / "_Assistant" / "living-summary"
        summary_md_path = summary_folder / "current.md"
        summary_json_path = summary_folder / "current.json"
        summary_md_existed = summary_md_path.is_file()
        summary_json_existed = summary_json_path.is_file()
        summary_md_before = summary_md_path.read_text(encoding="utf-8") if summary_md_existed else ""
        summary_json_before = summary_json_path.read_text(encoding="utf-8") if summary_json_existed else ""
        destination.parent.mkdir(parents=True, exist_ok=True)
        os.replace(old_package, destination)
        now = utc_now()
        event_id = stable_id("E")
        event_details = {
            "from_path": relative_to_root(old_package, self.settings.app.one_drive_root),
            "to_path": relative_to_root(destination, self.settings.app.one_drive_root),
            "memory_state": memory_state,
        }
        archived_previous_state = source["memory_state"] if memory_state == "removed" else None
        changes_active_summary = (
            (event_type == "removed_from_memory" and source["memory_state"] == "active")
            or (event_type == "restored_to_memory" and memory_state == "active")
        )
        manifest_path = destination / "manifest.json"
        lifecycle_path = destination / "Assistant" / "source-lifecycle.jsonl"
        durable_event = {
            "id": event_id,
            "source_id": int(source["id"]),
            "ingestion_id": source["ingestion_id"],
            "project_id": source["project_id"],
            "event_type": event_type,
            "from_project_id": source["project_id"],
            "to_project_id": target_project_id,
            "reason": normalize_text(reason)[:500],
            "details": event_details,
            "created_at": now,
        }
        try:
            update_manifest(
                destination,
                database_project_id=target_project_id,
                project_id=target_project["archive_id"],
                memory_state=memory_state,
                project_fit_confirmed=(
                    bool(source["project_fit_confirmed"])
                    if project_fit_confirmed is None else bool(project_fit_confirmed)
                ),
                archived_previous_memory_state=archived_previous_state,
            )
            atomic_write_text(
                lifecycle_path,
                lifecycle_before + ("" if not lifecycle_before or lifecycle_before.endswith("\n") else "\n")
                + _json(durable_event) + "\n",
            )
            if changes_active_summary:
                atomic_write_text(
                    summary_md_path,
                    f"# {source_project['name']} — Living Summary\n\nCurrent memory is awaiting regeneration.\n",
                )
                atomic_write_json(summary_json_path, {
                    "schema_version": SCHEMA_VERSION,
                    "project_id": source_project["archive_id"],
                    "revision": int(source_project["summary_revision"]) + 1,
                    "review_status": "unreviewed",
                    "generation_state": "stale",
                    "created_at": now,
                    "sections": [],
                })
            with self.db.transaction() as connection:
                for item in descendants:
                    updates: dict[str, Any] = {
                        "project_id": target_project_id,
                        "original_path": self._path_after_package_move(
                            item["original_path"], old_package, destination
                        ),
                        "ingestion_path": self._path_after_package_move(
                            item["ingestion_path"], old_package, destination
                        ),
                        "memory_state": memory_state,
                        "memory_state_changed_at": now,
                    }
                    if item["id"] == source_id:
                        updates["ingestion_path"] = str(destination)
                    if memory_state == "removed":
                        updates["archived_previous_memory_state"] = item["memory_state"]
                    else:
                        updates["archived_previous_memory_state"] = None
                    if project_fit_confirmed is not None:
                        updates["project_fit_confirmed"] = int(project_fit_confirmed)
                    assignments = ", ".join(f"{key} = ?" for key in updates)
                    connection.execute(
                        f"UPDATE sources SET {assignments} WHERE id = ?",
                        (*updates.values(), item["id"]),
                    )
                descendant_ids = [int(item["id"]) for item in descendants]
                placeholders = ",".join("?" for _ in descendant_ids)
                connection.execute(
                    f"UPDATE source_chunks SET project_id = ? WHERE source_id IN ({placeholders})",
                    (target_project_id, *descendant_ids),
                )
                if event_type == "removed_from_memory" and source["memory_state"] == "active":
                    latest = connection.execute(
                        """SELECT u.text FROM project_updates u LEFT JOIN sources s ON s.id = u.source_id
                           WHERE u.project_id = ? AND (u.source_id IS NULL OR s.memory_state = 'active')
                           ORDER BY u.created_at DESC, u.id DESC LIMIT 1""",
                        (source["project_id"],),
                    ).fetchone()
                    connection.execute(
                        """UPDATE projects SET latest_change = ?, updated_at = ?,
                           summary_revision = summary_revision + 1, summary_generation_state = 'stale',
                           summary_error = NULL, current_summary = '', summary_generated_at = NULL
                           WHERE id = ?""",
                        (latest["text"] if latest else None, now, source["project_id"]),
                    )
                elif event_type == "restored_to_memory" and memory_state == "active":
                    latest = connection.execute(
                        """SELECT u.text FROM project_updates u LEFT JOIN sources s ON s.id = u.source_id
                           WHERE u.project_id = ? AND (u.source_id IS NULL OR s.memory_state = 'active')
                           ORDER BY u.created_at DESC, u.id DESC LIMIT 1""",
                        (target_project_id,),
                    ).fetchone()
                    connection.execute(
                        """UPDATE projects SET summary_revision = summary_revision + 1,
                           summary_generation_state = 'stale', summary_error = NULL,
                           latest_change = ?, updated_at = ? WHERE id = ?""",
                        (latest["text"] if latest else None, now, target_project_id),
                    )
                self._record_source_lifecycle(
                    connection, source, event_type, reason,
                    from_project_id=source["project_id"], to_project_id=target_project_id,
                    details=event_details, event_id=event_id, created_at=now,
                )
        except Exception:
            # Each compensating write is best-effort so one locked OneDrive sidecar does not
            # prevent the other authoritative files or package location from being restored.
            for path, content in (
                (manifest_path, manifest_before),
                (lifecycle_path, lifecycle_before),
            ):
                try:
                    if destination.exists():
                        atomic_write_text(path, content)
                except OSError:
                    pass
            try:
                if summary_md_existed:
                    atomic_write_text(summary_md_path, summary_md_before)
                else:
                    summary_md_path.unlink(missing_ok=True)
            except OSError:
                pass
            try:
                if summary_json_existed:
                    atomic_write_text(summary_json_path, summary_json_before)
                else:
                    summary_json_path.unlink(missing_ok=True)
            except OSError:
                pass
            # Restore the directory last so any reverted sidecar bytes travel back with the package.
            if destination.exists() and not old_package.exists():
                old_package.parent.mkdir(parents=True, exist_ok=True)
                os.replace(destination, old_package)
            raise
        try:
            self._refresh_source_archive(source_id)
        except OSError:
            # The authoritative lifecycle/manifest state was already written before the database commit.
            pass
        return self.source_detail(source_id)

    def remove_source_from_memory(
        self, project_id: str, source_id: int, reason: str, *, exclude_review_id: int | None = None
    ) -> dict[str, Any]:
        clean_reason = normalize_text(reason) or "Removed from active project memory by the user."
        with self.db.connect() as connection:
            project = self._project(connection, project_id)
            source = connection.execute(
                "SELECT * FROM sources WHERE id = ? AND project_id = ?", (source_id, project_id)
            ).fetchone()
        if not source:
            raise NotFoundError("Source not found in this project")
        if source["parent_source_id"] is not None and source["source_type"] != "routed_segment":
            raise ValidationError("Remove the top-level source package instead")
        if source["memory_state"] == "removed":
            raise ConflictError("Source is already removed from active project memory")
        if not source["ingestion_path"]:
            raise ConflictError("This legacy source has no movable OneDrive package")
        destination = self.archive_paths["archive"] / project["archive_id"] / Path(source["ingestion_path"]).name
        previous_state = source["memory_state"]
        self._move_source_package(
            source_id, destination, memory_state="removed", event_type="removed_from_memory",
            reason=clean_reason,
        )
        now = utc_now()
        with self.db.transaction() as connection:
            connection.execute(
                """UPDATE review_items SET status = 'dismissed',
                   resolution_json = ?, resolved_at = ?
                   WHERE source_id = ? AND status = 'open' AND kind = 'project_fit'
                     AND (? IS NULL OR id <> ?)""",
                (
                    _json({"action": "remove", "reason": clean_reason}), now, source_id,
                    exclude_review_id, exclude_review_id,
                ),
            )
        self._write_knowledge_history(project_id)
        if previous_state == "active":
            try:
                self.regenerate_living_summary(project_id, advance_revision=False)
            except UnexpectedSummaryError:
                # The removal committed; its summary remains visibly failed and retryable.
                pass
        return self.source_detail(source_id)

    def restore_source_to_memory(
        self, project_id: str, source_id: int, *,
        reason: str = "Restored to active project memory by the user.",
    ) -> dict[str, Any]:
        with self.db.connect() as connection:
            project = self._project(connection, project_id)
            source = connection.execute(
                "SELECT * FROM sources WHERE id = ? AND project_id = ?", (source_id, project_id)
            ).fetchone()
        if not source:
            raise NotFoundError("Source not found in this project archive")
        if source["parent_source_id"] is not None and source["source_type"] != "routed_segment":
            raise ValidationError("Restore the top-level source package instead")
        if source["memory_state"] != "removed":
            raise ConflictError("Source is already available to the project")
        restored_state = source["archived_previous_memory_state"] or "pending"
        destination = Path(project["folder_path"]) / Path(source["ingestion_path"]).name
        result = self._move_source_package(
            source_id, destination, memory_state=restored_state,
            event_type="restored_to_memory", reason=reason,
            project_fit_confirmed=bool(source["project_fit_confirmed"]),
        )
        if restored_state == "active":
            self._write_knowledge_history(project_id)
            try:
                self.regenerate_living_summary(project_id, advance_revision=False)
            except UnexpectedSummaryError:
                # The restoration committed; its summary remains visibly failed and retryable.
                pass
        else:
            with self.db.transaction() as connection:
                connection.execute(
                    """UPDATE sources SET processing_state = 'captured', error_code = NULL,
                       error_message = NULL, retry_count = 0 WHERE id = ?""",
                    (source_id,),
                )
            try:
                update_manifest(Path(result["ingestion_path"]), processing_status="captured")
                self._refresh_source_archive(source_id, processing_status="captured")
            except OSError:
                # The package move and database reset already committed; derived sidecars are retryable.
                pass
            result = self.source_detail(source_id)
        return result

    def rebuild_project_knowledge(self, project_id: str) -> dict[str, Any]:
        with self.db.connect() as connection:
            self._project(connection, project_id)
            active_sources = connection.execute(
                """SELECT count(*) FROM sources
                   WHERE project_id = ? AND parent_source_id IS NULL AND memory_state = 'active'""",
                (project_id,),
            ).fetchone()[0]
            knowledge_items = connection.execute(
                """SELECT count(*) FROM knowledge_items k JOIN sources s ON s.id = k.source_id
                   WHERE k.project_id = ? AND s.memory_state = 'active'""",
                (project_id,),
            ).fetchone()[0]
        self._write_knowledge_history(project_id)
        summary = self.regenerate_living_summary(project_id, advance_revision=True)
        return {
            "project_id": project_id, "active_sources": active_sources,
            "knowledge_items": knowledge_items, "living_summary": summary,
        }

    def search_archive(self, query: str, *, project_id: str | None = None, limit: int = 50) -> list[dict[str, Any]]:
        clean = normalize_text(query)
        if not clean:
            return []
        pattern = _like_pattern(clean)
        limit = max(1, min(limit, 100))
        project_clause = " AND p.id = ?" if project_id else ""
        results: list[dict[str, Any]] = []
        with self.db.connect() as connection:
            for row in connection.execute(
                f"""SELECT p.id AS project_id, p.name AS project_name, p.name AS title,
                           'project' AS result_type, p.name AS excerpt, NULL AS source_id,
                           NULL AS source_type, p.updated_at AS source_date
                    FROM projects p WHERE p.name LIKE ? ESCAPE '\\'{project_clause}
                    LIMIT ?""",
                ((pattern, project_id, limit) if project_id else (pattern, limit)),
            ).fetchall():
                results.append(dict(row))
            summary_rows = connection.execute(
                """SELECT p.id AS project_id, p.name AS project_name, v.content_json, v.created_at
                   FROM projects p JOIN summary_versions v ON v.project_id = p.id
                   WHERE v.revision = p.summary_revision
                     AND (? IS NULL OR p.id = ?)""",
                (project_id, project_id),
            ).fetchall()
            for row in summary_rows:
                for section in _decode(row["content_json"], {"sections": []}).get("sections", []):
                    section_text = str(section.get("text", ""))
                    section_name = str(section.get("section", "Living Summary"))
                    if clean.casefold() in f"{section_name} {section_text}".casefold():
                        results.append({
                            "project_id": row["project_id"], "project_name": row["project_name"],
                            "title": f"{row['project_name']} — {section_name}",
                            "result_type": "living_summary_claim", "excerpt": section_text,
                            "source_id": None, "source_type": None, "source_date": row["created_at"],
                        })
            for row in connection.execute(
                f"""SELECT p.id AS project_id, p.name AS project_name,
                           coalesce(s.source_title, s.original_filename) AS title,
                           'source_ingestion' AS result_type, s.source_summary AS excerpt,
                           s.id AS source_id, s.source_type, coalesce(s.source_date, s.created_at) AS source_date
                    FROM sources s LEFT JOIN projects p ON p.id = s.project_id
                    WHERE (s.source_title LIKE ? ESCAPE '\\' OR s.original_filename LIKE ? ESCAPE '\\' OR s.source_summary LIKE ? ESCAPE '\\'
                           OR s.metadata_json LIKE ? ESCAPE '\\') AND s.memory_state <> 'removed'{project_clause}
                    ORDER BY s.created_at DESC LIMIT ?""",
                ((pattern, pattern, pattern, pattern, project_id, limit) if project_id
                 else (pattern, pattern, pattern, pattern, limit)),
            ).fetchall():
                results.append(dict(row))
            for row in connection.execute(
                f"""SELECT p.id AS project_id, p.name AS project_name, s.source_title AS title,
                           'knowledge_item' AS result_type, k.text AS excerpt, s.id AS source_id,
                           s.source_type, k.source_date
                    FROM knowledge_items k JOIN projects p ON p.id = k.project_id
                    JOIN sources s ON s.id = k.source_id
                    WHERE (k.text LIKE ? ESCAPE '\\' OR k.category LIKE ? ESCAPE '\\') AND s.memory_state = 'active'{project_clause}
                    ORDER BY k.created_at DESC LIMIT ?""",
                ((pattern, pattern, project_id, limit) if project_id else (pattern, pattern, limit)),
            ).fetchall():
                results.append(dict(row))
            for row in connection.execute(
                f"""SELECT p.id AS project_id, p.name AS project_name,
                           coalesce(s.source_title, s.original_filename) AS title,
                           'original_file' AS result_type, f.original_name AS excerpt,
                           s.id AS source_id, s.source_type, coalesce(s.source_date, s.created_at) AS source_date,
                           f.id AS original_file_id
                    FROM original_files f JOIN sources s ON s.id = f.source_id
                    LEFT JOIN projects p ON p.id = s.project_id
                    WHERE f.original_name LIKE ? ESCAPE '\\' AND s.memory_state <> 'removed'{project_clause}
                    ORDER BY f.created_at DESC LIMIT ?""",
                ((pattern, project_id, limit) if project_id else (pattern, limit)),
            ).fetchall():
                results.append(dict(row))
            if len(results) < limit:
                for row in connection.execute(
                    f"""SELECT p.id AS project_id, p.name AS project_name,
                               coalesce(s.source_title, s.original_filename) AS title,
                               'source_excerpt' AS result_type, c.text AS excerpt, s.id AS source_id,
                               s.source_type, coalesce(s.source_date, s.created_at) AS source_date,
                               c.locator
                        FROM source_chunks c JOIN sources s ON s.id = c.source_id
                        LEFT JOIN projects p ON p.id = c.project_id
                        WHERE c.text LIKE ? ESCAPE '\\' AND s.memory_state <> 'removed'{project_clause}
                        ORDER BY c.id DESC LIMIT ?""",
                    ((pattern, project_id, limit) if project_id else (pattern, limit)),
                ).fetchall():
                    results.append(dict(row))
        return results[:limit]

    def rebuild_index(self) -> dict[str, int]:
        """Rescan durable OneDrive sidecars into an empty or partially rebuilt local index."""
        counts = {"projects": 0, "sources": 0, "knowledge_items": 0, "summary_versions": 0, "errors": 0}
        project_map: dict[str, str] = {}
        for descriptor_path in self.archive_paths["projects"].glob("*/_Assistant/project.json"):
            try:
                descriptor = json.loads(descriptor_path.read_text(encoding="utf-8"))
                project_id = str(descriptor["project_id"])
                archive_id = str(descriptor["archive_id"])
                project_map[archive_id] = project_id
                with self.db.transaction() as connection:
                    group = connection.execute(
                        "SELECT id FROM portfolio_groups WHERE name = 'Unassigned' COLLATE NOCASE"
                    ).fetchone()
                    exists = connection.execute("SELECT 1 FROM projects WHERE id = ?", (project_id,)).fetchone()
                    connection.execute(
                        """INSERT OR IGNORE INTO projects(
                           id, archive_id, name, portfolio_group_id, folder_path, created_at, updated_at
                           ) VALUES (?, ?, ?, ?, ?, ?, ?)""",
                        (project_id, archive_id, descriptor.get("name") or descriptor_path.parents[1].name,
                         group["id"], str(descriptor_path.parents[1]), descriptor.get("created_at") or utc_now(),
                         descriptor.get("created_at") or utc_now()),
                    )
                if not exists:
                    counts["projects"] += 1
            except (OSError, ValueError, KeyError, sqlite3.Error):
                counts["errors"] += 1
        package_paths = list(self.archive_paths["projects"].glob("*/*/manifest.json"))
        package_paths += list(self.archive_paths["shared_intake"].glob("*/manifest.json"))
        package_paths += list(self.archive_paths["archive"].glob("*/*/manifest.json"))
        for manifest_path in package_paths:
            try:
                manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
                if not isinstance(manifest, dict):
                    raise ValueError("Archive manifest must contain a JSON object")
                package = manifest_path.parent
                index_file = package / "Assistant" / "index.json"
                index = json.loads(index_file.read_text(encoding="utf-8")) if index_file.is_file() else {}
                if not isinstance(index, dict):
                    raise ValueError("Archive index must contain a JSON object")
                email_metadata = index.get("email_metadata")
                if email_metadata is None:
                    email_metadata = {}
                if not isinstance(email_metadata, dict):
                    raise ValueError("Archive email metadata must contain a JSON object")
                rebuilt_metadata = {**index, **email_metadata}
                canonical = bool(manifest.get("canonical_source", True))
                ingestion_id = str(manifest.get("ingestion_id") or stable_id("I"))
                manifest_project_id = manifest.get("project_id")
                database_project_id = manifest.get("database_project_id")
                with self.db.connect() as connection:
                    candidate_project = (
                        connection.execute(
                            "SELECT archive_id FROM projects WHERE id = ?", (database_project_id,)
                        ).fetchone()
                        if database_project_id else None
                    )
                    if (
                        candidate_project and manifest_project_id
                        and candidate_project["archive_id"] != manifest_project_id
                    ):
                        candidate_project = None
                    if not candidate_project:
                        database_project_id = project_map.get(str(manifest_project_id))
                    if (manifest_project_id or manifest.get("database_project_id")) and not database_project_id:
                        counts["errors"] += 1
                        continue
                    existing = connection.execute(
                        "SELECT * FROM sources WHERE ingestion_path = ?", (str(package),)
                    ).fetchone()
                if existing:
                    source_id = int(existing["id"])
                    memory_state = self._manifest_memory_state(manifest)
                    if existing["memory_state"] == "active" and memory_state == "pending":
                        # A best-effort post-commit manifest refresh may be stale; never demote
                        # committed memory without an explicit removed lifecycle state.
                        memory_state = "active"
                    archived_previous_state = manifest.get("archived_previous_memory_state")
                    if archived_previous_state not in {"pending", "active"}:
                        archived_previous_state = "active" if memory_state == "removed" else None
                    with self.db.transaction() as connection:
                        connection.execute(
                            """UPDATE sources SET memory_state = ?, project_fit_confirmed = ?,
                               archived_previous_memory_state = ?,
                               memory_state_changed_at = coalesce(memory_state_changed_at, created_at)
                               WHERE id = ?""",
                            (
                                memory_state,
                                int(bool(manifest.get("project_fit_confirmed", memory_state == "active"))),
                                archived_previous_state, source_id,
                            ),
                        )
                else:
                    originals = manifest.get("original_files") if isinstance(manifest.get("original_files"), list) else []
                    primary_item = originals[0] if originals else None
                    primary = package / primary_item["relative_path"] if primary_item else package
                    if not canonical and manifest.get("canonical_source_path"):
                        canonical_package = self.settings.app.one_drive_root / str(manifest["canonical_source_path"])
                        canonical_manifest = json.loads((canonical_package / "manifest.json").read_text(encoding="utf-8"))
                        canonical_originals = canonical_manifest.get("original_files", [])
                        if canonical_originals:
                            primary = canonical_package / canonical_originals[0]["relative_path"]
                    digest = (
                        str(primary_item.get("sha256")) if primary_item and len(originals) == 1
                        else hashlib.sha256(_json([
                            (item.get("relative_path"), item.get("sha256")) for item in originals
                        ]).encode()).hexdigest()
                    )
                    state = str(manifest.get("processing_status") or "complete")
                    if state not in {"captured", "processing", "pending_ai", "complete", "needs_review", "unsupported", "error"}:
                        state = "complete"
                    memory_state = self._manifest_memory_state(manifest)
                    archived_previous_state = manifest.get("archived_previous_memory_state")
                    if archived_previous_state not in {"pending", "active"}:
                        archived_previous_state = "active" if memory_state == "removed" else None
                    with self.db.transaction() as connection:
                        cursor = connection.execute(
                            """INSERT INTO sources(
                               project_id, source_type, sha256, original_filename, original_path,
                               metadata_json, processing_state, created_at, processed_at, ingestion_id,
                               ingestion_path, source_title, source_date, capture_method, canonical_source,
                               linked_ingestion_id, source_summary, memory_state, project_fit_confirmed,
                               archived_previous_memory_state, memory_state_changed_at
                               ) VALUES (?, ?, ?, ?, ?, '{}', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                            (database_project_id, manifest.get("source_type") or "unknown", digest,
                             primary_item.get("original_name") if primary_item else manifest.get("title") or "Source unavailable",
                             str(primary), state, manifest.get("created_at") or utc_now(),
                             manifest.get("created_at") if state == "complete" else None,
                             ingestion_id if canonical else None, str(package), manifest.get("title"),
                             manifest.get("source_date"), manifest.get("capture_method") or "archive_rescan",
                             int(canonical), manifest.get("linked_ingestion_id"), "", memory_state,
                             int(bool(manifest.get("project_fit_confirmed", memory_state == "active"))),
                             archived_previous_state,
                             manifest.get("created_at") or utc_now()),
                        )
                        source_id = int(cursor.lastrowid)
                        for item in originals:
                            connection.execute(
                                """INSERT INTO original_files(
                                   source_id, relative_path, original_name, stored_name, size_bytes, sha256,
                                   is_attachment, created_at
                                   ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                                (source_id, item["relative_path"], item.get("original_name") or Path(item["relative_path"]).name,
                                 item.get("stored_name") or Path(item["relative_path"]).name,
                                 int(item.get("size_bytes") or 0), item.get("sha256") or "",
                                 int(bool(item.get("is_attachment"))), manifest.get("created_at") or utc_now()),
                            )
                    counts["sources"] += 1
                with self.db.transaction() as connection:
                    existing_chunks = connection.execute(
                        "SELECT count(*) FROM source_chunks WHERE source_id = ?", (source_id,)
                    ).fetchone()[0]
                    if not existing_chunks:
                        sequence = 0
                        for extracted in sorted((package / "Assistant" / "Extracted").glob("*.txt")):
                            text = extracted.read_text(encoding="utf-8")
                            for offset in range(0, len(text), 3000):
                                excerpt = text[offset:offset + 3000].strip()
                                if excerpt:
                                    connection.execute(
                                        """INSERT INTO source_chunks(
                                           source_id, project_id, sequence, text, locator, processing_state, processed_at
                                           ) VALUES (?, ?, ?, ?, ?, 'complete', ?)""",
                                        (source_id, database_project_id, sequence, excerpt,
                                         f"{extracted.name} characters {offset + 1}-{offset + len(excerpt)}", utc_now()),
                                    )
                                    sequence += 1
                    source_summary_file = package / "Assistant" / "source-summary.md"
                    connection.execute(
                        "UPDATE sources SET source_summary = ?, metadata_json = ? WHERE id = ?",
                        (index.get("summary") or (source_summary_file.read_text(encoding="utf-8") if source_summary_file.is_file() else ""),
                          _json(rebuilt_metadata), source_id),
                    )
                    lifecycle_file = package / "Assistant" / "source-lifecycle.jsonl"
                    if lifecycle_file.is_file():
                        for line in lifecycle_file.read_text(encoding="utf-8").splitlines():
                            event = json.loads(line)
                            connection.execute(
                                """INSERT OR IGNORE INTO source_lifecycle_events(
                                   id, source_id, ingestion_id, project_id, event_type,
                                   from_project_id, to_project_id, reason, details_json, created_at
                                   ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                                (
                                    event["id"], source_id, event.get("ingestion_id"),
                                    event.get("project_id") or database_project_id, event["event_type"],
                                    event.get("from_project_id"), event.get("to_project_id"),
                                    event.get("reason") or "Rebuilt from archive lifecycle history.",
                                    _json(event.get("details") or {}), event.get("created_at") or utc_now(),
                                ),
                            )
                counts["knowledge_items"] += self._rebuild_package_knowledge(
                    source_id, database_project_id, package
                )
            except (OSError, ValueError, KeyError, sqlite3.Error):
                counts["errors"] += 1
        for project_id in project_map.values():
            try:
                counts["summary_versions"] += self._rebuild_project_summaries(project_id)
                self._write_knowledge_history(project_id)
            except (OSError, ValueError, KeyError, sqlite3.Error):
                counts["errors"] += 1
        return counts

    def _rebuild_package_knowledge(self, source_id: int, project_id: str | None, package: Path) -> int:
        if not project_id:
            return 0
        citations_path = package / "Assistant" / "citations.json"
        knowledge_path = package / "Assistant" / "knowledge-items.json"
        citations = json.loads(citations_path.read_text(encoding="utf-8")) if citations_path.is_file() else []
        knowledge = json.loads(knowledge_path.read_text(encoding="utf-8")) if knowledge_path.is_file() else []
        count = 0
        with self.db.transaction() as connection:
            citation_map: dict[str, str] = {}
            used_chunk_ids = {
                int(row["chunk_id"])
                for row in connection.execute(
                    "SELECT chunk_id FROM citation_records WHERE source_id = ?", (source_id,)
                ).fetchall()
            }
            for citation in citations:
                citation_id = str(citation["citation_id"])
                existing = connection.execute(
                    "SELECT id, source_id, chunk_id FROM citation_records WHERE id = ?", (citation_id,)
                ).fetchone()
                if existing:
                    if int(existing["source_id"]) == source_id:
                        citation_map[citation_id] = citation_id
                        used_chunk_ids.add(int(existing["chunk_id"]))
                        continue
                    raise ValueError("Citation ID is already associated with a different source")
                citation_excerpt = str(citation.get("excerpt") or "")[:120]
                chunk = None
                if citation_excerpt:
                    candidates = connection.execute(
                        """SELECT * FROM source_chunks
                           WHERE source_id = ? AND instr(text, ?) > 0 ORDER BY id""",
                        (source_id, citation_excerpt),
                    ).fetchall()
                    chunk = next(
                        (candidate for candidate in candidates if int(candidate["id"]) not in used_chunk_ids),
                        None,
                    )
                if not chunk:
                    max_sequence = connection.execute(
                        "SELECT COALESCE(MAX(sequence), -1) FROM source_chunks WHERE source_id = ?",
                        (source_id,),
                    ).fetchone()[0]
                    cursor = connection.execute(
                        """INSERT INTO source_chunks(
                           source_id, project_id, sequence, text, locator, processing_state, processed_at
                           ) VALUES (?, ?, ?, ?, ?, 'complete', ?)""",
                        (source_id, project_id, max(1_000_000, int(max_sequence) + 1),
                         citation.get("excerpt") or "Citation excerpt unavailable",
                         citation.get("locator") or "Archive citation", utc_now()),
                    )
                    chunk_id = int(cursor.lastrowid)
                else:
                    chunk_id = int(chunk["id"])
                connection.execute(
                    """INSERT OR IGNORE INTO citation_records(
                       id, source_id, chunk_id, original_relative_path, display_name, source_type,
                       locator, excerpt, source_date, created_at
                       ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (citation_id, source_id, chunk_id, citation.get("original_relative_path") or "",
                     citation.get("display_name") or "Source", citation.get("source_type") or "unknown",
                     citation.get("locator") or "Archive citation", citation.get("excerpt") or "",
                     citation.get("source_date"), utc_now()),
                )
                # INSERT OR IGNORE may have been a no-op; confirm ownership before mapping the ID.
                rebuilt = connection.execute(
                    "SELECT id, source_id, chunk_id FROM citation_records WHERE id = ?", (citation_id,)
                ).fetchone()
                if rebuilt and int(rebuilt["source_id"]) == source_id:
                    citation_map[citation_id] = citation_id
                    used_chunk_ids.add(int(rebuilt["chunk_id"]))
                else:
                    raise ValueError("Citation record could not be rebuilt for its owning source")
            for item in knowledge:
                if connection.execute("SELECT 1 FROM knowledge_items WHERE id = ?", (item["knowledge_item_id"],)).fetchone():
                    continue
                ids = [value for value in item.get("citation_ids", []) if value in citation_map]
                if not ids:
                    continue
                connection.execute(
                    """INSERT INTO knowledge_items(
                       id, project_id, source_id, text, category, source_date, citation_ids_json,
                       review_status, supersedes_knowledge_item_id, created_at
                       ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (item["knowledge_item_id"], project_id, source_id, item["text"],
                     item.get("category") or "development", item.get("source_date"), _json(ids),
                     item.get("review_status") if item.get("review_status") in {"unreviewed", "approved", "flagged"} else "unreviewed",
                     item.get("supersedes_knowledge_item_id"), utc_now()),
                )
                connection.execute(
                    """INSERT INTO project_updates(
                       project_id, source_id, update_type, text, citations_json, created_at
                       ) SELECT ?, ?, 'knowledge', ?, '[]', ? WHERE NOT EXISTS (
                         SELECT 1 FROM project_updates WHERE project_id = ? AND source_id = ? AND text = ?
                       )""",
                    (project_id, source_id, item["text"], utc_now(), project_id, source_id, item["text"]),
                )
                count += 1
        return count

    def _rebuild_project_summaries(self, project_id: str) -> int:
        with self.db.connect() as connection:
            project = self._project(connection, project_id)
        folder = Path(project["folder_path"]) / "_Assistant" / "living-summary"
        count = 0
        for version_path in sorted((folder / "versions").glob("*.json")):
            payload = json.loads(version_path.read_text(encoding="utf-8"))
            revision = int(payload.get("revision") or 0)
            with self.db.transaction() as connection:
                cursor = connection.execute(
                    """INSERT OR IGNORE INTO summary_versions(
                       project_id, revision, content_json, markdown, review_status,
                       generation_state, created_at
                       ) VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (project_id, revision, _json({"sections": payload.get("sections", [])}),
                     "\n\n".join(f"## {item.get('section')}\n\n{item.get('text')}" for item in payload.get("sections", [])),
                     payload.get("review_status") if payload.get("review_status") in {"unreviewed", "approved", "flagged"} else "unreviewed",
                     payload.get("generation_state") if payload.get("generation_state") in {"current", "updating", "stale", "failed"} else "stale",
                     payload.get("created_at") or utc_now()),
                )
                count += cursor.rowcount
        current_path = folder / "current.json"
        if current_path.is_file():
            current = json.loads(current_path.read_text(encoding="utf-8"))
            summary_text = " ".join(item.get("text", "") for item in current.get("sections", []))
            with self.db.transaction() as connection:
                connection.execute(
                    """UPDATE projects SET current_summary = ?, summary_revision = ?,
                       summary_generation_state = ?, summary_review_status = ?, summary_generated_at = ?
                       WHERE id = ?""",
                    (summary_text, int(current.get("revision") or 0), current.get("generation_state") or "current",
                     current.get("review_status") or "unreviewed", current.get("created_at"), project_id),
                )
        return count

    def _apply_ai_action(
        self, connection: sqlite3.Connection, project_id: str, source_id: int,
        operation: dict[str, Any], allowed: set[tuple[int, int]], now: str,
    ) -> None:
        if not isinstance(operation, dict):
            raise LlmContractError("Action operation must be an object")
        kind = operation.get("operation")
        citations = self._validate_citations(operation.get("citations"), allowed)
        if kind == "create":
            description = normalize_text(str(operation.get("description", "")))
            assignee_type = operation.get("assignee_type")
            assignee_value = normalize_text(str(operation.get("assignee_value", "")))
            due = operation.get("due_date")
            try:
                confidence = float(operation.get("confidence", 0))
            except (TypeError, ValueError) as exc:
                raise LlmContractError("Action confidence must be a number") from exc
            complete = description and assignee_type in ASSIGNEE_TYPES and assignee_value and due and confidence >= 0.9
            if due:
                try:
                    date.fromisoformat(str(due))
                except ValueError:
                    complete = False
            if not complete:
                self._insert_review(
                    connection, "inferred_action", source_id, project_id,
                    "Who owns this action and when is it due?",
                    "The source did not provide enough explicit owner/due-date evidence.",
                    [operation], [], "No routing rule will be created.", now,
                )
                return
            existing = connection.execute(
                "SELECT 1 FROM action_items WHERE project_id = ? AND source_id = ? AND description = ?",
                (project_id, source_id, description),
            ).fetchone()
            if not existing:
                connection.execute(
                    """
                    INSERT INTO action_items(project_id, description, assignee_type, assignee_value,
                      due_date, state, source_id, citations_json, created_by, created_at, updated_at)
                    VALUES (?, ?, ?, ?, ?, 'open', ?, ?, 'ai', ?, ?)
                    """,
                    (project_id, description, assignee_type, assignee_value, due, source_id,
                     _json(citations), now, now),
                )
        elif kind == "progress":
            action_id = operation.get("action_item_id")
            progress = normalize_text(str(operation.get("progress_text", "")))
            row = connection.execute(
                "SELECT * FROM action_items WHERE id = ? AND project_id = ?", (action_id, project_id)
            ).fetchone()
            if not row or not progress:
                self._insert_review(
                    connection, "inferred_action", source_id, project_id,
                    "Which action item should this progress update apply to?",
                    "The referenced action item or progress text is incomplete.",
                    [operation], [], "No routing rule will be created.", now,
                )
                return
            connection.execute(
                "UPDATE action_items SET progress_text = ?, citations_json = ?, updated_at = ? WHERE id = ?",
                (progress, _json(citations), now, action_id),
            )
        elif kind == "request_close":
            self._insert_review(
                connection, "action_close", source_id, project_id,
                "Do you approve completing this action item?",
                "AI is never allowed to close an action item without explicit user approval.",
                [operation], [{"action_item_id": operation.get("action_item_id"), "choice": "complete"}],
                "No routing rule will be created.", now,
            )
        else:
            raise LlmContractError("Unsupported action item operation")

    def _insert_review(
        self, connection: sqlite3.Connection, kind: str, source_id: int | None, project_id: str | None,
        question: str, reason: str, evidence: Any, options: Any, memory_preview: str, now: str,
    ) -> int:
        cursor = connection.execute(
            """
            INSERT INTO review_items(kind, source_id, project_id, status, question, reason,
              evidence_json, options_json, memory_preview, created_at)
            VALUES (?, ?, ?, 'open', ?, ?, ?, ?, ?, ?)
            """,
            (kind, source_id, project_id, question, reason, _json(evidence), _json(options), memory_preview, now),
        )
        return int(cursor.lastrowid)

    def _create_review(
        self, *, kind: str, source_id: int | None, project_id: str | None, question: str,
        reason: str, evidence: Any, options: Any, memory_preview: str,
    ) -> int:
        with self.db.transaction() as connection:
            return self._insert_review(
                connection, kind, source_id, project_id, question, reason,
                evidence, options, memory_preview, utc_now(),
            )

    def import_snow(self, stream: BinaryIO, filename: str) -> dict[str, Any]:
        clean_name = safe_filename(filename, "snow-export")
        suffix = Path(clean_name).suffix.casefold()
        if suffix not in {".csv", ".xlsx"}:
            raise ValidationError("SNOW import accepts only CSV or XLSX files")
        directory = self._under_root(
            self.settings.app.one_drive_root / "_PortfolioAssistant" / "imports" / "snow"
        )
        directory.mkdir(parents=True, exist_ok=True)
        temp_path = self._under_root(directory / f".{uuid.uuid4().hex}.tmp{suffix}")
        limit = self.settings.app.max_file_mb * 1024 * 1024
        total = 0
        try:
            with temp_path.open("xb") as output:
                while True:
                    block = stream.read(1024 * 1024)
                    if not block:
                        break
                    total += len(block)
                    if total > limit:
                        raise ValidationError(f"Import exceeds the configured {self.settings.app.max_file_mb} MB limit")
                    output.write(block)
            digest = sha256_file(temp_path)
            final_path = self._under_root(directory / f"{datetime.now():%Y%m%d-%H%M%S}-{digest[:10]}-{clean_name}")
            if final_path.exists():
                temp_path.unlink()
            else:
                os.replace(temp_path, final_path)
            rows, columns = self._read_snow_rows(final_path)
            missing = [column for column in REQUIRED_SNOW_COLUMNS if column not in columns]
            if missing:
                raise ValidationError(f"Missing required SNOW columns: {', '.join(missing)}")
            result: dict[str, Any] = {
                "tickets_read": len(rows), "new_comments_applied": 0, "tickets_unchanged": 0,
                "review_or_error_count": 0, "pending_ai": 0, "affected_projects": [],
                "review_item_ids": [],
            }
            for row_number, raw in enumerate(rows, start=2):
                try:
                    row_result = self._import_snow_row(raw, row_number, final_path, digest)
                    result["new_comments_applied"] += row_result["new_comments_applied"]
                    result["tickets_unchanged"] += int(row_result["unchanged"])
                    result["pending_ai"] += int(row_result["pending_ai"])
                    if row_result.get("project_id") and row_result["project_id"] not in result["affected_projects"]:
                        result["affected_projects"].append(row_result["project_id"])
                    if row_result.get("review_id"):
                        result["review_or_error_count"] += 1
                        result["review_item_ids"].append(row_result["review_id"])
                    elif row_result.get("error"):
                        result["review_or_error_count"] += 1
                except (ValidationError, ValueError) as exc:
                    review_id = self._create_review(
                        kind="snow_invalid_row", source_id=None, project_id=None,
                        question=f"How should SNOW import row {row_number} be handled?",
                        reason=_safe_error(exc), evidence=[{"row": row_number, "values": self._recognized_snow(raw)}],
                        options=[], memory_preview="No routing rule will be created.",
                    )
                    result["review_or_error_count"] += 1
                    result["review_item_ids"].append(review_id)
            return result
        finally:
            temp_path.unlink(missing_ok=True)

    def _read_snow_rows(self, path: Path) -> tuple[list[dict[str, Any]], list[str]]:
        if path.suffix.casefold() == ".csv":
            data = path.read_bytes()
            text = None
            for encoding in ("utf-8-sig", "cp1252"):
                try:
                    text = data.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if text is None:
                raise ValidationError("CSV encoding must be UTF-8 or Windows-1252")
            reader = csv.DictReader(io.StringIO(text), restkey="__extra_columns__", restval=None)
            columns = [str(column).strip() for column in (reader.fieldnames or [])]
            rows = []
            for raw in reader:
                row = {str(key).strip(): value for key, value in raw.items()}
                row["__ragged_row__"] = bool(row.get("__extra_columns__")) or any(
                    row.get(column) is None for column in columns
                )
                rows.append(row)
            return rows, columns
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            # Some ServiceNow-generated workbooks underreport their used range as A1 even
            # though the worksheet contains a full table. Read-only mode otherwise trusts
            # that metadata and silently returns only the first header cell.
            sheet.reset_dimensions()
            iterator = sheet.iter_rows(values_only=True)
            header = next(iterator, None)
            if not header:
                raise ValidationError("XLSX export is empty")
            header_width = max((index + 1 for index, value in enumerate(header) if value is not None), default=0)
            columns = [str(value).strip() if value is not None else "" for value in header[:header_width]]
            rows = []
            for values in iterator:
                if not any(value is not None for value in values):
                    continue
                row = dict(zip(columns, values[:header_width]))
                row["__ragged_row__"] = any(value is not None for value in values[header_width:])
                rows.append(row)
            return rows, columns
        finally:
            workbook.close()

    def _recognized_snow(self, row: dict[str, Any]) -> dict[str, Any]:
        result: dict[str, Any] = {}
        for column in RECOGNIZED_SNOW_COLUMNS:
            value = row.get(column)
            if isinstance(value, (datetime, date)):
                value = value.isoformat()
            result[column] = "" if value is None else str(value)
        return result

    def _parse_snow_datetime(self, value: Any, field: str) -> str:
        if isinstance(value, datetime):
            parsed = value
        elif isinstance(value, date):
            parsed = datetime.combine(value, time.min)
        else:
            text = normalize_text(str(value or ""))
            if not text:
                raise ValidationError(f"{field} is required")
            parsed = None
            for candidate in (text, text.replace("Z", "+00:00")):
                try:
                    parsed = datetime.fromisoformat(candidate)
                    break
                except ValueError:
                    continue
            if parsed is None:
                for fmt in ("%m/%d/%Y %H:%M:%S", "%m/%d/%Y %I:%M:%S %p", "%Y-%m-%d %H:%M:%S"):
                    try:
                        parsed = datetime.strptime(text, fmt)
                        break
                    except ValueError:
                        continue
            if parsed is None:
                raise ValidationError(f"{field} is not a recognized date/time")
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone.utc).isoformat(timespec="seconds")

    def _parse_snow_entries(self, cell: str) -> list[dict[str, str]]:
        matches = list(SNOW_ENTRY_HEADER.finditer(cell))
        if not matches or cell[:matches[0].start()].strip():
            raise ValidationError("Comments and Work notes could not be split at deterministic entry boundaries")
        entries = []
        for index, match in enumerate(matches):
            end = matches[index + 1].start() if index + 1 < len(matches) else len(cell)
            body = cell[match.end():end].strip()
            if not body:
                raise ValidationError("A SNOW comment entry has no body")
            stamp = datetime.strptime(match.group("stamp"), "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
            entries.append({
                "timestamp": stamp.isoformat(timespec="seconds"),
                "author": normalize_text(match.group("author")),
                "entry_type": normalize_text(match.group("kind")),
                "body": body,
            })
        return entries

    def _entry_hash(self, number: str, entry: dict[str, str]) -> str:
        material = "\x1f".join((
            number.casefold(), entry["timestamp"], entry["author"].casefold(),
            entry["entry_type"].casefold(), normalize_text(entry["body"]),
        ))
        return hashlib.sha256(material.encode("utf-8")).hexdigest()

    def _import_snow_row(
        self, raw: dict[str, Any], row_number: int, export_path: Path, export_sha: str
    ) -> dict[str, Any]:
        if raw.get("__ragged_row__"):
            raise ValidationError("SNOW row has a different number of values than the header")
        number = normalize_text(str(raw.get("Number") or ""))
        name = normalize_text(str(raw.get("Short description") or ""))
        assignment = normalize_text(str(raw.get("Assignment group") or ""))
        comments = str(raw.get("Comments and Work notes") or "")
        if not number or not name or not assignment or not comments.strip():
            raise ValidationError("Number, Short description, Assignment group, and Comments and Work notes are required")
        updated_at = self._parse_snow_datetime(raw.get("Updated"), "Updated")
        metadata = self._recognized_snow(raw)
        with self.db.connect() as connection:
            project = connection.execute("SELECT * FROM projects WHERE snow_number = ?", (number,)).fetchone()
        if project is None:
            created = self.create_project(
                name, snow_number=number, assignment_group=assignment, snow_metadata=metadata
            )
            project_id = created["id"]
            with self.db.connect() as connection:
                project = connection.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
        else:
            project_id = project["id"]
        cell_hash = hashlib.sha256(comments.encode("utf-8")).hexdigest()
        if project["snow_comments_cell_hash"] == cell_hash:
            return {"project_id": project_id, "new_comments_applied": 0, "unchanged": True, "pending_ai": False}
        if project["snow_updated_at"]:
            stored = datetime.fromisoformat(project["snow_updated_at"])
            incoming = datetime.fromisoformat(updated_at)
            if incoming < stored:
                return {"project_id": project_id, "new_comments_applied": 0, "unchanged": True, "pending_ai": False}
        try:
            entries = self._parse_snow_entries(comments)
        except ValidationError as exc:
            source_id = self._insert_snow_source(
                project_id, number, cell_hash, export_path, export_sha,
                metadata={"cell_hash": cell_hash, "row_number": row_number, "parse_failed": True},
                state="needs_review",
            )
            with self.db.transaction() as connection:
                connection.execute(
                    "INSERT INTO source_chunks(source_id, project_id, sequence, text, locator, processing_state) VALUES (?, ?, 0, ?, 'unparsed cumulative comments', 'needs_review')",
                    (source_id, project_id, comments),
                )
                review_id = self._insert_review(
                    connection, "snow_unparseable_comments", source_id, project_id,
                    "Where are the entry boundaries in this cumulative SNOW comment cell?",
                    str(exc), [{"source_id": source_id, "excerpt": comments[:1200]}], [],
                    "No routing rule will be created.", utc_now(),
                )
            return {"project_id": project_id, "new_comments_applied": 0, "unchanged": False, "pending_ai": False, "review_id": review_id}
        now = utc_now()
        with self.db.transaction() as connection:
            current = self._project(connection, project_id)
            next_name = current["name"] if current["name_manually_overridden"] else name
            connection.execute(
                """
                UPDATE projects SET name = ?, assignment_group = ?, snow_state = ?, snow_priority = ?,
                  snow_updated_at = ?, snow_metadata_json = ?, updated_at = ? WHERE id = ?
                """,
                (next_name, assignment, normalize_text(str(raw.get("State") or "")) or None,
                 normalize_text(str(raw.get("Priority") or "")) or None, updated_at,
                 _json(metadata), now, project_id),
            )
        prepared = []
        with self.db.connect() as connection:
            for entry in entries:
                entry_hash = self._entry_hash(number, entry)
                exists = connection.execute(
                    """SELECT 1 FROM source_chunks c JOIN sources s ON s.id = c.source_id
                       WHERE s.project_id = ? AND c.entry_hash = ? AND c.processing_state = 'complete'""",
                    (project_id, entry_hash),
                ).fetchone()
                if not exists:
                    prepared.append((entry_hash, entry))
        if not prepared:
            with self.db.transaction() as connection:
                connection.execute(
                    "UPDATE projects SET snow_comments_cell_hash = ? WHERE id = ?", (cell_hash, project_id)
                )
            return {"project_id": project_id, "new_comments_applied": 0, "unchanged": True, "pending_ai": False}
        with self.db.connect() as connection:
            existing_source = connection.execute(
                "SELECT id, processing_state FROM sources WHERE project_id = ? AND native_id = ?",
                (project_id, f"snow:{number}:{cell_hash}"),
            ).fetchone()
        if existing_source:
            state = self.process_source(int(existing_source["id"]))
            return {
                "project_id": project_id,
                "new_comments_applied": len(prepared) if state == "complete" else 0,
                "unchanged": False,
                "pending_ai": state == "pending_ai",
                "error": state == "error",
                "review_id": self._latest_source_review(int(existing_source["id"])) if state == "needs_review" else None,
            }
        source_id = self._insert_snow_source(
            project_id, number, cell_hash, export_path, export_sha,
            metadata={"cell_hash": cell_hash, "row_number": row_number}, state="captured",
        )
        prepared.sort(key=lambda pair: pair[1]["timestamp"])
        with self.db.transaction() as connection:
            for sequence, (entry_hash, entry) in enumerate(prepared):
                connection.execute(
                    """
                    INSERT INTO source_chunks(source_id, project_id, sequence, text, locator,
                      entry_hash, comment_at, author, entry_type, processing_state)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'captured')
                    """,
                    (source_id, project_id, sequence, entry["body"],
                     f"SNOW {number}, {entry['timestamp']} - {entry['author']} ({entry['entry_type']})",
                     entry_hash, entry["timestamp"], entry["author"], entry["entry_type"]),
                )
        state = self.process_source(source_id)
        return {
            "project_id": project_id,
            "new_comments_applied": len(prepared) if state == "complete" else 0,
            "unchanged": False,
            "pending_ai": state == "pending_ai",
            "error": state == "error",
            "review_id": self._latest_source_review(source_id) if state == "needs_review" else None,
        }

    def _insert_snow_source(
        self, project_id: str, number: str, cell_hash: str, export_path: Path, export_sha: str,
        *, metadata: dict[str, Any], state: str,
    ) -> int:
        with self.db.transaction() as connection:
            existing = connection.execute(
                "SELECT id FROM sources WHERE project_id = ? AND native_id = ?",
                (project_id, f"snow:{number}:{cell_hash}"),
            ).fetchone()
            if existing:
                return int(existing["id"])
            cursor = connection.execute(
                """
                INSERT INTO sources(project_id, source_type, native_id, sha256, original_filename,
                  original_path, metadata_json, processing_state, memory_state,
                  project_fit_confirmed, memory_state_changed_at, created_at)
                VALUES (?, 'snow_comments', ?, ?, ?, ?, ?, ?, 'pending', 1, ?, ?)
                """,
                (project_id, f"snow:{number}:{cell_hash}", export_sha, export_path.name,
                 str(export_path), _json(metadata), state, utc_now(), utc_now()),
            )
        return int(cursor.lastrowid)

    def _latest_source_review(self, source_id: int) -> int | None:
        with self.db.connect() as connection:
            row = connection.execute(
                "SELECT id FROM review_items WHERE source_id = ? ORDER BY id DESC LIMIT 1", (source_id,)
            ).fetchone()
        return int(row["id"]) if row else None

    def process_multi_source(self, source_id: int) -> str:
        with self.db.transaction() as connection:
            source = connection.execute("SELECT * FROM sources WHERE id = ?", (source_id,)).fetchone()
            if not source:
                raise NotFoundError("Source not found")
            if source["project_id"] is not None:
                raise ValidationError("This source already belongs to a project")
            if source["processing_state"] == "complete":
                return "complete"
            if source["processing_state"] == "processing":
                return "processing"
            claimed = connection.execute(
                """UPDATE sources SET processing_state = 'processing', error_code = NULL, error_message = NULL
                   WHERE id = ? AND processing_state <> 'processing'""",
                (source_id,),
            )
            if claimed.rowcount != 1:
                return "processing"
        try:
            self._ensure_extracted(source_id)
            evidence = self._source_evidence(source_id)
            if not evidence:
                raise UnsupportedSource("No supported text could be extracted")
            with self.db.connect() as connection:
                source = connection.execute("SELECT * FROM sources WHERE id = ?", (source_id,)).fetchone()
                projects = [dict(row) for row in connection.execute(
                    "SELECT id, name, snow_number FROM projects ORDER BY name"
                ).fetchall()]
            deterministic = self._match_routing_rule(source, evidence)
            if deterministic:
                routed_evidence = evidence
                dropped = 0
                segments = [{
                    "text": item["text"][:900], "project_id": deterministic["target_project_id"],
                    "confidence": 1.0, "reason": f"Matched local routing rule #{deterministic['id']}.",
                    "citations": [{"source_id": item["source_id"], "chunk_id": item["chunk_id"]}],
                    "matched_rule_id": deterministic["id"],
                } for item in evidence]
            else:
                routed_evidence, dropped = self._bounded_evidence(evidence)
                self._record_evidence_window(source_id, dropped)
                if not routed_evidence:
                    raise LlmContractError("No complete evidence chunk fits the configured evidence limit")
                routed = self.llm.route(routed_evidence, projects)
                segments = routed.get("segments") if isinstance(routed, dict) else None
                if not isinstance(segments, list) or not segments:
                    raise LlmContractError("Routing response did not contain segments")
            allowed = {(int(item["source_id"]), int(item["chunk_id"])) for item in routed_evidence}
            now = utc_now()
            with self.db.transaction() as connection:
                for segment in segments:
                    if not isinstance(segment, dict) or not normalize_text(str(segment.get("text", ""))):
                        raise LlmContractError("Routing segment is missing text")
                    citations = self._validate_citations(segment.get("citations"), allowed)
                    target = segment.get("project_id")
                    if target and not connection.execute("SELECT 1 FROM projects WHERE id = ?", (target,)).fetchone():
                        target = None
                    filename_phrase = normalize_text(Path(source["original_filename"]).stem).casefold()[:120]
                    suggested_rule = {"rule_type": "filename_phrase", "pattern": filename_phrase, "context": {}}
                    evidence_payload = [{
                        "text": normalize_text(segment["text"]), "citations": citations,
                        "confidence": float(segment.get("confidence", 0)), "reason": str(segment.get("reason", "")),
                        "proposed_project_id": target, "suggested_rule": suggested_rule,
                        "matched_rule_id": segment.get("matched_rule_id"),
                    }]
                    options = [{"project_id": project["id"], "label": project["name"]} for project in projects]
                    self._insert_review(
                        connection, "multi_project_route", source_id, target,
                        "Which project should receive this cited segment?",
                        str(segment.get("reason") or "Multi-project intake always requires confirmation."),
                        evidence_payload, options,
                        f"Remember filename phrase “{filename_phrase}” for the confirmed project.", now,
                    )
                connection.execute(
                    "UPDATE sources SET processing_state = 'needs_review', model_id = ?, processed_at = ? WHERE id = ?",
                    (self.llm.model_for("multi_project_routing"), now, source_id),
                )
            self._refresh_source_archive(source_id, processing_status="needs_review")
            return "needs_review"
        except UnsupportedSource as exc:
            self._set_source_state(source_id, "unsupported", "unsupported_type", exc)
            return "unsupported"
        except LlmUnavailable as exc:
            self._set_source_state(source_id, "pending_ai", "llm_unavailable", exc, increment_retry=True)
            return "pending_ai"
        except (ExtractionFailure, LlmContractError, ValidationError) as exc:
            self._set_source_state(source_id, "error", "multi_project_processing_failed", exc)
            return "error"
        except Exception as exc:  # noqa: BLE001 - must not strand the source in 'processing'
            # Anything else (OSError from a locked OneDrive path, sqlite3 errors, a
            # source removed mid-flight) would otherwise escape into the background
            # worker and leave this source permanently in 'processing'.
            self._set_source_state(source_id, "error", "multi_project_processing_failed", exc)
            return "error"

    def _match_routing_rule(self, source: sqlite3.Row, evidence: list[dict[str, Any]]) -> dict[str, Any] | None:
        combined = "\n".join(item["text"] for item in evidence).casefold()
        filename = source["original_filename"].casefold()
        metadata = _decode(source["metadata_json"], {})
        with self.db.connect() as connection:
            rules = connection.execute("SELECT * FROM routing_rules WHERE active = 1 ORDER BY id").fetchall()
        for row in rules:
            rule = dict(row)
            pattern = rule["pattern"].casefold()
            context = _decode(rule["context_json"], {})
            matched = False
            if rule["rule_type"] in {"ticket_number", "project_name"}:
                matched = pattern in combined
            elif rule["rule_type"] == "filename_phrase":
                matched = pattern in filename
            elif rule["rule_type"] == "sender_subject":
                sender = str(metadata.get("sender", "")).casefold()
                subject = str(metadata.get("subject", "")).casefold()
                matched = pattern in subject and str(context.get("sender_domain", "")).casefold() in sender
            elif rule["rule_type"] == "meeting_workstream":
                matched = pattern == str(source["meeting_name"] or "").casefold() and str(context.get("workstream", "")).casefold() in combined
            if matched:
                return rule
        return None

    def list_reviews(self, status: str = "open") -> list[dict[str, Any]]:
        if status not in {"open", "resolved", "dismissed", "all"}:
            raise ValidationError("Invalid review status")
        where = "" if status == "all" else "WHERE r.status = ?"
        params: tuple[Any, ...] = () if status == "all" else (status,)
        with self.db.connect() as connection:
            rows = connection.execute(
                f"""
                SELECT r.*, p.name AS project_name, s.original_filename
                FROM review_items r
                LEFT JOIN projects p ON p.id = r.project_id
                LEFT JOIN sources s ON s.id = r.source_id
                {where} ORDER BY r.created_at DESC, r.id DESC
                """,
                params,
            ).fetchall()
            decoded = [self._decode_review(row) for row in rows]
            for review in decoded:
                for evidence in review["evidence"]:
                    if isinstance(evidence, dict) and isinstance(evidence.get("citations"), list):
                        evidence["citations"] = self._enrich_citations(connection, evidence["citations"])
        return decoded

    def _decode_review(self, row: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
        item = dict(row)
        item["evidence"] = _decode(item.pop("evidence_json", "[]"), [])
        item["options"] = _decode(item.pop("options_json", "[]"), [])
        item["resolution"] = _decode(item.pop("resolution_json", None), None)
        return item

    def resolve_review(self, review_id: int, resolution: dict[str, Any]) -> dict[str, Any]:
        with self.db.connect() as connection:
            pending_review = connection.execute(
                "SELECT kind FROM review_items WHERE id = ?", (review_id,)
            ).fetchone()
        if pending_review and pending_review["kind"] == "project_fit":
            return self._resolve_project_fit_review(review_id, resolution)
        action = resolution.get("action", "apply")
        if action not in {"apply", "dismiss"}:
            raise ValidationError("Review action must be apply or dismiss")
        now = utc_now()
        routing_publication: dict[str, Any] = {}
        try:
            with self.db.transaction() as connection:
                review = connection.execute("SELECT * FROM review_items WHERE id = ?", (review_id,)).fetchone()
                if not review:
                    raise NotFoundError("Review item not found")
                if review["status"] != "open":
                    raise ConflictError("Review item is already resolved")
                if action == "dismiss":
                    connection.execute(
                        "UPDATE review_items SET status = 'dismissed', resolution_json = ?, resolved_at = ? WHERE id = ?",
                        (_json(resolution), now, review_id),
                    )
                # Delete this compatibility branch after all pre-003 review rows have been resolved or
                # after an archive rebuild, which never recreates cross_project_evidence reviews.
                elif review["kind"] in {"multi_project_route", "cross_project_evidence"}:
                    self._resolve_routing_review(
                        connection, review, resolution, now,
                        learn_rule=review["kind"] == "multi_project_route",
                        publication=routing_publication,
                    )
                elif review["kind"] == "action_close":
                    action_id = int(resolution.get("action_item_id") or 0)
                    item = connection.execute(
                        "SELECT * FROM action_items WHERE id = ? AND project_id = ?", (action_id, review["project_id"])
                    ).fetchone()
                    if not item:
                        raise ValidationError("Action item was not found in this project")
                    connection.execute(
                        "UPDATE action_items SET state = 'complete', completed_at = ?, updated_at = ? WHERE id = ?",
                        (now, now, action_id),
                    )
                    connection.execute(
                        "UPDATE review_items SET status = 'resolved', resolution_json = ?, resolved_at = ? WHERE id = ?",
                        (_json(resolution), now, review_id),
                    )
                elif review["kind"] == "project_field_recommendation":
                    field = resolution.get("field")
                    value = resolution.get("value")
                    if field not in {"status", "priority"}:
                        raise ValidationError("Only a user-confirmed status or priority recommendation can be applied")
                    valid = STATUSES if field == "status" else PRIORITIES
                    if value not in valid:
                        raise ValidationError(f"Invalid project {field}")
                    project = self._project(connection, review["project_id"])
                    connection.execute(
                        f"UPDATE projects SET {field} = ?, updated_at = ?, completed_at = ? WHERE id = ?",
                        (value, now, now if field == "status" and value == "Complete" else project["completed_at"], review["project_id"]),
                    )
                    text = f"User approved changing {field} from {project[field]} to {value}."
                    connection.execute(
                        "INSERT INTO project_updates(project_id, source_id, update_type, text, citations_json, created_at) VALUES (?, ?, 'manual_field', ?, '[]', ?)",
                        (review["project_id"], review["source_id"], text, now),
                    )
                    connection.execute(
                        "UPDATE review_items SET status = 'resolved', resolution_json = ?, resolved_at = ? WHERE id = ?",
                        (_json(resolution), now, review_id),
                    )
                elif review["kind"] == "inferred_action":
                    evidence = _decode(review["evidence_json"], [])
                    proposed = evidence[0] if evidence and isinstance(evidence[0], dict) else {}
                    description = normalize_text(str(resolution.get("description") or proposed.get("description") or ""))
                    assignee_type = resolution.get("assignee_type") or proposed.get("assignee_type")
                    assignee_value = normalize_text(str(resolution.get("assignee_value") or proposed.get("assignee_value") or ""))
                    due = resolution.get("due_date") or proposed.get("due_date")
                    if not description or assignee_type not in ASSIGNEE_TYPES or not assignee_value or not due:
                        raise ValidationError("Confirmed action description, assignee, and due date are required")
                    try:
                        date.fromisoformat(str(due))
                    except ValueError as exc:
                        raise ValidationError("Confirmed action due date must be an ISO date") from exc
                    citations = proposed.get("citations") if isinstance(proposed.get("citations"), list) else []
                    connection.execute(
                        """INSERT INTO action_items(project_id, description, assignee_type, assignee_value,
                           due_date, state, source_id, citations_json, created_by, created_at, updated_at)
                           VALUES (?, ?, ?, ?, ?, 'open', ?, ?, 'user', ?, ?)""",
                        (review["project_id"], description, assignee_type, assignee_value, due,
                         review["source_id"], _json(citations), now, now),
                    )
                    connection.execute(
                        "UPDATE review_items SET status = 'resolved', resolution_json = ?, resolved_at = ? WHERE id = ?",
                        (_json(resolution), now, review_id),
                    )
                else:
                    connection.execute(
                        "UPDATE review_items SET status = 'resolved', resolution_json = ?, resolved_at = ? WHERE id = ?",
                        (_json(resolution), now, review_id),
                    )
                source_id = review["source_id"]
                if source_id:
                    root = connection.execute("SELECT * FROM sources WHERE id = ?", (source_id,)).fetchone()
                    # The second condition finishes legacy direct-routing reviews created before migration 003.
                    if root and (root["project_id"] is None or review["kind"] == "cross_project_evidence"):
                        open_count = connection.execute(
                            "SELECT count(*) FROM review_items WHERE source_id = ? AND status = 'open'", (source_id,)
                        ).fetchone()[0]
                        if open_count == 0:
                            connection.execute(
                                "UPDATE sources SET processing_state = 'complete', processed_at = ? WHERE id = ?",
                                (now, source_id),
                            )
                updated = connection.execute(
                    """SELECT r.*, p.name AS project_name, s.original_filename
                       FROM review_items r LEFT JOIN projects p ON p.id = r.project_id
                       LEFT JOIN sources s ON s.id = r.source_id WHERE r.id = ?""",
                    (review_id,),
                ).fetchone()
        except Exception:
            self._discard_routing_staging(routing_publication.get("staging_path"))
            raise
        routing_published = True
        if routing_publication.get("staging_path"):
            try:
                routing_published = self._publish_routing_staging(
                    routing_publication["staging_path"]
                )
            except Exception:
                routing_published = False
                LOGGER.warning(
                    "Routing review %s committed but package publication is pending recovery",
                    review_id,
                    exc_info=True,
                )
        decoded = self._decode_review(updated)
        # cross_project_evidence is retained here solely for pre-migration review compatibility.
        if action == "apply" and review["kind"] in {"multi_project_route", "cross_project_evidence"}:
            if review["source_id"]:
                self._refresh_source_archive(int(review["source_id"]))
            routed = decoded.get("resolution") or {}
            derived_source_id = routed.get("derived_source_id")
            target_project_id = routed.get("target_project_id")
            if derived_source_id and target_project_id:
                if routing_published:
                    self._refresh_source_archive(int(derived_source_id), processing_status="complete")
                self._write_knowledge_history(str(target_project_id))
                try:
                    self.regenerate_living_summary(str(target_project_id), advance_revision=False)
                except UnexpectedSummaryError:
                    # Routing committed; return that result while its summary remains visibly failed.
                    pass
        return decoded

    def _resolve_project_fit_review(
        self, review_id: int, resolution: dict[str, Any]
    ) -> dict[str, Any]:
        action = str(resolution.get("action") or "")
        if action not in {"keep", "move", "remove"}:
            raise ValidationError("Project-fit review action must be keep, move, or remove")
        with self.db.connect() as connection:
            review = connection.execute(
                "SELECT * FROM review_items WHERE id = ?", (review_id,)
            ).fetchone()
            if not review:
                raise NotFoundError("Review item not found")
            if review["status"] != "open":
                raise ConflictError("Review item is already resolved")
            source = connection.execute(
                "SELECT * FROM sources WHERE id = ?", (review["source_id"],)
            ).fetchone()
        if not source:
            raise ValidationError("Review source no longer exists")
        if source["memory_state"] != "pending":
            raise ConflictError("Only a pending source can be reassigned before processing")
        reason = normalize_text(str(resolution.get("reason") or "User resolved the project-fit review."))
        if action == "remove":
            self.remove_source_from_memory(
                str(source["project_id"]), int(source["id"]), reason,
                exclude_review_id=review_id,
            )
        elif action == "move":
            target_project_id = str(resolution.get("target_project_id") or "")
            if not target_project_id or target_project_id == source["project_id"]:
                raise ValidationError("Choose a different project before moving this source")
            with self.db.connect() as connection:
                target = self._project(connection, target_project_id)
            destination = Path(target["folder_path"]) / Path(source["ingestion_path"]).name
            self._move_source_package(
                int(source["id"]), destination, memory_state="pending",
                event_type="moved_before_processing", reason=reason,
                to_project_id=target_project_id, project_fit_confirmed=True,
            )
        else:
            with self.db.transaction() as connection:
                current = connection.execute(
                    "SELECT * FROM sources WHERE id = ?", (source["id"],)
                ).fetchone()
                connection.execute(
                    """UPDATE sources SET project_fit_confirmed = 1,
                       memory_state_changed_at = ? WHERE id = ?""",
                    (utc_now(), source["id"]),
                )
                self._record_source_lifecycle(
                    connection, current, "project_fit_confirmed", reason,
                    from_project_id=current["project_id"], to_project_id=current["project_id"],
                    details={"confirmed_by": "user", "review_id": review_id},
                )
            self._refresh_source_archive(int(source["id"]))
        now = utc_now()
        with self.db.transaction() as connection:
            connection.execute(
                """UPDATE review_items SET status = 'resolved', resolution_json = ?, resolved_at = ?
                   WHERE id = ?""",
                (_json(resolution), now, review_id),
            )
            if action != "remove":
                connection.execute(
                    """UPDATE sources SET processing_state = 'captured', error_code = NULL,
                       error_message = NULL WHERE id = ?""",
                    (source["id"],),
                )
            updated = connection.execute(
                """SELECT r.*, p.name AS project_name, s.original_filename
                   FROM review_items r LEFT JOIN projects p ON p.id = r.project_id
                   LEFT JOIN sources s ON s.id = r.source_id WHERE r.id = ?""",
                (review_id,),
            ).fetchone()
        decoded = self._decode_review(updated)
        if action != "remove":
            state = self.process_source(int(source["id"]))
            decoded["source_processing_state"] = state
            with self.db.connect() as connection:
                current_source = connection.execute(
                    "SELECT project_id FROM sources WHERE id = ?", (source["id"],)
                ).fetchone()
            decoded["target_project_id"] = current_source["project_id"]
        return decoded

    def _resolve_routing_review(
        self, connection: sqlite3.Connection, review: sqlite3.Row,
        resolution: dict[str, Any], now: str, *, learn_rule: bool,
        publication: dict[str, Any],
    ) -> None:
        target_project_id = str(resolution.get("target_project_id") or "")
        target = self._project(connection, target_project_id)
        evidence = _decode(review["evidence_json"], [])
        if not evidence:
            raise ValidationError("Routing review has no preserved evidence")
        segment = evidence[0]
        confirmed_rule: dict[str, Any] | None = None
        if learn_rule:
            rule = resolution.get("rule") or segment.get("suggested_rule") or {}
            rule_type = str(rule.get("rule_type") or "")
            pattern = normalize_text(str(rule.get("pattern") or "")).casefold()
            context = rule.get("context") if isinstance(rule.get("context"), dict) else {}
            # Validate every user-controlled rule field before creating or editing
            # the linked archive package. A rejected rule must leave no sidecar for
            # disaster recovery to mistake for confirmed memory.
            self._validate_routing_rule(rule_type, pattern, context, target)
            confirmed_rule = {
                "rule_type": rule_type, "pattern": pattern, "context": context,
            }
        citations = segment.get("citations", [])
        source = connection.execute("SELECT * FROM sources WHERE id = ?", (review["source_id"],)).fetchone()
        if not source:
            raise ValidationError("Routing source no longer exists")
        allowed_rows = connection.execute(
            "SELECT id, source_id FROM source_chunks WHERE source_id = ?", (source["id"],)
        ).fetchall()
        allowed = {(int(row["source_id"]), int(row["id"])) for row in allowed_rows}
        valid_citations = self._validate_citations(citations, allowed)
        created = datetime.fromisoformat(str(source["created_at"]).replace("Z", "+00:00"))
        linked_ingestion_id = source["ingestion_id"] or self._legacy_ingestion_id(source)
        linked_package = self._under_root(
            Path(target["folder_path"]) /
            ingestion_folder_name(
                created, "linked-source", source["source_title"] or source["original_filename"],
                linked_ingestion_id,
            )
        )
        package_identity = {
            "ingestion_id": linked_ingestion_id,
            "linked_ingestion_id": linked_ingestion_id,
            "database_project_id": target_project_id,
            "project_id": target["archive_id"],
            "source_type": "linked-source",
        }
        base_linked_segments: list[Any] = []
        if linked_package.exists():
            if not self._routing_manifest_matches(
                read_json(linked_package / "manifest.json", None), package_identity
            ):
                raise ConflictError("Routing destination contains unrelated data")
            existing_segments = read_json(
                linked_package / "Assistant" / "linked-segments.json", None
            )
            if not isinstance(existing_segments, list):
                raise ConflictError("Linked routing segment metadata is invalid")
            base_linked_segments = existing_segments
        legacy_incomplete = self._under_root(
            linked_package.parent / f"_INCOMPLETE_LINK_{review['id']}"
        )
        if legacy_incomplete.exists():
            if not self._routing_manifest_matches(
                read_json(legacy_incomplete / "manifest.json", None), package_identity
            ):
                raise ConflictError("Legacy routing staging contains unrelated data")
            self._quarantine_routing_path(
                legacy_incomplete, f"legacy-incomplete-review-{review['id']}"
            )
        staging = self._under_root(
            self._routing_staging_root() / f"routing-review-{review['id']}"
        )
        if staging.exists():
            self._quarantine_routing_path(staging, f"superseded-review-{review['id']}")
        staging.mkdir(parents=False)
        publication["staging_path"] = staging
        mode = "segments" if linked_package.exists() else "package"
        atomic_write_json(staging / "publication.json", {
            "review_id": int(review["id"]),
            "final_relative_path": relative_to_root(
                linked_package, self.settings.app.one_drive_root
            ),
            "mode": mode,
            "package_identity": package_identity,
        })
        if not linked_package.exists():
            staged_package = staging / "package"
            staged_package.mkdir(parents=False)
            self._initial_assistant_files(
                staged_package, linked_ingestion_id, target_project_id,
                source["source_title"] or source["original_filename"],
            )
            atomic_write_json(staged_package / "Assistant" / "linked-segments.json", [])
            atomic_write_json(staged_package / "manifest.json", {
                "schema_version": SCHEMA_VERSION,
                "ingestion_id": linked_ingestion_id,
                "routing_review_id": int(review["id"]),
                "project_id": target["archive_id"],
                "database_project_id": target_project_id,
                "source_type": "linked-source",
                "title": source["source_title"] or source["original_filename"],
                "created_at": now,
                "source_date": source["source_date"],
                "capture_method": "routing_review",
                "canonical_source": False,
                "linked_ingestion_id": linked_ingestion_id,
                "canonical_source_path": relative_to_root(Path(source["ingestion_path"]), self.settings.app.one_drive_root),
                "processing_status": "complete",
                "memory_state": "active",
                "project_fit_confirmed": True,
                "original_files": [],
                "assistant_files": [
                    "Assistant/linked-segments.json", "Assistant/source-summary.md",
                    "Assistant/knowledge-items.json", "Assistant/citations.json",
                    "Assistant/source-lifecycle.jsonl",
                ],
                "extractor_version": "1.0",
                "errors": [],
            })
        derived_native = f"routed-review:{review['id']}"
        existing = connection.execute(
            "SELECT id FROM sources WHERE project_id = ? AND native_id = ?",
            (target_project_id, derived_native),
        ).fetchone()
        if existing:
            derived_source_id = int(existing["id"])
        else:
            cursor = connection.execute(
                """
                INSERT INTO sources(project_id, parent_source_id, source_type, native_id, sha256,
                  original_filename, original_path, metadata_json, meeting_name, meeting_date,
                  processing_state, model_id, created_at, processed_at, ingestion_path, source_title,
                  source_date, capture_method, canonical_source, linked_ingestion_id, source_summary,
                  memory_state, project_fit_confirmed, memory_state_changed_at)
                VALUES (?, ?, 'routed_segment', ?, ?, ?, ?, ?, ?, ?, 'complete', ?, ?, ?, ?, ?, ?,
                        'routing_review', 0, ?, ?, 'active', 1, ?)
                """,
                (target_project_id, source["id"], derived_native, source["sha256"],
                 source["original_filename"], source["original_path"],
                 _json({"review_id": review["id"]}), source["meeting_name"], source["meeting_date"],
                 source["model_id"], now, now, str(linked_package),
                 source["source_title"] or source["original_filename"], source["source_date"],
                 linked_ingestion_id, normalize_text(str(segment["text"])), now),
            )
            derived_source_id = int(cursor.lastrowid)
            derived_citations = []
            for sequence, citation in enumerate(valid_citations):
                chunk = connection.execute(
                    "SELECT * FROM source_chunks WHERE id = ? AND source_id = ?",
                    (citation["chunk_id"], citation["source_id"]),
                ).fetchone()
                inserted = connection.execute(
                    """
                    INSERT INTO source_chunks(source_id, project_id, sequence, text, locator, processing_state, processed_at)
                    VALUES (?, ?, ?, ?, ?, 'complete', ?)
                    """,
                    (derived_source_id, target_project_id, sequence, chunk["text"], chunk["locator"], now),
                )
                derived_citations.append({"source_id": derived_source_id, "chunk_id": int(inserted.lastrowid)})
            text = normalize_text(str(segment["text"]))
            citation_payload = []
            for original_citation, derived_citation in zip(valid_citations, derived_citations, strict=True):
                original_chunk = connection.execute(
                    "SELECT * FROM source_chunks WHERE id = ?", (original_citation["chunk_id"],)
                ).fetchone()
                try:
                    canonical_relative = relative_to_root(
                        Path(source["original_path"]), Path(source["ingestion_path"])
                    )
                except ValueError:
                    canonical_relative = Path(source["original_path"]).name
                citation_id = self._store_citation_record(
                    connection,
                    source_id=derived_source_id,
                    chunk_id=derived_citation["chunk_id"],
                    original_relative_path=canonical_relative,
                    display_name=source["original_filename"],
                    source_type=source["source_type"],
                    locator=original_chunk["locator"],
                    excerpt=original_chunk["text"][:1200],
                    source_date=source["source_date"] or source["created_at"],
                    created_at=now,
                )
                citation_payload.append({**derived_citation, "citation_id": citation_id})
            connection.execute(
                "INSERT INTO project_updates(project_id, source_id, update_type, text, citations_json, model_id, created_at) VALUES (?, ?, 'routed_review', ?, ?, ?, ?)",
                (target_project_id, derived_source_id, text, _json(citation_payload), source["model_id"], now),
            )
            connection.execute(
                """INSERT INTO knowledge_items(
                   id, project_id, source_id, text, category, source_date, citation_ids_json,
                   review_status, created_at
                   ) VALUES (?, ?, ?, ?, ?, ?, ?, 'unreviewed', ?)""",
                (stable_id("K"), target_project_id, derived_source_id, text,
                 self._knowledge_category(text), source["source_date"] or source["created_at"],
                 _json([item["citation_id"] for item in citation_payload]), now),
            )
            connection.execute(
                """UPDATE projects SET latest_change = ?, updated_at = ?,
                   summary_revision = summary_revision + 1, summary_generation_state = 'stale'
                   WHERE id = ?""",
                (text, now, target_project_id),
            )
            linked_segments = list(base_linked_segments)
            linked_segments = [
                item for item in linked_segments
                if not isinstance(item, dict) or item.get("review_id") != review["id"]
            ]
            linked_segments.append({
                "review_id": review["id"], "text": text, "citations": citation_payload,
                "confirmed_at": now, "target_project_id": target["archive_id"],
            })
            if mode == "package":
                linked_segments_path = staging / "package" / "Assistant" / "linked-segments.json"
                atomic_write_json(linked_segments_path, linked_segments)
            else:
                atomic_write_json(staging / "linked-segments.json", linked_segments)
        stored_resolution = {**resolution, "derived_source_id": derived_source_id}
        if confirmed_rule is not None:
            connection.execute(
                """
                INSERT OR IGNORE INTO routing_rules(rule_type, pattern, context_json, target_project_id,
                  created_from_review_id, active, created_at)
                VALUES (?, ?, ?, ?, ?, 1, ?)
                """,
                (
                    confirmed_rule["rule_type"], confirmed_rule["pattern"],
                    _json(confirmed_rule["context"]), target_project_id, review["id"], now,
                ),
            )
            stored_resolution["confirmed_rule"] = confirmed_rule
        connection.execute(
            "UPDATE review_items SET project_id = ?, status = 'resolved', resolution_json = ?, resolved_at = ? WHERE id = ?",
            (target_project_id, _json(stored_resolution), now, review["id"]),
        )

    def _validate_routing_rule(
        self, rule_type: str, pattern: str, context: dict[str, Any], target: sqlite3.Row
    ) -> None:
        if rule_type not in RULE_TYPES or not 3 <= len(pattern) <= 120:
            raise ValidationError("A valid narrow routing rule preview must be confirmed")
        if any(character in pattern for character in ("/", "\\", "\x00")):
            raise ValidationError("Routing rule patterns cannot contain paths")
        if rule_type == "ticket_number" and not re.fullmatch(r"[a-z]{2,12}\d{3,20}", pattern):
            raise ValidationError("Ticket-number rules require one exact ticket number")
        if rule_type == "project_name" and pattern != normalize_text(target["name"]).casefold():
            raise ValidationError("Project-name rules require the exact confirmed project name")
        if rule_type == "sender_subject":
            domain = normalize_text(str(context.get("sender_domain", ""))).casefold()
            if not domain or any(character.isspace() for character in domain):
                raise ValidationError("Sender/subject rules require one sender domain and subject phrase")
        if rule_type == "meeting_workstream" and not normalize_text(str(context.get("workstream", ""))):
            raise ValidationError("Meeting rules require a recurring meeting name and named workstream")

    def list_routing_rules(self) -> list[dict[str, Any]]:
        with self.db.connect() as connection:
            rows = connection.execute(
                """SELECT r.*, p.name AS target_project_name
                   FROM routing_rules r JOIN projects p ON p.id = r.target_project_id
                   ORDER BY r.created_at DESC"""
            ).fetchall()
        result = []
        for row in rows:
            item = dict(row)
            item["active"] = bool(item["active"])
            item["context"] = _decode(item.pop("context_json", "{}"), {})
            result.append(item)
        return result

    def set_routing_rule_active(self, rule_id: int, active: bool) -> dict[str, Any]:
        with self.db.transaction() as connection:
            if not connection.execute("SELECT 1 FROM routing_rules WHERE id = ?", (rule_id,)).fetchone():
                raise NotFoundError("Routing rule not found")
            try:
                connection.execute("UPDATE routing_rules SET active = ? WHERE id = ?", (int(active), rule_id))
            except sqlite3.IntegrityError as exc:
                raise ConflictError("An identical active routing rule already exists") from exc
        return next(item for item in self.list_routing_rules() if item["id"] == rule_id)

    def list_actions(self, project_id: str) -> list[dict[str, Any]]:
        with self.db.connect() as connection:
            self._project(connection, project_id)
            rows = connection.execute(
                """SELECT a.* FROM action_items a LEFT JOIN sources s ON s.id = a.source_id
                   WHERE a.project_id = ?
                     AND (a.source_id IS NULL OR a.created_by = 'user' OR s.memory_state = 'active')
                   ORDER BY a.state = 'complete', a.due_date, a.id""",
                (project_id,),
            ).fetchall()
            decoded = self._decode_active_actions(connection, rows)
        return decoded

    def create_action(self, project_id: str, values: dict[str, Any]) -> dict[str, Any]:
        description = normalize_text(str(values.get("description", "")))
        assignee_type = values.get("assignee_type", "me")
        assignee_value = normalize_text(str(values.get("assignee_value", "me")))
        due = values.get("due_date") or None
        state = values.get("state", "open")
        if not description or assignee_type not in ASSIGNEE_TYPES or not assignee_value or state not in ACTION_STATES:
            raise ValidationError("Action description, assignee, and state are invalid")
        if due:
            try:
                date.fromisoformat(str(due))
            except ValueError as exc:
                raise ValidationError("due_date must be an ISO date") from exc
        now = utc_now()
        with self.db.transaction() as connection:
            self._project(connection, project_id)
            cursor = connection.execute(
                """
                INSERT INTO action_items(project_id, description, assignee_type, assignee_value,
                  due_date, state, progress_text, created_by, created_at, updated_at, completed_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, 'user', ?, ?, ?)
                """,
                (project_id, description, assignee_type, assignee_value, due, state,
                 normalize_text(str(values.get("progress_text", ""))) or None,
                 now, now, now if state == "complete" else None),
            )
            row = connection.execute("SELECT * FROM action_items WHERE id = ?", (cursor.lastrowid,)).fetchone()
            decoded = self._decode_active_action(connection, row)
        return decoded

    def update_action(self, project_id: str, action_id: int, values: dict[str, Any]) -> dict[str, Any]:
        allowed = {"description", "assignee_type", "assignee_value", "due_date", "state", "progress_text"}
        if set(values) - allowed:
            raise ValidationError("Unknown action item field")
        if values.get("state") == "complete":
            raise ValidationError("Use the explicit complete action endpoint")
        with self.db.transaction() as connection:
            row = connection.execute(
                "SELECT * FROM action_items WHERE id = ? AND project_id = ?", (action_id, project_id)
            ).fetchone()
            if not row:
                raise NotFoundError("Action item not found")
            merged = dict(row)
            merged.update(values)
            if merged["assignee_type"] not in ASSIGNEE_TYPES or merged["state"] not in {"open", "blocked"}:
                raise ValidationError("Invalid action item assignee or state")
            if merged.get("due_date"):
                try:
                    date.fromisoformat(str(merged["due_date"]))
                except ValueError as exc:
                    raise ValidationError("due_date must be an ISO date") from exc
            connection.execute(
                """UPDATE action_items SET description = ?, assignee_type = ?, assignee_value = ?,
                   due_date = ?, state = ?, progress_text = ?, updated_at = ?, completed_at = NULL WHERE id = ?""",
                (normalize_text(str(merged["description"])), merged["assignee_type"],
                 normalize_text(str(merged["assignee_value"])), merged.get("due_date"), merged["state"],
                 normalize_text(str(merged.get("progress_text") or "")) or None, utc_now(), action_id),
            )
            if "progress_text" in values:
                connection.execute(
                    "UPDATE action_items SET citations_json = '[]' WHERE id = ?", (action_id,)
                )
            updated = connection.execute("SELECT * FROM action_items WHERE id = ?", (action_id,)).fetchone()
            decoded = self._decode_active_action(connection, updated)
        return decoded

    def complete_action(self, project_id: str, action_id: int, confirmed: bool) -> dict[str, Any]:
        if not confirmed:
            raise ValidationError("Explicit user confirmation is required")
        now = utc_now()
        with self.db.transaction() as connection:
            row = connection.execute(
                "SELECT * FROM action_items WHERE id = ? AND project_id = ?", (action_id, project_id)
            ).fetchone()
            if not row:
                raise NotFoundError("Action item not found")
            connection.execute(
                "UPDATE action_items SET state = 'complete', completed_at = ?, updated_at = ? WHERE id = ?",
                (now, now, action_id),
            )
            updated = connection.execute("SELECT * FROM action_items WHERE id = ?", (action_id,)).fetchone()
            decoded = self._decode_active_action(connection, updated)
        return decoded

    def ask_project(self, project_id: str, question: str) -> dict[str, Any]:
        clean = normalize_text(question)
        if not clean:
            raise ValidationError("Question cannot be blank")
        with self.db.connect() as connection:
            project = self._project(connection, project_id)
            recent_updates = [self._decode_update(row) for row in connection.execute(
                """SELECT u.* FROM project_updates u LEFT JOIN sources s ON s.id = u.source_id
                   WHERE u.project_id = ? AND (u.source_id IS NULL OR s.memory_state = 'active')
                   ORDER BY u.created_at DESC, u.id DESC LIMIT 10""",
                (project_id,),
            ).fetchall()]
        chunks = self.db.search_chunks(project_id, clean, limit=12)
        if not chunks:
            return {
                "answer": "", "claims": [],
                "uncertainty": "No matching evidence was found in this project.",
                "model_id": self.llm.model_id, "evidence_dropped_chunks": 0,
            }
        evidence = [{
            "source_id": int(row["source_id"]), "chunk_id": int(row["id"]), "text": row["text"],
            "locator": row["locator"], "original_filename": row["original_filename"],
            "meeting_name": row.get("meeting_name"), "meeting_date": row.get("meeting_date"),
        } for row in chunks]
        bounded, dropped = self._bounded_evidence(evidence)
        if not bounded:
            return {
                "answer": "", "claims": [],
                "uncertainty": "No matching project evidence was found within the configured evidence limit.",
                "model_id": self.llm.model_id, "evidence_dropped_chunks": dropped,
            }
        result = self.llm.chat(
            clean,
            _json({
                "current_summary": (
                    project["current_summary"] if project["summary_generation_state"] == "current" else ""
                ),
                "recent_updates": recent_updates,
            }),
            bounded,
        )
        if not isinstance(result, dict) or not isinstance(result.get("answer"), str):
            raise LlmContractError("Chat response is malformed")
        allowed = {(item["source_id"], item["chunk_id"]) for item in bounded}
        claims = result.get("claims", [])
        if not isinstance(claims, list):
            raise LlmContractError("Chat claims must be an array")
        enriched: list[dict[str, Any]] = []
        for claim in claims:
            if not isinstance(claim, dict) or not normalize_text(str(claim.get("text", ""))):
                raise LlmContractError("Every chat claim requires text")
            citations = self._validate_citations(claim.get("citations"), allowed)
            claim_copy = {"text": str(claim.get("text", "")), "citations": []}
            for citation in citations:
                matching = next(item for item in bounded if item["source_id"] == citation["source_id"] and item["chunk_id"] == citation["chunk_id"])
                claim_copy["citations"].append({**citation, **{
                    "locator": matching["locator"], "original_filename": matching["original_filename"],
                    "meeting_name": matching.get("meeting_name"), "meeting_date": matching.get("meeting_date"),
                    "excerpt": matching["text"][:600],
                }})
            enriched.append(claim_copy)
        if result["answer"].strip() and not enriched:
            raise LlmContractError("A substantive chat answer must contain claim citations")
        if not result["answer"].strip() and not normalize_text(str(result.get("uncertainty") or "")):
            raise LlmContractError("Chat must return either a cited answer or an uncertainty statement")
        return {
            "answer": result["answer"], "claims": enriched, "uncertainty": result.get("uncertainty"),
            "model_id": self.llm.model_id, "evidence_dropped_chunks": dropped,
        }

    def get_chunk(self, chunk_id: int) -> dict[str, Any]:
        with self.db.connect() as connection:
            row = connection.execute(
                """SELECT c.*, s.original_filename, s.meeting_name, s.meeting_date
                   FROM source_chunks c JOIN sources s ON s.id = c.source_id WHERE c.id = ?""",
                (chunk_id,),
            ).fetchone()
            if not row:
                raise NotFoundError("Source excerpt not found")
        return dict(row)

    def get_source_original(self, source_id: int) -> tuple[Path, str]:
        with self.db.connect() as connection:
            row = connection.execute("SELECT * FROM sources WHERE id = ?", (source_id,)).fetchone()
            if not row:
                raise NotFoundError("Source not found")
        path = self._under_root(Path(row["original_path"]))
        if not path.is_file():
            raise NotFoundError("Preserved original is unavailable")
        return path, safe_filename(row["original_filename"], "source")

    def get_original_file(self, original_file_id: int) -> tuple[Path, str]:
        with self.db.connect() as connection:
            row = connection.execute(
                """SELECT f.*, s.ingestion_path FROM original_files f
                   JOIN sources s ON s.id = f.source_id WHERE f.id = ?""",
                (original_file_id,),
            ).fetchone()
            if not row:
                raise NotFoundError("Original file not found")
        path = self._under_root(Path(row["ingestion_path"]) / row["relative_path"])
        if not path.is_file():
            raise NotFoundError("Preserved original is unavailable locally; make it available in OneDrive and retry")
        if sha256_file(path) != row["sha256"]:
            raise ConflictError("Preserved original no longer matches its recorded SHA-256 hash")
        return path, safe_filename(row["original_name"], row["stored_name"])

    def get_citation_original(self, citation_id: str) -> tuple[Path, str]:
        with self.db.connect() as connection:
            row = connection.execute(
                """SELECT c.*, s.ingestion_path, s.linked_ingestion_id, s.original_filename,
                          s.original_path AS source_original_path,
                          s.sha256 AS source_sha256
                   FROM citation_records c JOIN sources s ON s.id = c.source_id WHERE c.id = ?""",
                (citation_id,),
            ).fetchone()
            if not row:
                raise NotFoundError("Citation not found")
            package = Path(row["ingestion_path"]) if row["ingestion_path"] else None
            source_original_path = Path(row["source_original_path"])
            source_sha256 = str(row["source_sha256"])
            original_source_id = int(row["source_id"])
            if row["linked_ingestion_id"]:
                canonical = connection.execute(
                    """SELECT id, ingestion_path, original_path, sha256 FROM sources
                       WHERE ingestion_id = ? AND canonical_source = 1""",
                    (row["linked_ingestion_id"],),
                ).fetchone()
                if canonical:
                    package = (
                        Path(canonical["ingestion_path"])
                        if canonical["ingestion_path"] else None
                    )
                    source_original_path = Path(canonical["original_path"])
                    source_sha256 = str(canonical["sha256"])
                    original_source_id = int(canonical["id"])
            original = None
            if package is not None:
                original = connection.execute(
                    """SELECT sha256 FROM original_files
                       WHERE source_id = ? AND relative_path = ?""",
                    (original_source_id, row["original_relative_path"]),
                ).fetchone()
        if package is None:
            # ServiceNow rows intentionally reference one shared import rather than
            # copying a multi-ticket export into every project package. Their
            # authoritative original and hash live directly on sources.
            path = self._under_root(source_original_path)
            expected_sha256 = source_sha256
        else:
            path = self._under_root(package / row["original_relative_path"])
            expected_sha256 = str(original["sha256"]) if original else ""
        if not path.is_file():
            raise NotFoundError("Cited original is unavailable locally; make it available in OneDrive and retry")
        if not expected_sha256 or sha256_file(path) != expected_sha256:
            raise ConflictError("Cited original no longer matches its recorded SHA-256 hash")
        return path, safe_filename(row["display_name"], row["original_filename"])

    def run_daily(self, run_date: date | None = None) -> dict[str, Any]:
        local_now = datetime.now().astimezone()
        target = run_date or local_now.date()
        with self.db.connect() as connection:
            existing = connection.execute(
                "SELECT * FROM daily_runs WHERE run_date = ?", (target.isoformat(),)
            ).fetchone()
        if existing:
            return self._decode_daily(existing)
        local_tz = local_now.tzinfo
        window_start_local = datetime.combine(target - timedelta(days=1), time.min, tzinfo=local_tz)
        window_end_local = datetime.combine(target, time.min, tzinfo=local_tz)
        window_start = window_start_local.astimezone(timezone.utc).isoformat()
        window_end = window_end_local.astimezone(timezone.utc).isoformat()
        with self.db.connect() as connection:
            rows = connection.execute(
                """
                SELECT u.id AS update_id, u.project_id, u.text, u.update_type, u.citations_json,
                       u.created_at, p.name AS project_name
                FROM project_updates u JOIN projects p ON p.id = u.project_id
                LEFT JOIN sources s ON s.id = u.source_id
                WHERE u.created_at >= ? AND u.created_at < ?
                  AND (u.source_id IS NULL OR s.memory_state = 'active')
                ORDER BY u.created_at, u.id
                """,
                (window_start, window_end),
            ).fetchall()
            review_count = int(connection.execute(
                "SELECT count(*) FROM review_items WHERE status = 'open'"
            ).fetchone()[0])
            action_count = int(connection.execute(
                """SELECT count(*) FROM action_items a LEFT JOIN sources s ON s.id = a.source_id
                   WHERE a.updated_at >= ? AND a.updated_at < ?
                     AND (a.source_id IS NULL OR a.created_by = 'user' OR s.memory_state = 'active')""",
                (window_start, window_end),
            ).fetchone()[0])
            action_rows = self._decode_active_actions(connection, connection.execute(
                """SELECT a.id AS action_item_id, a.project_id, a.description, a.state,
                          a.progress_text, a.citations_json, a.created_at, a.updated_at,
                          a.source_id, a.created_by, p.name AS project_name
                   FROM action_items a JOIN projects p ON p.id = a.project_id
                   LEFT JOIN sources s ON s.id = a.source_id
                   WHERE a.updated_at >= ? AND a.updated_at < ?
                     AND (a.source_id IS NULL OR a.created_by = 'user' OR s.memory_state = 'active')
                   ORDER BY a.updated_at, a.id""",
                (window_start, window_end),
            ).fetchall())
        evidence = [{
            "kind": "project_update", "update_id": int(row["update_id"]), "project_id": row["project_id"],
            "project_name": row["project_name"], "text": row["text"],
            "update_type": row["update_type"], "citations": _decode(row["citations_json"], []),
        } for row in rows]
        action_evidence = [{
            "kind": "action_item", "action_item_id": int(row["action_item_id"]),
            "project_id": row["project_id"], "project_name": row["project_name"],
            "description": row["description"], "state": row["state"],
            "progress_text": row["progress_text"], "updated_at": row["updated_at"],
        } for row in action_rows]
        combined_evidence = [*evidence, *action_evidence]
        counts = {
            "projects_changed": len({item["project_id"] for item in combined_evidence}),
            "updates": len(evidence), "action_changes": action_count, "open_reviews": review_count,
        }
        result = self.llm.daily(combined_evidence, counts)
        if not isinstance(result, dict) or not isinstance(result.get("summary"), str):
            raise LlmContractError("Daily update response is malformed")
        supplied_ids = {item["update_id"] for item in evidence}
        raw_ids = result.get("update_ids", [])
        if not isinstance(raw_ids, list):
            raise LlmContractError("Daily update citations must be an array")
        try:
            cited_ids = [int(value) for value in raw_ids]
        except (TypeError, ValueError) as exc:
            raise LlmContractError("Daily update citation identifiers are invalid") from exc
        if any(value not in supplied_ids for value in cited_ids):
            raise LlmContractError("Daily update cited an update outside the prior-day window")
        if evidence and not cited_ids:
            raise LlmContractError("Daily project statements require committed update citations")
        now = utc_now()
        with self.db.transaction() as connection:
            connection.execute(
                """
                INSERT INTO daily_runs(run_date, window_start, window_end, summary_text,
                  project_changes_json, counts_json, model_id, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(run_date) DO UPDATE SET
                  window_start=excluded.window_start, window_end=excluded.window_end,
                  summary_text=excluded.summary_text, project_changes_json=excluded.project_changes_json,
                  counts_json=excluded.counts_json, model_id=excluded.model_id, updated_at=excluded.updated_at
                """,
                (target.isoformat(), window_start, window_end, result["summary"],
                 _json({"updates": evidence, "action_changes": action_evidence, "cited_update_ids": cited_ids}), _json(counts),
                 self.llm.model_id, now, now),
            )
            row = connection.execute("SELECT * FROM daily_runs WHERE run_date = ?", (target.isoformat(),)).fetchone()
        return self._decode_daily(row)

    def get_daily(self, run_date: date | None = None) -> dict[str, Any] | None:
        target = (run_date or datetime.now().astimezone().date()).isoformat()
        with self.db.connect() as connection:
            row = connection.execute("SELECT * FROM daily_runs WHERE run_date = ?", (target,)).fetchone()
        return self._decode_daily(row) if row else None

    def _decode_daily(self, row: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
        item = dict(row)
        item["project_changes"] = _decode(item.pop("project_changes_json", "{}"), {})
        item["counts"] = _decode(item.pop("counts_json", "{}"), {})
        return item

    def scheduler_status(self) -> dict[str, Any]:
        task_name = "CHIO Portfolio Assistant Morning Update"
        try:
            result = subprocess.run(
                ["schtasks", "/Query", "/TN", task_name, "/FO", "LIST"],
                capture_output=True, text=True, timeout=10, check=False,
            )
            installed = result.returncode == 0
        except (FileNotFoundError, OSError, subprocess.TimeoutExpired):
            installed = False
        return {
            "installed": installed,
            "message": "Morning task installed" if installed else "Morning task not installed",
            "run_time": self.settings.app.daily_run_time,
        }

    def configuration_status(self, *, test_llm: bool = False) -> dict[str, Any]:
        credential = self.llm.credential_status()
        status = {
            "database_path": str(self.settings.app.database_path),
            "one_drive_root": str(self.settings.app.one_drive_root),
            "bind": f"{self.settings.app.bind_host}:{self.settings.app.bind_port}",
            "retrieval_mode": self.db.fts_mode,
            "llm_adapter": self.settings.llm.adapter,
            "llm_model": self.llm.model_id,
            "llm_judgment_model": self.llm.judgment_model_id,
            "llm_endpoint": self.llm.endpoint_url,
            "api_key_present": credential["configured"],
            "api_key_source": credential["source"],
            "api_key_environment_override": credential["environment_override"],
            "api_key_local_saved": credential["local_key_present"],
            "credential_error": credential.get("credential_error", False),
            "model_preference_error": self.llm.model_preference_error,
            "scheduler": self.scheduler_status(),
        }
        if test_llm:
            try:
                status["llm_test"] = self.llm.test_connection()
            except (LlmUnavailable, LlmContractError) as exc:
                status["llm_test"] = {"ok": False, "error": _safe_error(exc)}
        return status

    def save_llm_api_key(self, api_key: str) -> dict[str, Any]:
        result = self.llm.save_api_key(api_key)
        self._invalidate_llm_model_catalog()
        return result

    def remove_llm_api_key(self) -> dict[str, Any]:
        result = self.llm.remove_api_key()
        self._invalidate_llm_model_catalog()
        return result

    def save_llm_models(self, routine_model: str, judgment_model: str) -> dict[str, str]:
        return self.llm.save_models(routine_model, judgment_model)

    def _invalidate_llm_model_catalog(self) -> None:
        with self._llm_model_catalog_lock:
            self._llm_model_catalog_generation += 1
            self._llm_model_catalog_result = None

    @staticmethod
    def _copy_llm_model_catalog_result(result: dict[str, Any]) -> dict[str, Any]:
        copied = dict(result)
        if "available_models" in copied:
            copied["available_models"] = list(copied["available_models"])
        return copied

    def refresh_llm_model_catalog(self, *, force: bool = False) -> dict[str, Any]:
        while True:
            while True:
                with self._llm_model_catalog_lock:
                    if self._llm_model_catalog_result is not None and not force:
                        return self._copy_llm_model_catalog_result(self._llm_model_catalog_result)
                    pending = self._llm_model_catalog_inflight
                    if pending is None:
                        pending = threading.Event()
                        self._llm_model_catalog_inflight = pending
                        generation = self._llm_model_catalog_generation
                        stale_result = self._llm_model_catalog_result
                        break
                pending.wait()

            result: dict[str, Any] | None = None
            try:
                credential = self.llm.credential_status()
                if not credential["configured"]:
                    result = {
                        "ok": False, "configured": False,
                        "error": "No GenAI.mil API key is configured",
                    }
                else:
                    try:
                        available_models = self.llm.list_models()
                    except (LlmUnavailable, LlmContractError) as exc:
                        result = {
                            "ok": False, "configured": True, "error": _safe_error(exc),
                        }
                        if stale_result and stale_result.get("available_models"):
                            result["available_models"] = list(stale_result["available_models"])
                            result["stale"] = True
                    else:
                        result = {
                            "ok": True, "configured": True,
                            "available_models": list(available_models),
                        }
            finally:
                with self._llm_model_catalog_lock:
                    generation_is_current = generation == self._llm_model_catalog_generation
                    if result is not None and generation_is_current:
                        self._llm_model_catalog_result = self._copy_llm_model_catalog_result(result)
                    if self._llm_model_catalog_inflight is pending:
                        self._llm_model_catalog_inflight = None
                    pending.set()

            if generation_is_current:
                return self._copy_llm_model_catalog_result(result)

    def llm_health(self) -> dict[str, Any]:
        catalog = self.refresh_llm_model_catalog(force=True)
        if not catalog["configured"]:
            return {
                "ok": False, "configured": False, "error": "No GenAI.mil API key is configured",
                "routine_model": self.llm.model_id,
                "judgment_model": self.llm.judgment_model_id,
            }
        if not catalog["ok"]:
            return {
                **catalog,
                "routine_model": self.llm.model_id,
                "judgment_model": self.llm.judgment_model_id,
            }
        available_models = catalog["available_models"]
        missing = [
            model for model in {self.llm.model_id, self.llm.judgment_model_id}
            if model not in available_models
        ]
        if missing:
            return {
                "ok": False, "configured": True,
                "error": "Choose available models in Settings before testing the API",
                "available_models": available_models,
                "unavailable_configured_models": sorted(missing, key=str.casefold),
                "routine_model": self.llm.model_id,
                "judgment_model": self.llm.judgment_model_id,
            }

        def probe(model_id: str) -> dict[str, Any]:
            try:
                return self.llm.test_connection(model_id)
            except (LlmUnavailable, LlmContractError) as exc:
                return {
                    "ok": False, "configured_model": model_id,
                    "error": _safe_error(exc),
                }

        routine = probe(self.llm.model_id)
        judgment = (
            routine if self.llm.judgment_model_id == self.llm.model_id
            else probe(self.llm.judgment_model_id)
        )
        result = {
            "ok": routine.get("ok") is True and judgment.get("ok") is True,
            "configured": True,
            "available_models": available_models,
            "routine": routine,
            "judgment": judgment,
            "routine_model": self.llm.model_id,
            "judgment_model": self.llm.judgment_model_id,
        }
        failures = []
        for label, health in (("Routine model", routine), ("Judgment model", judgment)):
            if health.get("ok") is not True:
                detail = health.get("error") or "did not report a successful response"
                failures.append(f"{label}: {detail}")
        if failures:
            result["error"] = "; ".join(failures)
        return result
