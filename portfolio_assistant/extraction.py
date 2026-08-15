from __future__ import annotations

import codecs
import hashlib
import re
from dataclasses import dataclass, field
from email import policy
from email.message import EmailMessage
from email.parser import BytesParser
from pathlib import Path, PureWindowsPath
from typing import Any

import extract_msg
from bs4 import BeautifulSoup
from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


SUPPORTED_SUFFIXES = {".eml", ".msg", ".txt", ".md", ".vtt", ".srt", ".docx", ".pdf", ".xlsx"}
TRANSCRIPT_SUFFIXES = {".vtt", ".srt"}
WINDOWS_RESERVED_NAMES = frozenset(
    {"CON", "PRN", "AUX", "NUL"}
    | {f"COM{index}" for index in range(1, 10)}
    | {f"LPT{index}" for index in range(1, 10)}
)


class UnsupportedSource(ValueError):
    pass


class ExtractionFailure(ValueError):
    pass


@dataclass
class AttachmentData:
    filename: str
    data: bytes
    content_type: str = "application/octet-stream"


@dataclass
class ExtractedSource:
    source_type: str
    native_id: str | None
    metadata: dict[str, Any]
    chunks: list[dict[str, Any]] = field(default_factory=list)
    attachments: list[AttachmentData] = field(default_factory=list)


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def safe_filename(value: str, fallback: str = "source") -> str:
    """Reduce an untrusted name to one safe Windows filename.

    Windows separator semantics are applied on every platform so the stored name
    never depends on the host OS: ``Path().name`` does not treat a backslash as a
    separator on POSIX, which leaves traversal segments such as ``..\\..\\evil.txt``
    embedded in the stored name.
    """
    raw = str(value or fallback).replace("/", "\\")
    name = PureWindowsPath(raw).name
    name = re.sub(r"[\x00-\x1f<>:\"/\\|?*]", "_", name).strip(" .")
    name = name[:180].strip(" .")
    if not name:
        return fallback
    if PureWindowsPath(name).stem.upper() in WINDOWS_RESERVED_NAMES:
        # CON, NUL, COM1 and friends resolve to devices rather than files, so an
        # archive copy written under that name silently discards its bytes.
        name = f"_{name}"
    return name


def decode_text_bytes(path: Path) -> str:
    """Decode a plain-text source, honouring byte-order marks before falling back.

    cp1252 accepts any byte sequence, so without the BOM check a UTF-16 file
    decodes to NUL-riddled mojibake instead of raising.
    """
    raw = path.read_bytes()
    # UTF-32 LE starts with the UTF-16 LE mark, so it has to be tested first.
    for mark, encoding in (
        (codecs.BOM_UTF32_LE, "utf-32"), (codecs.BOM_UTF32_BE, "utf-32"),
        (codecs.BOM_UTF16_LE, "utf-16"), (codecs.BOM_UTF16_BE, "utf-16"),
    ):
        if raw.startswith(mark):
            try:
                return raw.decode(encoding)
            except UnicodeDecodeError as exc:
                raise ExtractionFailure("The text file encoding is unsupported") from exc
    for encoding in ("utf-8-sig", "cp1252"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ExtractionFailure("The text file encoding is unsupported")


def html_to_text(value: str) -> str:
    soup = BeautifulSoup(value, "html.parser")
    for element in soup(["script", "style", "iframe", "object", "embed"]):
        element.decompose()
    return "\n".join(line.strip() for line in soup.get_text("\n").splitlines() if line.strip())


def _bounded(chunks: list[dict[str, Any]], max_bytes: int) -> list[dict[str, Any]]:
    total = sum(len(item["text"].encode("utf-8")) for item in chunks)
    if total > max_bytes:
        raise ExtractionFailure("Extracted text exceeds the configured size limit")
    return chunks


def _chunk_lines(text: str, prefix: str = "lines", max_chars: int = 3000) -> list[dict[str, Any]]:
    lines = text.splitlines()
    chunks: list[dict[str, Any]] = []
    buffer: list[str] = []
    start = 1
    for index, line in enumerate(lines, start=1):
        if buffer and sum(len(part) + 1 for part in buffer) + len(line) > max_chars:
            chunks.append({"text": "\n".join(buffer).strip(), "locator": f"{prefix} {start}-{index - 1}"})
            buffer = []
            start = index
        buffer.append(line)
    if buffer:
        chunks.append({"text": "\n".join(buffer).strip(), "locator": f"{prefix} {start}-{len(lines)}"})
    return [item for item in chunks if item["text"]]


def _extract_eml(path: Path, max_attachments: int) -> ExtractedSource:
    message = BytesParser(policy=policy.default).parsebytes(path.read_bytes())
    if not isinstance(message, EmailMessage):
        raise ExtractionFailure("The saved email could not be parsed")
    plain: list[str] = []
    html: list[str] = []
    attachments: list[AttachmentData] = []
    for part in message.walk():
        if part.is_multipart():
            continue
        disposition = part.get_content_disposition()
        filename = part.get_filename()
        if disposition == "attachment" or filename:
            if len(attachments) >= max_attachments:
                raise ExtractionFailure("Email attachment count exceeds the configured limit")
            attachments.append(AttachmentData(
                safe_filename(filename or f"attachment-{len(attachments) + 1}"),
                part.get_payload(decode=True) or b"",
                part.get_content_type(),
            ))
            continue
        try:
            content = part.get_content()
        except (LookupError, UnicodeError):
            payload = part.get_payload(decode=True) or b""
            content = payload.decode("utf-8", errors="replace")
        if part.get_content_type() == "text/plain":
            plain.append(str(content))
        elif part.get_content_type() == "text/html":
            html.append(html_to_text(str(content)))
    body = "\n\n".join(plain).strip() or "\n\n".join(html).strip()
    if not body:
        body = "(Saved email has no textual body.)"
    metadata = {
        "subject": str(message.get("Subject", "")),
        "sender": str(message.get("From", "")),
        "recipients": str(message.get("To", "")),
        "cc": str(message.get("Cc", "")),
        "timestamp": str(message.get("Date", "")),
        "message_id": str(message.get("Message-ID", "")),
        "in_reply_to": str(message.get("In-Reply-To", "")),
        "references": str(message.get("References", "")),
    }
    native = metadata["message_id"].strip() or None
    return ExtractedSource("eml", native, metadata, _chunk_lines(body, "Email body lines"), attachments)


def _extract_msg(path: Path, max_attachments: int) -> ExtractedSource:
    message = None
    try:
        message = extract_msg.Message(str(path), delayAttachments=True)
        body = (message.body or "").strip()
        html_body = getattr(message, "htmlBody", None)
        if not body and html_body:
            if isinstance(html_body, bytes):
                html_body = html_body.decode("utf-8", errors="replace")
            body = html_to_text(str(html_body))
        attachments: list[AttachmentData] = []
        for attachment in message.attachments:
            if len(attachments) >= max_attachments:
                raise ExtractionFailure("Email attachment count exceeds the configured limit")
            data = getattr(attachment, "data", None)
            if not isinstance(data, bytes):
                continue
            filename = getattr(attachment, "longFilename", None) or getattr(attachment, "shortFilename", None)
            attachments.append(AttachmentData(safe_filename(filename or f"attachment-{len(attachments) + 1}"), data))
        header = getattr(message, "headerDict", {}) or {}
        metadata = {
            "subject": message.subject or "",
            "sender": message.sender or "",
            "recipients": message.to or "",
            "cc": message.cc or "",
            "timestamp": str(message.date or ""),
            "message_id": str(header.get("Message-Id") or header.get("Message-ID") or ""),
            "conversation_topic": str(header.get("Thread-Topic") or ""),
            "conversation_index": str(header.get("Thread-Index") or ""),
        }
        native = metadata["message_id"].strip() or None
        return ExtractedSource(
            "msg", native, metadata,
            _chunk_lines(body or "(Saved email has no textual body.)", "Email body lines"),
            attachments,
        )
    except ExtractionFailure:
        raise
    except Exception as exc:
        raise ExtractionFailure("The Outlook MSG file could not be parsed") from exc
    finally:
        if message is not None:
            try:
                message.close()
            except Exception:
                pass


def _extract_docx(path: Path) -> ExtractedSource:
    try:
        document = Document(str(path))
        paragraphs = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]
    except Exception as exc:
        raise ExtractionFailure("The DOCX file could not be parsed") from exc
    if not paragraphs:
        raise UnsupportedSource("The DOCX file contains no extractable text")
    text = "\n".join(paragraphs)
    return ExtractedSource("docx", None, {}, _chunk_lines(text, "paragraphs"))


def _extract_pdf(path: Path) -> ExtractedSource:
    try:
        reader = PdfReader(str(path))
        page_count = len(reader.pages)
        chunks = []
        for number, page in enumerate(reader.pages, start=1):
            text = (page.extract_text() or "").strip()
            if text:
                chunks.extend(_chunk_lines(text, f"page {number} lines"))
    except Exception as exc:
        raise ExtractionFailure("The PDF file could not be parsed") from exc
    if not chunks:
        raise UnsupportedSource("The PDF has no text layer; OCR is not supported")
    return ExtractedSource("pdf", None, {"pages": page_count}, chunks)


def _extract_xlsx(path: Path) -> ExtractedSource:
    """Read displayed workbook values without evaluating formulas or macros."""
    workbook = None
    try:
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        chunks: list[dict[str, Any]] = []
        sheet_names: list[str] = []
        for worksheet in workbook.worksheets:
            sheet_names.append(worksheet.title)
            lines: list[str] = []
            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                values = [normalize_text(str(value)) for value in row if value is not None and str(value).strip()]
                if values:
                    lines.append(f"Row {row_number}: " + " | ".join(values))
            if lines:
                chunks.extend(_chunk_lines("\n".join(lines), f"sheet {worksheet.title} rows"))
    except Exception as exc:
        raise ExtractionFailure("The XLSX workbook could not be parsed") from exc
    finally:
        if workbook is not None:
            workbook.close()
    if not chunks:
        raise UnsupportedSource("The XLSX workbook contains no extractable displayed values")
    return ExtractedSource("xlsx", None, {"worksheets": sheet_names}, chunks)


def inspect_native_id(path: Path) -> str | None:
    suffix = path.suffix.casefold()
    if suffix == ".eml":
        try:
            message = BytesParser(policy=policy.default).parsebytes(path.read_bytes(), headersonly=True)
            return str(message.get("Message-ID", "")).strip() or None
        except Exception:
            return None
    if suffix == ".msg":
        message = None
        try:
            message = extract_msg.Message(str(path), delayAttachments=True)
            header = getattr(message, "headerDict", {}) or {}
            return str(header.get("Message-Id") or header.get("Message-ID") or "").strip() or None
        except Exception:
            return None
        finally:
            if message is not None:
                try:
                    message.close()
                except Exception:
                    pass
    return None


def extract_source(path: Path, *, max_attachments: int, max_text_bytes: int) -> ExtractedSource:
    suffix = path.suffix.casefold()
    if suffix not in SUPPORTED_SUFFIXES:
        raise UnsupportedSource(f"{suffix or 'This file type'} is not supported")
    if suffix == ".eml":
        result = _extract_eml(path, max_attachments)
    elif suffix == ".msg":
        result = _extract_msg(path, max_attachments)
    elif suffix == ".docx":
        result = _extract_docx(path)
    elif suffix == ".pdf":
        result = _extract_pdf(path)
    elif suffix == ".xlsx":
        result = _extract_xlsx(path)
    else:
        text = decode_text_bytes(path)
        if not text.strip():
            raise UnsupportedSource("The text file is empty")
        result = ExtractedSource(suffix.lstrip("."), None, {}, _chunk_lines(text, "lines"))
    result.chunks = _bounded(result.chunks, max_text_bytes)
    return result
